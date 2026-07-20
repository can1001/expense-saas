"""Excel(xlsx) 생성 공용 유틸 — 워크북 헤더 스타일 + StreamingResponse 헬퍼.

ExcelJS(Next.js) 산출물과 근사 재현(헤더 굵게·파랑 배경·컬럼 폭)이 목표이며,
컬럼 순서·헤더 텍스트·데이터 정확성이 우선이다.
"""

import io
from collections.abc import Iterable

from fastapi.responses import StreamingResponse
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

HEADER_FONT = Font(bold=True, color="FFFFFFFF")
HEADER_FILL = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(*(Side(style="thin"),) * 4)


def style_header_row(worksheet: Worksheet, row: int = 1) -> None:
    """헤더 행에 굵게+파랑 배경+테두리 스타일을 적용한다."""
    for cell in worksheet[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
    worksheet.row_dimensions[row].height = 25


def set_column_widths(worksheet: Worksheet, widths: Iterable[int]) -> None:
    for idx, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(idx)].width = width


def workbook_to_xlsx_response(workbook: Workbook, filename: str) -> StreamingResponse:
    """워크북을 xlsx 바이트로 직렬화해 다운로드용 StreamingResponse 로 반환."""
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type=XLSX_CONTENT_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
