"""관리자 대시보드/연도 설정 현황/보고서/실행·이력/역할·초대 라우터.
(app/api/admin/dashboard, app/api/admin/year-setup-status,
 app/api/admin/budget-execution, app/api/admin/cumulative-report,
 app/api/admin/quarterly-report(+export),
 app/api/admin/hr-admin-execution, app/api/admin/manager-exceptions,
 app/api/admin/change-history,
 app/api/admin/roles(+[id]), app/api/admin/invitations 이전)
mount prefix: /api/admin
"""

import io
import re
import secrets
from datetime import datetime, timedelta, timezone
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import Response
from openpyxl.workbook import Workbook
from pydantic import BaseModel
from sqlalchemy import and_, exists, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import aliased

from expense_api.core.auth.permissions import (
    PERMISSIONS,
    is_protected_system_role,
    sanitize_permissions,
)
from expense_api.core.db.session import get_session
from expense_api.core.dependencies.auth import CurrentUser
from expense_api.core.dependencies.authz import effective_permissions, require_permission
from expense_api.core.dependencies.tenant import require_tenant_id
from expense_api.core.excel import THIN_BORDER, XLSX_CONTENT_TYPE, set_column_widths, style_header_row
from expense_api.core.models.budget import (
    BudgetCategory,
    BudgetDetail,
    BudgetDetailYear,
    BudgetDetailYearHistory,
    BudgetSubcategory,
    Committee,
    Department,
    DepartmentBudgetDetail,
)
from expense_api.core.models.expense import Expense, ExpenseItem
from expense_api.core.models.user import Invitation, Role, User, UserYearRole, UserYearRoleHistory

router = APIRouter()

# 초대 유효기간 (일) — lib/services/invitation.ts INVITATION_EXPIRES_DAYS 이전
INVITATION_EXPIRES_DAYS = 7
_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


def _current_year() -> int:
    return datetime.now(timezone.utc).year


@router.get("/dashboard")
async def get_dashboard(
    year: int | None = None,
    user: CurrentUser = Depends(require_permission(PERMISSIONS.ADMIN_DASHBOARD_READ)),
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
) -> dict:
    yr = year or _current_year()

    now = datetime.now(timezone.utc)
    current_month = now.month
    month_start = datetime(yr, current_month, 1)
    month_end = (
        datetime(yr, current_month + 1, 1) if current_month < 12 else datetime(yr + 1, 1, 1)
    )

    budget_row = (
        await session.execute(
            select(
                func.sum(BudgetDetailYear.budgetAmount), func.sum(BudgetDetailYear.usedAmount)
            ).where(
                BudgetDetailYear.tenantId == tenant_id,
                BudgetDetailYear.year == yr,
                BudgetDetailYear.isActive.is_(True),
            )
        )
    ).one()
    total_budget = budget_row[0] or 0
    total_used = budget_row[1] or 0
    execution_rate = round(total_used / total_budget * 1000) / 10 if total_budget > 0 else 0

    pending_approvals = (
        await session.execute(
            select(func.count(Expense.id)).where(
                Expense.tenantId == tenant_id,
                Expense.status.in_(["PENDING", "APPROVED_STEP_1", "APPROVED_STEP_2"]),
            )
        )
    ).scalar_one()

    monthly_expense = (
        await session.execute(
            select(func.sum(Expense.requestAmount)).where(
                Expense.tenantId == tenant_id,
                Expense.status == "APPROVED_FINAL",
                Expense.requestDate >= month_start,
                Expense.requestDate < month_end,
            )
        )
    ).scalar_one() or 0

    pending_payments = (
        await session.execute(
            select(func.count(Expense.id)).where(
                Expense.tenantId == tenant_id,
                Expense.status == "APPROVED_FINAL",
                Expense.paymentStatus != "COMPLETED",
            )
        )
    ).scalar_one()

    recent_expenses = (
        (
            await session.execute(
                select(Expense)
                .where(Expense.tenantId == tenant_id, Expense.status != "DRAFT")
                .order_by(Expense.createdAt.desc())
                .limit(5)
            )
        )
        .scalars()
        .all()
    )

    yearly_row = (
        await session.execute(
            select(func.sum(Expense.requestAmount), func.count(Expense.id)).where(
                Expense.tenantId == tenant_id,
                Expense.status == "APPROVED_FINAL",
                Expense.requestDate >= datetime(yr, 1, 1),
                Expense.requestDate < datetime(yr + 1, 1, 1),
            )
        )
    ).one()

    return {
        "year": yr,
        "kpi": {
            "executionRate": execution_rate,
            "totalBudget": total_budget,
            "totalUsed": total_used,
            "pendingApprovals": pending_approvals,
            "monthlyExpense": monthly_expense,
            "pendingPayments": pending_payments,
        },
        "yearly": {
            "totalExpense": yearly_row[0] or 0,
            "expenseCount": yearly_row[1] or 0,
        },
        "recentExpenses": [
            {
                "id": e.id,
                "applicantName": e.applicantName,
                "requestAmount": e.requestAmount,
                "status": e.status,
                "requestDate": e.requestDate,
                "department": e.department,
                "committee": e.committee,
                "createdAt": e.createdAt,
            }
            for e in recent_expenses
        ],
    }


@router.get("/year-setup-status")
async def get_year_setup_status(
    year: int | None = None,
    user: CurrentUser = Depends(require_permission(PERMISSIONS.ADMIN_DASHBOARD_READ)),
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
) -> dict:
    yr = year or _current_year()

    total_active_users = (
        await session.execute(
            select(func.count(User.id)).where(User.tenantId == tenant_id, User.isActive.is_(True))
        )
    ).scalar_one()

    users_with_role = (
        await session.execute(
            select(func.count(func.distinct(UserYearRole.userId)))
            .join(User, User.id == UserYearRole.userId)
            .where(
                UserYearRole.tenantId == tenant_id,
                UserYearRole.year == yr,
                User.isActive.is_(True),
            )
        )
    ).scalar_one()

    role_breakdown_rows = (
        await session.execute(
            select(UserYearRole.role, func.count(UserYearRole.id))
            .join(User, User.id == UserYearRole.userId)
            .where(
                UserYearRole.tenantId == tenant_id,
                UserYearRole.year == yr,
                User.isActive.is_(True),
            )
            .group_by(UserYearRole.role)
        )
    ).all()

    total_budget_details = (
        await session.execute(
            select(func.count(DepartmentBudgetDetail.id))
            .join(BudgetDetail, BudgetDetail.id == DepartmentBudgetDetail.budgetDetailId)
            .where(
                DepartmentBudgetDetail.tenantId == tenant_id,
                DepartmentBudgetDetail.isActive.is_(True),
                BudgetDetail.isActive.is_(True),
            )
        )
    ).scalar_one()

    details_with_manager = (
        await session.execute(
            select(func.count(BudgetDetailYear.id))
            .join(BudgetDetail, BudgetDetail.id == BudgetDetailYear.budgetDetailId)
            .where(
                BudgetDetailYear.tenantId == tenant_id,
                BudgetDetailYear.year == yr,
                BudgetDetailYear.managerId.is_not(None),
                BudgetDetail.isActive.is_(True),
            )
        )
    ).scalar_one()

    details_with_budget = (
        await session.execute(
            select(func.count(BudgetDetailYear.id))
            .join(BudgetDetail, BudgetDetail.id == BudgetDetailYear.budgetDetailId)
            .where(
                BudgetDetailYear.tenantId == tenant_id,
                BudgetDetailYear.year == yr,
                BudgetDetailYear.budgetAmount > 0,
                BudgetDetail.isActive.is_(True),
            )
        )
    ).scalar_one()

    budget_sum = (
        await session.execute(
            select(func.sum(BudgetDetailYear.budgetAmount))
            .join(BudgetDetail, BudgetDetail.id == BudgetDetailYear.budgetDetailId)
            .where(
                BudgetDetailYear.tenantId == tenant_id,
                BudgetDetailYear.year == yr,
                BudgetDetail.isActive.is_(True),
            )
        )
    ).scalar_one() or 0

    manager_year_exists = exists(
        select(BudgetDetailYear.id).where(
            BudgetDetailYear.budgetDetailId == BudgetDetail.id,
            BudgetDetailYear.year == yr,
            BudgetDetailYear.managerId.is_not(None),
        )
    )
    missing_managers_rows = (
        await session.execute(
            select(DepartmentBudgetDetail, BudgetDetail)
            .join(BudgetDetail, BudgetDetail.id == DepartmentBudgetDetail.budgetDetailId)
            .where(
                DepartmentBudgetDetail.tenantId == tenant_id,
                DepartmentBudgetDetail.isActive.is_(True),
                BudgetDetail.isActive.is_(True),
                ~manager_year_exists,
            )
            .limit(10)
        )
    ).all()

    budget_year_exists = exists(
        select(BudgetDetailYear.id).where(
            BudgetDetailYear.budgetDetailId == BudgetDetail.id,
            BudgetDetailYear.year == yr,
            BudgetDetailYear.budgetAmount > 0,
        )
    )
    missing_budgets_rows = (
        await session.execute(
            select(DepartmentBudgetDetail, BudgetDetail)
            .join(BudgetDetail, BudgetDetail.id == DepartmentBudgetDetail.budgetDetailId)
            .where(
                DepartmentBudgetDetail.tenantId == tenant_id,
                DepartmentBudgetDetail.isActive.is_(True),
                BudgetDetail.isActive.is_(True),
                ~budget_year_exists,
            )
            .limit(10)
        )
    ).all()

    missing_managers = await _to_missing_detail_dtos(session, missing_managers_rows)
    missing_budgets = await _to_missing_detail_dtos(session, missing_budgets_rows)

    role_year_exists = exists(
        select(UserYearRole.id).where(
            UserYearRole.userId == User.id,
            UserYearRole.year == yr,
        )
    )
    missing_roles_rows = (
        await session.execute(
            select(User)
            .where(User.tenantId == tenant_id, User.isActive.is_(True), ~role_year_exists)
            .limit(10)
        )
    ).scalars().all()

    return {
        "year": yr,
        "summary": {
            "roleSetup": {
                "total": total_active_users,
                "completed": users_with_role,
                "rate": round(users_with_role / total_active_users * 100)
                if total_active_users > 0
                else 0,
                "breakdown": [{"role": role, "count": count} for role, count in role_breakdown_rows],
            },
            "managerAssignment": {
                "total": total_budget_details,
                "completed": details_with_manager,
                "rate": round(details_with_manager / total_budget_details * 100)
                if total_budget_details > 0
                else 0,
            },
            "budgetInput": {
                "total": total_budget_details,
                "completed": details_with_budget,
                "rate": round(details_with_budget / total_budget_details * 100)
                if total_budget_details > 0
                else 0,
                "totalAmount": budget_sum,
            },
        },
        "missing": {
            "roles": [{"id": u.id, "username": u.username, "userid": u.userid} for u in missing_roles_rows],
            "managers": missing_managers,
            "budgets": missing_budgets,
        },
    }


async def _to_missing_detail_dtos(
    session: AsyncSession, rows: list
) -> list[dict]:
    """DepartmentBudgetDetail+BudgetDetail 조인 결과 → committee/department/category/subcategory 라벨 부착."""
    from expense_api.core.models.budget import BudgetCategory, BudgetSubcategory, Committee, Department

    out: list[dict] = []
    for dept_detail, detail in rows:
        dept = await session.get(Department, dept_detail.departmentId)
        committee = await session.get(Committee, dept.committeeId) if dept else None
        subcategory = await session.get(BudgetSubcategory, detail.subcategoryId)
        category = await session.get(BudgetCategory, subcategory.categoryId) if subcategory else None
        out.append(
            {
                "id": detail.id,
                "name": detail.name,
                "committee": committee.name if committee else None,
                "department": dept.name if dept else None,
                "category": category.name if category else None,
                "subcategory": subcategory.name if subcategory else None,
            }
        )
    return out


# ── D2: 보고서 ──────────────────────────────────────────────────────


def _quarter_date_range(year: int, quarter: int) -> tuple[datetime, datetime]:
    start_month = (quarter - 1) * 3 + 1
    end_month = start_month + 2
    start = datetime(year, start_month, 1)
    next_month_first = (
        datetime(year + 1, 1, 1) if end_month == 12 else datetime(year, end_month + 1, 1)
    )
    end = (next_month_first - timedelta(days=1)).replace(hour=23, minute=59, second=59)
    return start, end


def _year_date_range(year: int) -> tuple[datetime, datetime]:
    return datetime(year, 1, 1), datetime(year, 12, 31, 23, 59, 59, 999000)


@router.get("/budget-execution")
async def get_budget_execution(
    year: int | None = None,
    user: CurrentUser = Depends(require_permission(PERMISSIONS.REPORT_BUDGET_EXEC_READ)),
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
) -> dict:
    yr = year or _current_year()
    start, end = _year_date_range(yr)

    all_depts = (
        (
            await session.execute(
                select(Department)
                .where(Department.tenantId == tenant_id, Department.isActive.is_(True))
                .order_by(Department.sortOrder)
            )
        )
        .scalars()
        .all()
    )
    depts_by_committee: dict[str, list[Department]] = {}
    for d in all_depts:
        depts_by_committee.setdefault(d.committeeId, []).append(d)

    committees = (
        (
            await session.execute(
                select(Committee)
                .where(
                    Committee.tenantId == tenant_id,
                    Committee.isActive.is_(True),
                    ~Committee.name.contains("행정위"),
                    ~Committee.name.contains("인사위"),
                )
                .order_by(Committee.sortOrder)
            )
        )
        .scalars()
        .all()
    )
    hr_admin_committees = (
        (
            await session.execute(
                select(Committee)
                .where(
                    Committee.tenantId == tenant_id,
                    Committee.isActive.is_(True),
                    or_(Committee.name.contains("행정위"), Committee.name.contains("인사위")),
                )
                .order_by(Committee.sortOrder)
            )
        )
        .scalars()
        .all()
    )

    budget_rows = (
        await session.execute(
            select(DepartmentBudgetDetail.departmentId, BudgetDetailYear.budgetAmount)
            .select_from(DepartmentBudgetDetail)
            .join(Department, Department.id == DepartmentBudgetDetail.departmentId)
            .join(BudgetDetail, BudgetDetail.id == DepartmentBudgetDetail.budgetDetailId)
            .join(
                BudgetDetailYear,
                and_(
                    BudgetDetailYear.budgetDetailId == BudgetDetail.id,
                    BudgetDetailYear.year == yr,
                    BudgetDetailYear.isActive.is_(True),
                ),
            )
            .where(
                DepartmentBudgetDetail.tenantId == tenant_id,
                DepartmentBudgetDetail.isActive.is_(True),
                Department.isActive.is_(True),
                BudgetDetail.isActive.is_(True),
            )
        )
    ).all()
    department_budget_map: dict[str, int] = {}
    for dept_id, amount in budget_rows:
        department_budget_map[dept_id] = department_budget_map.get(dept_id, 0) + (amount or 0)

    expense_rows = (
        await session.execute(
            select(Expense.department, func.sum(Expense.requestAmount))
            .where(
                Expense.tenantId == tenant_id,
                Expense.status == "APPROVED_FINAL",
                Expense.requestDate >= start,
                Expense.requestDate <= end,
            )
            .group_by(Expense.department)
        )
    ).all()
    department_spent_map = {dept: amount or 0 for dept, amount in expense_rows}

    total_budget = 0
    total_spent = 0
    committee_data = []
    for comm in committees:
        comm_budget = 0
        comm_spent = 0
        departments_payload = []
        for dept in depts_by_committee.get(comm.id, []):
            budget = department_budget_map.get(dept.id, 0)
            spent = department_spent_map.get(dept.name, 0)
            execution_rate = round(spent / budget * 100) if budget > 0 else 0
            comm_budget += budget
            comm_spent += spent
            departments_payload.append(
                {
                    "id": dept.id,
                    "name": dept.name,
                    "budget": budget,
                    "spent": spent,
                    "executionRate": execution_rate,
                }
            )
        total_budget += comm_budget
        total_spent += comm_spent
        comm_execution_rate = round(comm_spent / comm_budget * 100) if comm_budget > 0 else 0
        committee_data.append(
            {
                "id": comm.id,
                "name": comm.name,
                "budget": comm_budget,
                "spent": comm_spent,
                "executionRate": comm_execution_rate,
                "departments": departments_payload,
            }
        )

    total_execution_rate = round(total_spent / total_budget * 100) if total_budget > 0 else 0

    hr_admin_budget = 0
    hr_admin_spent = 0
    for comm in hr_admin_committees:
        for dept in depts_by_committee.get(comm.id, []):
            hr_admin_budget += department_budget_map.get(dept.id, 0)
            hr_admin_spent += department_spent_map.get(dept.name, 0)

    grand_total_budget = total_budget + hr_admin_budget
    grand_total_spent = total_spent + hr_admin_spent
    grand_total_execution_rate = (
        round(grand_total_spent / grand_total_budget * 100) if grand_total_budget > 0 else 0
    )
    ministry_budget_ratio = (
        round(total_budget / grand_total_budget * 100) if grand_total_budget > 0 else 0
    )

    return {
        "year": yr,
        "summary": {
            "totalBudget": total_budget,
            "totalSpent": total_spent,
            "executionRate": total_execution_rate,
            "grandTotalBudget": grand_total_budget,
            "grandTotalSpent": grand_total_spent,
            "grandTotalExecutionRate": grand_total_execution_rate,
            "ministryBudgetRatio": ministry_budget_ratio,
        },
        "committees": committee_data,
    }


@router.get("/cumulative-report")
async def get_cumulative_report(
    year: int | None = None,
    toQuarter: int | None = None,
    user: CurrentUser = Depends(require_permission(PERMISSIONS.REPORT_CUMULATIVE_READ)),
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
) -> dict:
    yr = year or _current_year()
    to_quarter = toQuarter or 4

    total_budget = (
        await session.execute(
            select(func.sum(BudgetDetailYear.budgetAmount)).where(
                BudgetDetailYear.tenantId == tenant_id,
                BudgetDetailYear.year == yr,
                BudgetDetailYear.isActive.is_(True),
            )
        )
    ).scalar_one() or 0

    quarterly_breakdown = []
    cumulative_spent = 0
    for q in range(1, to_quarter + 1):
        q_start, q_end = _quarter_date_range(yr, q)
        spent = (
            await session.execute(
                select(func.sum(Expense.requestAmount)).where(
                    Expense.tenantId == tenant_id,
                    Expense.status == "APPROVED_FINAL",
                    Expense.requestDate >= q_start,
                    Expense.requestDate <= q_end,
                )
            )
        ).scalar_one() or 0
        cumulative_spent += spent
        ratio = round(spent / total_budget * 1000) / 10 if total_budget > 0 else 0
        quarterly_breakdown.append({"quarter": q, "spent": spent, "ratio": ratio})

    year_start, _ = _quarter_date_range(yr, 1)
    _, quarter_end = _quarter_date_range(yr, to_quarter)

    dept_budget_rows = (
        await session.execute(
            select(Department.name, Committee.name, BudgetDetailYear.budgetAmount)
            .select_from(BudgetDetailYear)
            .join(
                DepartmentBudgetDetail,
                DepartmentBudgetDetail.budgetDetailId == BudgetDetailYear.budgetDetailId,
            )
            .join(Department, Department.id == DepartmentBudgetDetail.departmentId)
            .join(Committee, Committee.id == Department.committeeId)
            .where(
                BudgetDetailYear.tenantId == tenant_id,
                BudgetDetailYear.year == yr,
                BudgetDetailYear.isActive.is_(True),
                DepartmentBudgetDetail.tenantId == tenant_id,
            )
        )
    ).all()
    dept_budget_map: dict[str, dict] = {}
    for dept_name, comm_name, amount in dept_budget_rows:
        key = f"{comm_name}|{dept_name}"
        entry = dept_budget_map.get(key)
        if entry:
            entry["budget"] += amount or 0
        else:
            dept_budget_map[key] = {
                "committee": comm_name,
                "department": dept_name,
                "budget": amount or 0,
            }

    expense_dept_rows = (
        await session.execute(
            select(Expense.committee, Expense.department, func.sum(Expense.requestAmount))
            .where(
                Expense.tenantId == tenant_id,
                Expense.status == "APPROVED_FINAL",
                Expense.requestDate >= year_start,
                Expense.requestDate <= quarter_end,
            )
            .group_by(Expense.committee, Expense.department)
        )
    ).all()
    dept_expense_map = {
        f"{committee or ''}|{department or ''}": amount or 0
        for committee, department, amount in expense_dept_rows
    }

    by_department = []
    for key, data in dept_budget_map.items():
        spent = dept_expense_map.get(key, 0)
        remaining = data["budget"] - spent
        execution_rate = round(spent / data["budget"] * 1000) / 10 if data["budget"] > 0 else 0
        by_department.append(
            {
                "committee": data["committee"],
                "department": data["department"],
                "budget": data["budget"],
                "cumulativeSpent": spent,
                "remaining": remaining,
                "executionRate": execution_rate,
            }
        )
    by_department.sort(key=lambda d: (d["committee"], d["department"]))

    remaining = total_budget - cumulative_spent
    execution_rate = round(cumulative_spent / total_budget * 1000) / 10 if total_budget > 0 else 0

    return {
        "year": yr,
        "toQuarter": to_quarter,
        "summary": {
            "totalBudget": total_budget,
            "cumulativeSpent": cumulative_spent,
            "remaining": remaining,
            "executionRate": execution_rate,
        },
        "quarterlyBreakdown": quarterly_breakdown,
        "byDepartment": by_department,
    }


def _quarterly_base_conditions(
    tenant_id: str, start: datetime, end: datetime, department: str, category: str
) -> list:
    conditions = [
        Expense.tenantId == tenant_id,
        Expense.status == "APPROVED_FINAL",
        Expense.requestDate >= start,
        Expense.requestDate <= end,
    ]
    if department:
        conditions.append(Expense.department == department)
    if category:
        conditions.append(
            Expense.id.in_(
                select(ExpenseItem.expenseId).where(ExpenseItem.budgetCategory == category)
            )
        )
    return conditions


def _quarterly_report_defaults() -> tuple[int, int]:
    """이전 분기를 기본값으로 (1분기에는 전년도 4분기)."""
    now = datetime.now(timezone.utc)
    actual_quarter = (now.month - 1) // 3 + 1
    default_quarter = 4 if actual_quarter == 1 else actual_quarter - 1
    default_year = now.year - 1 if actual_quarter == 1 else now.year
    return default_year, default_quarter


async def _load_quarterly_budget_tree(
    session: AsyncSession, tenant_id: str, year: int
) -> tuple[list[BudgetCategory], dict[str, list[BudgetSubcategory]], dict[str, list[tuple]]]:
    budget_categories = (
        (
            await session.execute(
                select(BudgetCategory)
                .where(BudgetCategory.tenantId == tenant_id, BudgetCategory.isActive.is_(True))
                .order_by(BudgetCategory.sortOrder)
            )
        )
        .scalars()
        .all()
    )
    subcategories = (
        (
            await session.execute(
                select(BudgetSubcategory)
                .where(
                    BudgetSubcategory.tenantId == tenant_id, BudgetSubcategory.isActive.is_(True)
                )
                .order_by(BudgetSubcategory.sortOrder)
            )
        )
        .scalars()
        .all()
    )
    subs_by_category: dict[str, list[BudgetSubcategory]] = {}
    for s in subcategories:
        subs_by_category.setdefault(s.categoryId, []).append(s)

    detail_rows = (
        await session.execute(
            select(BudgetDetail, BudgetDetailYear.budgetAmount)
            .join(
                BudgetDetailYear,
                and_(
                    BudgetDetailYear.budgetDetailId == BudgetDetail.id,
                    BudgetDetailYear.year == year,
                    BudgetDetailYear.isActive.is_(True),
                ),
                isouter=True,
            )
            .where(BudgetDetail.tenantId == tenant_id, BudgetDetail.isActive.is_(True))
            .order_by(BudgetDetail.sortOrder)
        )
    ).all()
    details_by_subcategory: dict[str, list[tuple]] = {}
    for detail, budget_amount in detail_rows:
        details_by_subcategory.setdefault(detail.subcategoryId, []).append(
            (detail, budget_amount or 0)
        )

    return budget_categories, subs_by_category, details_by_subcategory


def _quarterly_budget_totals(
    budget_categories: list[BudgetCategory],
    subs_by_category: dict[str, list[BudgetSubcategory]],
    details_by_subcategory: dict[str, list[tuple]],
) -> tuple[dict[str, int], dict[str, int], int]:
    budget_by_category: dict[str, int] = {}
    budget_by_subcategory: dict[str, int] = {}
    total_budget_amount = 0
    for cat in budget_categories:
        cat_total = 0
        for sub in subs_by_category.get(cat.id, []):
            sub_total = sum(amt for _, amt in details_by_subcategory.get(sub.id, []))
            budget_by_subcategory[f"{cat.name}|{sub.name}"] = sub_total
            cat_total += sub_total
        budget_by_category[cat.name] = cat_total
        total_budget_amount += cat_total
    return budget_by_category, budget_by_subcategory, total_budget_amount


@router.get("/quarterly-report")
async def get_quarterly_report(
    year: int | None = None,
    quarter: int | None = None,
    department: str = "",
    category: str = "",
    paymentStatus: str = "",
    user: CurrentUser = Depends(require_permission(PERMISSIONS.REPORT_QUARTERLY_READ)),
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
) -> dict:
    default_year, default_quarter = _quarterly_report_defaults()
    yr = year or default_year
    q = quarter or default_quarter

    start_date, end_date = _quarter_date_range(yr, q)
    year_start_date, year_end_date = _year_date_range(yr)

    base_conditions = _quarterly_base_conditions(tenant_id, start_date, end_date, department, category)
    final_where = list(base_conditions)
    if paymentStatus:
        final_where.append(Expense.paymentStatus == paymentStatus)

    total_count, total_sum = (
        await session.execute(
            select(func.count(Expense.id), func.sum(Expense.requestAmount)).where(*final_where)
        )
    ).one()
    total_sum = total_sum or 0

    completed_sum = (
        await session.execute(
            select(func.sum(Expense.requestAmount)).where(
                *base_conditions, Expense.paymentStatus == "COMPLETED"
            )
        )
    ).scalar_one() or 0
    pending_sum = (
        await session.execute(
            select(func.sum(Expense.requestAmount)).where(
                *base_conditions, Expense.paymentStatus == "PENDING"
            )
        )
    ).scalar_one() or 0

    expense_rows = (
        await session.execute(select(Expense.requestDate, Expense.requestAmount).where(*final_where))
    ).all()
    monthly_map = {m: {"count": 0, "amount": 0} for m in range(1, 4)}
    for request_date, amount in expense_rows:
        month = (request_date.month - 1) - (q - 1) * 3 + 1
        if 1 <= month <= 3:
            monthly_map[month]["count"] += 1
            monthly_map[month]["amount"] += amount

    total_amount = total_sum
    by_month = []
    for month, data in monthly_map.items():
        actual_month = (q - 1) * 3 + month
        ratio = round(data["amount"] / total_amount * 1000) / 10 if total_amount > 0 else 0
        by_month.append(
            {
                "month": actual_month,
                "monthLabel": f"{actual_month}월",
                "count": data["count"],
                "amount": data["amount"],
                "ratio": ratio,
            }
        )

    dept_agg_rows = (
        await session.execute(
            select(
                Expense.committee,
                Expense.department,
                func.count(Expense.id),
                func.sum(Expense.requestAmount),
            )
            .where(*final_where)
            .group_by(Expense.committee, Expense.department)
            .order_by(Expense.committee, Expense.department)
        )
    ).all()

    committees_with_depts = (
        (
            await session.execute(
                select(Committee)
                .where(Committee.tenantId == tenant_id, Committee.isActive.is_(True))
                .order_by(Committee.sortOrder)
            )
        )
        .scalars()
        .all()
    )
    all_depts = (
        (
            await session.execute(
                select(Department)
                .where(Department.tenantId == tenant_id, Department.isActive.is_(True))
                .order_by(Department.sortOrder)
            )
        )
        .scalars()
        .all()
    )
    depts_by_committee: dict[str, list[Department]] = {}
    for d in all_depts:
        depts_by_committee.setdefault(d.committeeId, []).append(d)

    item_rows = (
        await session.execute(
            select(
                ExpenseItem.budgetCategory,
                ExpenseItem.budgetSubcategory,
                ExpenseItem.budgetDetail,
                ExpenseItem.amount,
                Expense.committee,
                Expense.department,
            )
            .join(Expense, Expense.id == ExpenseItem.expenseId)
            .where(*final_where)
        )
    ).all()

    dept_hierarchy: dict[str, dict] = {}
    for cat, sub, detail, amount, committee, dept_name in item_rows:
        dept_key = f"{committee}|{dept_name}"
        cat_map = dept_hierarchy.setdefault(dept_key, {})
        cat_data = cat_map.setdefault(cat, {"count": 0, "amount": 0, "subcategories": {}})
        cat_data["count"] += 1
        cat_data["amount"] += amount
        sub_data = cat_data["subcategories"].setdefault(sub, {"count": 0, "amount": 0, "details": {}})
        sub_data["count"] += 1
        sub_data["amount"] += amount
        detail_data = sub_data["details"].setdefault(detail, {"count": 0, "amount": 0})
        detail_data["count"] += 1
        detail_data["amount"] += amount

    committee_map: dict[str, dict] = {}
    for committee, dept_name, count, amount in dept_agg_rows:
        amount = amount or 0
        entry = committee_map.setdefault(
            committee, {"committee": committee, "count": 0, "amount": 0, "departments": []}
        )
        entry["count"] += count
        entry["amount"] += amount

    for committee, dept_name, count, amount in dept_agg_rows:
        dept_amount = amount or 0
        dept_key = f"{committee}|{dept_name}"
        cat_map = dept_hierarchy.get(dept_key, {})
        comm_entry = committee_map[committee]

        category_details = []
        for cat_name, cat_data in cat_map.items():
            sub_list = []
            for sub_name, sub_data in cat_data["subcategories"].items():
                detail_list = []
                for detail_name, detail_data in sub_data["details"].items():
                    detail_list.append(
                        {
                            "detail": detail_name,
                            "count": detail_data["count"],
                            "amount": detail_data["amount"],
                            "ratio": round(detail_data["amount"] / dept_amount * 1000) / 10
                            if dept_amount > 0
                            else 0,
                        }
                    )
                detail_list.sort(key=lambda d: d["detail"])
                sub_list.append(
                    {
                        "subcategory": sub_name,
                        "count": sub_data["count"],
                        "amount": sub_data["amount"],
                        "ratio": round(sub_data["amount"] / dept_amount * 1000) / 10
                        if dept_amount > 0
                        else 0,
                        "details": detail_list,
                    }
                )
            sub_list.sort(key=lambda s: s["subcategory"])
            category_details.append(
                {
                    "category": cat_name,
                    "count": cat_data["count"],
                    "amount": cat_data["amount"],
                    "ratio": round(cat_data["amount"] / dept_amount * 1000) / 10
                    if dept_amount > 0
                    else 0,
                    "subcategories": sub_list,
                }
            )
        category_details.sort(key=lambda c: c["category"])

        dept_ratio = (
            round(dept_amount / comm_entry["amount"] * 1000) / 10 if comm_entry["amount"] > 0 else 0
        )
        comm_entry["departments"].append(
            {
                "department": dept_name,
                "count": count,
                "amount": dept_amount,
                "deptRatio": dept_ratio,
                "categoryDetails": category_details,
            }
        )

    for comm in committees_with_depts:
        entry = committee_map.setdefault(
            comm.name, {"committee": comm.name, "count": 0, "amount": 0, "departments": []}
        )
        existing_names = {d["department"] for d in entry["departments"]}
        for dept in depts_by_committee.get(comm.id, []):
            if dept.name not in existing_names:
                entry["departments"].append(
                    {
                        "department": dept.name,
                        "count": 0,
                        "amount": 0,
                        "deptRatio": 0,
                        "categoryDetails": [],
                    }
                )

    by_department = []
    for comm_data in committee_map.values():
        comm_data["departments"].sort(key=lambda d: d["department"])
        by_department.append(
            {
                **comm_data,
                "ratio": round(comm_data["amount"] / total_amount * 1000) / 10
                if total_amount > 0
                else 0,
            }
        )
    by_department.sort(key=lambda c: c["committee"])

    budget_categories, subs_by_category, details_by_subcategory = await _load_quarterly_budget_tree(
        session, tenant_id, yr
    )
    budget_by_category, budget_by_subcategory, total_budget_amount = _quarterly_budget_totals(
        budget_categories, subs_by_category, details_by_subcategory
    )

    category_agg_rows = (
        await session.execute(
            select(
                ExpenseItem.budgetCategory,
                ExpenseItem.budgetSubcategory,
                ExpenseItem.budgetDetail,
                func.count(ExpenseItem.id),
                func.sum(ExpenseItem.amount),
            )
            .join(Expense, Expense.id == ExpenseItem.expenseId)
            .where(*final_where)
            .group_by(ExpenseItem.budgetCategory, ExpenseItem.budgetSubcategory, ExpenseItem.budgetDetail)
            .order_by(ExpenseItem.budgetCategory, ExpenseItem.budgetSubcategory, ExpenseItem.budgetDetail)
        )
    ).all()

    yearly_item_rows = (
        await session.execute(
            select(
                ExpenseItem.budgetCategory,
                ExpenseItem.budgetSubcategory,
                ExpenseItem.budgetDetail,
                func.sum(ExpenseItem.amount),
            )
            .join(Expense, Expense.id == ExpenseItem.expenseId)
            .where(
                Expense.tenantId == tenant_id,
                Expense.status == "APPROVED_FINAL",
                Expense.requestDate >= year_start_date,
                Expense.requestDate <= year_end_date,
            )
            .group_by(ExpenseItem.budgetCategory, ExpenseItem.budgetSubcategory, ExpenseItem.budgetDetail)
        )
    ).all()
    yearly_spent_by_category: dict[str, int] = {}
    yearly_spent_by_subcategory: dict[str, int] = {}
    yearly_spent_by_detail: dict[str, int] = {}
    for cat, sub, detail, amount in yearly_item_rows:
        amount = amount or 0
        yearly_spent_by_category[cat] = yearly_spent_by_category.get(cat, 0) + amount
        sub_key = f"{cat}|{sub}"
        yearly_spent_by_subcategory[sub_key] = yearly_spent_by_subcategory.get(sub_key, 0) + amount
        yearly_spent_by_detail[f"{sub_key}|{detail}"] = amount

    category_map: dict[str, dict] = {}
    item_total_amount = 0
    for cat, sub, detail, count, amount in category_agg_rows:
        amount = amount or 0
        item_total_amount += amount
        cat_entry = category_map.setdefault(
            cat,
            {
                "category": cat,
                "count": 0,
                "spentAmount": 0,
                "budgetAmount": budget_by_category.get(cat, 0),
                "subcategories": [],
            },
        )
        cat_entry["count"] += count
        cat_entry["spentAmount"] += amount

        sub_entry = next((s for s in cat_entry["subcategories"] if s["subcategory"] == sub), None)
        if sub_entry is None:
            sub_entry = {
                "subcategory": sub,
                "count": 0,
                "spentAmount": 0,
                "budgetAmount": budget_by_subcategory.get(f"{cat}|{sub}", 0),
                "details": [],
            }
            cat_entry["subcategories"].append(sub_entry)
        sub_entry["count"] += count
        sub_entry["spentAmount"] += amount
        sub_entry["details"].append(
            {"detail": detail, "count": count, "spentAmount": amount, "budgetAmount": 0}
        )

    for cat in budget_categories:
        cat_subs = subs_by_category.get(cat.id, [])
        if cat.name not in category_map:
            category_map[cat.name] = {
                "category": cat.name,
                "count": 0,
                "spentAmount": 0,
                "budgetAmount": budget_by_category.get(cat.name, 0),
                "subcategories": [
                    {
                        "subcategory": sub.name,
                        "count": 0,
                        "spentAmount": 0,
                        "budgetAmount": budget_by_subcategory.get(f"{cat.name}|{sub.name}", 0),
                        "details": [
                            {"detail": detail.name, "count": 0, "spentAmount": 0, "budgetAmount": 0}
                            for detail, _ in details_by_subcategory.get(sub.id, [])
                        ],
                    }
                    for sub in cat_subs
                ],
            }
        else:
            existing_cat = category_map[cat.name]
            for sub in cat_subs:
                existing_sub = next(
                    (s for s in existing_cat["subcategories"] if s["subcategory"] == sub.name), None
                )
                if existing_sub is None:
                    existing_sub = {
                        "subcategory": sub.name,
                        "count": 0,
                        "spentAmount": 0,
                        "budgetAmount": budget_by_subcategory.get(f"{cat.name}|{sub.name}", 0),
                        "details": [],
                    }
                    existing_cat["subcategories"].append(existing_sub)
                for detail, _ in details_by_subcategory.get(sub.id, []):
                    exists_detail = any(d["detail"] == detail.name for d in existing_sub["details"])
                    if not exists_detail:
                        existing_sub["details"].append(
                            {"detail": detail.name, "count": 0, "spentAmount": 0, "budgetAmount": 0}
                        )

    by_category = []
    for cat_data in sorted(category_map.values(), key=lambda c: c["category"]):
        yearly_spent_amount = yearly_spent_by_category.get(cat_data["category"], 0)
        yearly_remaining_amount = cat_data["budgetAmount"] - yearly_spent_amount
        yearly_execution_rate = (
            round(yearly_spent_amount / cat_data["budgetAmount"] * 1000) / 10
            if cat_data["budgetAmount"] > 0
            else 0
        )
        quarterly_budget = round(cat_data["budgetAmount"] / 4)
        quarterly_remaining = quarterly_budget - cat_data["spentAmount"]
        quarterly_execution_rate = (
            round(cat_data["spentAmount"] / quarterly_budget * 1000) / 10 if quarterly_budget > 0 else 0
        )

        subcategories_payload = []
        for sub in cat_data["subcategories"]:
            sub_key = f"{cat_data['category']}|{sub['subcategory']}"
            sub_yearly_spent = yearly_spent_by_subcategory.get(sub_key, 0)
            sub_yearly_remaining = sub["budgetAmount"] - sub_yearly_spent
            sub_yearly_rate = (
                round(sub_yearly_spent / sub["budgetAmount"] * 1000) / 10
                if sub["budgetAmount"] > 0
                else 0
            )
            sub_quarterly_budget = round(sub["budgetAmount"] / 4)
            sub_quarterly_remaining = sub_quarterly_budget - sub["spentAmount"]
            sub_quarterly_rate = (
                round(sub["spentAmount"] / sub_quarterly_budget * 1000) / 10
                if sub_quarterly_budget > 0
                else 0
            )
            details_payload = []
            for detail in sorted(sub["details"], key=lambda d: d["detail"]):
                detail_key = f"{sub_key}|{detail['detail']}"
                detail_yearly_spent = yearly_spent_by_detail.get(detail_key, 0)
                details_payload.append(
                    {
                        **detail,
                        "yearlySpentAmount": detail_yearly_spent,
                        "ratio": round(detail["spentAmount"] / item_total_amount * 1000) / 10
                        if item_total_amount > 0
                        else 0,
                    }
                )
            subcategories_payload.append(
                {
                    **sub,
                    "yearlySpentAmount": sub_yearly_spent,
                    "yearlyRemainingAmount": sub_yearly_remaining,
                    "yearlyExecutionRate": sub_yearly_rate,
                    "quarterlyBudget": sub_quarterly_budget,
                    "quarterlyRemaining": sub_quarterly_remaining,
                    "quarterlyExecutionRate": sub_quarterly_rate,
                    "ratio": round(sub["spentAmount"] / item_total_amount * 1000) / 10
                    if item_total_amount > 0
                    else 0,
                    "details": details_payload,
                }
            )
        by_category.append(
            {
                **cat_data,
                "yearlySpentAmount": yearly_spent_amount,
                "yearlyRemainingAmount": yearly_remaining_amount,
                "yearlyExecutionRate": yearly_execution_rate,
                "quarterlyBudget": quarterly_budget,
                "quarterlyRemaining": quarterly_remaining,
                "quarterlyExecutionRate": quarterly_execution_rate,
                "ratio": round(cat_data["spentAmount"] / item_total_amount * 1000) / 10
                if item_total_amount > 0
                else 0,
                "subcategories": subcategories_payload,
            }
        )

    yearly_spent = (
        await session.execute(
            select(func.sum(Expense.requestAmount)).where(
                Expense.tenantId == tenant_id,
                Expense.status == "APPROVED_FINAL",
                Expense.requestDate >= year_start_date,
                Expense.requestDate <= year_end_date,
            )
        )
    ).scalar_one() or 0

    filter_dept_rows = (
        await session.execute(
            select(Expense.committee, Expense.department)
            .where(
                Expense.tenantId == tenant_id,
                Expense.status == "APPROVED_FINAL",
                Expense.requestDate >= start_date,
                Expense.requestDate <= end_date,
            )
            .distinct()
            .order_by(Expense.committee, Expense.department)
        )
    ).all()
    filter_category_rows = (
        (
            await session.execute(
                select(ExpenseItem.budgetCategory)
                .join(Expense, Expense.id == ExpenseItem.expenseId)
                .where(
                    Expense.tenantId == tenant_id,
                    Expense.status == "APPROVED_FINAL",
                    Expense.requestDate >= start_date,
                    Expense.requestDate <= end_date,
                )
                .distinct()
                .order_by(ExpenseItem.budgetCategory)
            )
        )
        .scalars()
        .all()
    )

    quarterly_spent = total_sum
    yearly_remaining = total_budget_amount - yearly_spent
    yearly_execution_rate_total = (
        round(yearly_spent / total_budget_amount * 1000) / 10 if total_budget_amount > 0 else 0
    )
    quarterly_budget_total = round(total_budget_amount / 4)
    quarterly_remaining_total = quarterly_budget_total - quarterly_spent
    quarterly_execution_rate_total = (
        round(quarterly_spent / quarterly_budget_total * 1000) / 10 if quarterly_budget_total > 0 else 0
    )

    return {
        "year": yr,
        "quarter": q,
        "period": {
            "startDate": start_date.date().isoformat(),
            "endDate": end_date.date().isoformat(),
        },
        "summary": {
            "totalExpenses": total_count,
            "totalAmount": total_sum,
            "completedAmount": completed_sum,
            "pendingAmount": pending_sum,
        },
        "budgetSummary": {
            "totalBudget": total_budget_amount,
            "yearlySpent": yearly_spent,
            "yearlyRemaining": yearly_remaining,
            "yearlyExecutionRate": yearly_execution_rate_total,
            "quarterlyBudget": quarterly_budget_total,
            "quarterlySpent": quarterly_spent,
            "quarterlyRemaining": quarterly_remaining_total,
            "quarterlyExecutionRate": quarterly_execution_rate_total,
        },
        "byMonth": by_month,
        "byDepartment": by_department,
        "byCategory": by_category,
        "filterOptions": {
            "departments": [{"committee": c, "department": d} for c, d in filter_dept_rows],
            "categories": list(filter_category_rows),
        },
    }


@router.get("/quarterly-report/export")
async def export_quarterly_report(
    year: int | None = None,
    quarter: int | None = None,
    department: str = "",
    category: str = "",
    paymentStatus: str = "",
    user: CurrentUser = Depends(require_permission(PERMISSIONS.REPORT_QUARTERLY_READ)),
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
) -> Response:
    perms = await effective_permissions(user, session)
    if PERMISSIONS.REPORT_EXPORT not in perms:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "권한이 없습니다.")

    default_year, default_quarter = _quarterly_report_defaults()
    yr = year or default_year
    q = quarter or default_quarter

    start_date, end_date = _quarter_date_range(yr, q)
    year_start_date, year_end_date = _year_date_range(yr)

    base_conditions = _quarterly_base_conditions(tenant_id, start_date, end_date, department, category)
    final_where = list(base_conditions)
    if paymentStatus:
        final_where.append(Expense.paymentStatus == paymentStatus)

    total_count, total_sum = (
        await session.execute(
            select(func.count(Expense.id), func.sum(Expense.requestAmount)).where(*final_where)
        )
    ).one()
    total_sum = total_sum or 0

    expense_rows = (
        await session.execute(select(Expense.requestDate, Expense.requestAmount).where(*final_where))
    ).all()
    monthly_map = {m: {"count": 0, "amount": 0} for m in range(1, 4)}
    for request_date, amount in expense_rows:
        month = (request_date.month - 1) - (q - 1) * 3 + 1
        if 1 <= month <= 3:
            monthly_map[month]["count"] += 1
            monthly_map[month]["amount"] += amount

    dept_agg_rows = (
        await session.execute(
            select(
                Expense.committee,
                Expense.department,
                func.count(Expense.id),
                func.sum(Expense.requestAmount),
            )
            .where(*final_where)
            .group_by(Expense.committee, Expense.department)
            .order_by(Expense.committee, Expense.department)
        )
    ).all()

    category_agg_rows = (
        await session.execute(
            select(
                ExpenseItem.budgetCategory,
                ExpenseItem.budgetSubcategory,
                ExpenseItem.budgetDetail,
                func.count(ExpenseItem.id),
                func.sum(ExpenseItem.amount),
            )
            .join(Expense, Expense.id == ExpenseItem.expenseId)
            .where(*final_where)
            .group_by(ExpenseItem.budgetCategory, ExpenseItem.budgetSubcategory, ExpenseItem.budgetDetail)
            .order_by(ExpenseItem.budgetCategory, ExpenseItem.budgetSubcategory, ExpenseItem.budgetDetail)
        )
    ).all()

    budget_categories, subs_by_category, details_by_subcategory = await _load_quarterly_budget_tree(
        session, tenant_id, yr
    )
    budget_by_category, budget_by_subcategory, total_budget_amount = _quarterly_budget_totals(
        budget_categories, subs_by_category, details_by_subcategory
    )

    yearly_spent = (
        await session.execute(
            select(func.sum(Expense.requestAmount)).where(
                Expense.tenantId == tenant_id,
                Expense.status == "APPROVED_FINAL",
                Expense.requestDate >= year_start_date,
                Expense.requestDate <= year_end_date,
            )
        )
    ).scalar_one() or 0

    yearly_item_rows = (
        await session.execute(
            select(
                ExpenseItem.budgetCategory,
                ExpenseItem.budgetSubcategory,
                ExpenseItem.budgetDetail,
                func.sum(ExpenseItem.amount),
            )
            .join(Expense, Expense.id == ExpenseItem.expenseId)
            .where(
                Expense.tenantId == tenant_id,
                Expense.status == "APPROVED_FINAL",
                Expense.requestDate >= year_start_date,
                Expense.requestDate <= year_end_date,
            )
            .group_by(ExpenseItem.budgetCategory, ExpenseItem.budgetSubcategory, ExpenseItem.budgetDetail)
        )
    ).all()
    yearly_spent_by_subcategory: dict[str, int] = {}
    yearly_spent_by_detail: dict[str, int] = {}
    for cat, sub, detail, amount in yearly_item_rows:
        amount = amount or 0
        sub_key = f"{cat}|{sub}"
        yearly_spent_by_subcategory[sub_key] = yearly_spent_by_subcategory.get(sub_key, 0) + amount
        yearly_spent_by_detail[f"{sub_key}|{detail}"] = amount

    total_amount = total_sum
    quarterly_spent = total_sum
    yearly_remaining = total_budget_amount - yearly_spent
    yearly_execution_rate = (
        round(yearly_spent / total_budget_amount * 1000) / 10 if total_budget_amount > 0 else 0
    )
    quarterly_budget_total = round(total_budget_amount / 4)
    quarterly_remaining_total = quarterly_budget_total - quarterly_spent
    quarterly_execution_rate_total = (
        round(quarterly_spent / quarterly_budget_total * 1000) / 10 if quarterly_budget_total > 0 else 0
    )

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "요약"
    summary_sheet.append(["항목", "값"])
    style_header_row(summary_sheet)
    set_column_widths(summary_sheet, [20, 20])
    summary_rows = [
        ("조회 연도", f"{yr}년"),
        ("조회 분기", f"{q}분기"),
        ("조회 기간", f"{start_date.date().isoformat()} ~ {end_date.date().isoformat()}"),
        ("총 건수", total_count),
        ("", ""),
        ("[ 연간 예산 현황 ]", ""),
        ("연간 예산액", total_budget_amount),
        ("연간 지출액", yearly_spent),
        ("연간 잔액", yearly_remaining),
        ("연간 집행률", f"{yearly_execution_rate}%"),
        ("", ""),
        (f"[ {q}분기 예산 현황 ]", ""),
        ("분기 예산액", quarterly_budget_total),
        ("분기 지출액", quarterly_spent),
        ("분기 잔액", quarterly_remaining_total),
        ("분기 집행률", f"{quarterly_execution_rate_total}%"),
    ]
    for item, value in summary_rows:
        summary_sheet.append([item, value])
    for excel_row in summary_sheet.iter_rows(min_row=2, max_row=summary_sheet.max_row):
        for cell in excel_row:
            cell.border = THIN_BORDER
        value_cell = excel_row[1]
        if isinstance(value_cell.value, (int, float)):
            value_cell.number_format = "#,##0"

    monthly_sheet = workbook.create_sheet("월별 지출")
    monthly_sheet.append(["월", "건수", "금액", "비율(%)"])
    style_header_row(monthly_sheet)
    set_column_widths(monthly_sheet, [10, 10, 20, 12])
    for month, data in monthly_map.items():
        actual_month = (q - 1) * 3 + month
        ratio = round(data["amount"] / total_amount * 1000) / 10 if total_amount > 0 else 0
        monthly_sheet.append([f"{actual_month}월", data["count"], data["amount"], ratio])
    for excel_row in monthly_sheet.iter_rows(min_row=2, max_row=monthly_sheet.max_row):
        for cell in excel_row:
            cell.border = THIN_BORDER
        excel_row[2].number_format = "#,##0"

    dept_sheet = workbook.create_sheet("부서별 지출")
    dept_sheet.append(["위원회", "사역팀(부)", "건수", "금액", "비율(%)"])
    style_header_row(dept_sheet)
    set_column_widths(dept_sheet, [15, 15, 10, 20, 12])
    for committee, department_name, count, amount in dept_agg_rows:
        amount = amount or 0
        ratio = round(amount / total_amount * 1000) / 10 if total_amount > 0 else 0
        dept_sheet.append([committee, department_name, count, amount, ratio])
    for excel_row in dept_sheet.iter_rows(min_row=2, max_row=dept_sheet.max_row):
        for cell in excel_row:
            cell.border = THIN_BORDER
        excel_row[3].number_format = "#,##0"

    cat_sheet = workbook.create_sheet("분기별예산대비지출")
    cat_sheet.append(
        [
            "예산(항)",
            "예산(목)",
            "분기예산",
            "분기지출",
            "분기잔액",
            "분기집행률(%)",
            "건수",
            "(연간예산)",
            "(연간지출)",
            "(연간잔액)",
            "(연간집행률%)",
        ]
    )
    style_header_row(cat_sheet)
    set_column_widths(cat_sheet, [20, 20, 15, 15, 15, 14, 10, 15, 15, 15, 14])

    spent_by_subcategory: dict[str, dict] = {}
    spent_by_detail: dict[str, dict] = {}
    for cat, sub, detail, count, amount in category_agg_rows:
        amount = amount or 0
        sub_key = f"{cat}|{sub}"
        detail_key = f"{sub_key}|{detail}"
        existing = spent_by_subcategory.get(sub_key)
        if existing:
            existing["count"] += count
            existing["amount"] += amount
        else:
            spent_by_subcategory[sub_key] = {"count": count, "amount": amount}
        spent_by_detail[detail_key] = {
            "count": count,
            "amount": amount,
            "category": cat,
            "subcategory": sub,
            "detail": detail,
        }

    for cat in budget_categories:
        for sub in subs_by_category.get(cat.id, []):
            key = f"{cat.name}|{sub.name}"
            budget = budget_by_subcategory.get(key, 0)
            spent_entry = spent_by_subcategory.get(key)
            spent_amount = spent_entry["amount"] if spent_entry else 0
            count = spent_entry["count"] if spent_entry else 0
            q_budget = round(budget / 4)
            q_remaining = q_budget - spent_amount
            q_exec_rate = round(spent_amount / q_budget * 1000) / 10 if q_budget > 0 else 0
            yearly_sub_spent = yearly_spent_by_subcategory.get(key, 0)
            y_remaining = budget - yearly_sub_spent
            y_exec_rate = round(yearly_sub_spent / budget * 1000) / 10 if budget > 0 else 0
            if budget > 0 or spent_amount > 0:
                cat_sheet.append(
                    [
                        cat.name,
                        sub.name,
                        q_budget,
                        spent_amount,
                        q_remaining,
                        q_exec_rate,
                        count,
                        budget,
                        yearly_sub_spent,
                        y_remaining,
                        y_exec_rate,
                    ]
                )
    for excel_row in cat_sheet.iter_rows(min_row=2, max_row=cat_sheet.max_row):
        for cell in excel_row:
            cell.border = THIN_BORDER
        for idx in (2, 3, 4, 7, 8, 9):
            excel_row[idx].number_format = "#,##0"

    detail_sheet = workbook.create_sheet("세목별상세")
    detail_sheet.append(["예산(항)", "예산(목)", "예산(세목)", "분기지출", "건수", "연간지출", "비율(%)"])
    style_header_row(detail_sheet)
    set_column_widths(detail_sheet, [20, 20, 25, 15, 10, 15, 12])

    detail_entries = []
    for key, data in spent_by_detail.items():
        yearly_detail_spent = yearly_spent_by_detail.get(key, 0)
        ratio = round(data["amount"] / total_amount * 1000) / 10 if total_amount > 0 else 0
        detail_entries.append(
            {
                "category": data["category"],
                "subcategory": data["subcategory"],
                "detail": data["detail"],
                "spentAmount": data["amount"],
                "count": data["count"],
                "yearlySpent": yearly_detail_spent,
                "ratio": ratio,
            }
        )
    detail_entries.sort(key=lambda d: (d["category"], d["subcategory"], d["detail"]))
    for item in detail_entries:
        detail_sheet.append(
            [
                item["category"],
                item["subcategory"],
                item["detail"],
                item["spentAmount"],
                item["count"],
                item["yearlySpent"],
                item["ratio"],
            ]
        )
    for excel_row in detail_sheet.iter_rows(min_row=2, max_row=detail_sheet.max_row):
        for cell in excel_row:
            cell.border = THIN_BORDER
        excel_row[3].number_format = "#,##0"
        excel_row[5].number_format = "#,##0"

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    filename = f"분기별회계보고_{yr}년_{q}분기.xlsx"
    return Response(
        content=buffer.getvalue(),
        media_type=XLSX_CONTENT_TYPE,
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


# ── D3: 실행·이력 ────────────────────────────────────────────────────


def _format_display_name(name: str) -> str:
    """언더스코어를 공백으로 변환하고 정리."""
    return re.sub(r"\s+", " ", name.replace("_", " ")).strip()


async def _load_committee_budget_rows(
    session: AsyncSession, tenant_id: str, committee_id: str, year: int
) -> list[tuple[str, str | None, str | None, int]]:
    """committee 산하 활성 부서의 세목 목록 → (세목명, 목명, 항명, 해당연도 예산액)."""
    rows = (
        await session.execute(
            select(
                BudgetDetail.name,
                BudgetSubcategory.name,
                BudgetCategory.name,
                BudgetDetailYear.budgetAmount,
            )
            .select_from(DepartmentBudgetDetail)
            .join(Department, Department.id == DepartmentBudgetDetail.departmentId)
            .join(BudgetDetail, BudgetDetail.id == DepartmentBudgetDetail.budgetDetailId)
            .outerjoin(BudgetSubcategory, BudgetSubcategory.id == BudgetDetail.subcategoryId)
            .outerjoin(BudgetCategory, BudgetCategory.id == BudgetSubcategory.categoryId)
            .outerjoin(
                BudgetDetailYear,
                and_(
                    BudgetDetailYear.budgetDetailId == BudgetDetail.id,
                    BudgetDetailYear.year == year,
                    BudgetDetailYear.isActive.is_(True),
                ),
            )
            .where(
                DepartmentBudgetDetail.tenantId == tenant_id,
                DepartmentBudgetDetail.isActive.is_(True),
                Department.committeeId == committee_id,
                Department.isActive.is_(True),
            )
        )
    ).all()
    return [(name, sub, cat, amount or 0) for name, sub, cat, amount in rows]


@router.get("/hr-admin-execution")
async def get_hr_admin_execution(
    year: int | None = None,
    user: CurrentUser = Depends(require_permission(PERMISSIONS.REPORT_HR_ADMIN_READ)),
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
) -> dict:
    yr = year or _current_year()
    start, end = _year_date_range(yr)

    personnel_committee = (
        (
            await session.execute(
                select(Committee).where(
                    Committee.tenantId == tenant_id,
                    Committee.isActive.is_(True),
                    Committee.name.contains("인사위"),
                )
            )
        )
        .scalars()
        .first()
    )
    admin_committee = (
        (
            await session.execute(
                select(Committee).where(
                    Committee.tenantId == tenant_id,
                    Committee.isActive.is_(True),
                    Committee.name.contains("행정위"),
                )
            )
        )
        .scalars()
        .first()
    )

    expense_item_rows = (
        await session.execute(
            select(
                ExpenseItem.budgetCategory,
                ExpenseItem.budgetSubcategory,
                ExpenseItem.budgetDetail,
                func.sum(ExpenseItem.amount),
            )
            .join(Expense, Expense.id == ExpenseItem.expenseId)
            .where(
                Expense.tenantId == tenant_id,
                Expense.status == "APPROVED_FINAL",
                Expense.requestDate >= start,
                Expense.requestDate <= end,
            )
            .group_by(ExpenseItem.budgetCategory, ExpenseItem.budgetSubcategory, ExpenseItem.budgetDetail)
        )
    ).all()
    spent_by_detail = {
        f"{cat}|{sub}|{detail}": amount or 0 for cat, sub, detail, amount in expense_item_rows
    }

    personnel_items: list[dict] = []
    personnel_total_budget = 0
    personnel_total_spent = 0
    if personnel_committee:
        subcategory_map: dict[str, dict] = {}
        for detail_name, sub_name, cat_name, budget in await _load_committee_budget_rows(
            session, tenant_id, personnel_committee.id, yr
        ):
            subcat_name = sub_name or detail_name
            cat_label = cat_name or ""
            key = f"{cat_label}|{subcat_name}"
            spent_key = f"{cat_label}|{subcat_name}|{detail_name}"
            entry = subcategory_map.setdefault(key, {"budget": 0, "spent": 0})
            entry["budget"] += budget
            entry["spent"] += spent_by_detail.get(spent_key, 0)

        for key, data in subcategory_map.items():
            _, subcat_name = key.split("|", 1)
            execution_rate = round(data["spent"] / data["budget"] * 100) if data["budget"] > 0 else 0
            personnel_items.append(
                {
                    "name": _format_display_name(subcat_name),
                    "budget": data["budget"],
                    "spent": data["spent"],
                    "executionRate": execution_rate,
                }
            )
            personnel_total_budget += data["budget"]
            personnel_total_spent += data["spent"]

    admin_groups: list[dict] = []
    admin_total_budget = 0
    admin_total_spent = 0
    if admin_committee:
        category_map: dict[str, dict[str, dict]] = {}
        for detail_name, sub_name, cat_name, budget in await _load_committee_budget_rows(
            session, tenant_id, admin_committee.id, yr
        ):
            cat_label = cat_name or "기타"
            subcat_name = sub_name or detail_name
            spent_key = f"{cat_label}|{subcat_name}|{detail_name}"
            subcat_map = category_map.setdefault(cat_label, {})
            entry = subcat_map.setdefault(subcat_name, {"budget": 0, "spent": 0})
            entry["budget"] += budget
            entry["spent"] += spent_by_detail.get(spent_key, 0)

        for cat_label, subcat_map in category_map.items():
            items = []
            group_budget = 0
            group_spent = 0
            for subcat_name, data in subcat_map.items():
                execution_rate = round(data["spent"] / data["budget"] * 100) if data["budget"] > 0 else 0
                items.append(
                    {
                        "name": _format_display_name(subcat_name),
                        "budget": data["budget"],
                        "spent": data["spent"],
                        "executionRate": execution_rate,
                    }
                )
                group_budget += data["budget"]
                group_spent += data["spent"]
            group_execution_rate = round(group_spent / group_budget * 100) if group_budget > 0 else 0
            admin_groups.append(
                {
                    "name": _format_display_name(cat_label),
                    "items": items,
                    "subtotal": {
                        "budget": group_budget,
                        "spent": group_spent,
                        "executionRate": group_execution_rate,
                    },
                }
            )
            admin_total_budget += group_budget
            admin_total_spent += group_spent

    personnel_execution_rate = (
        round(personnel_total_spent / personnel_total_budget * 100) if personnel_total_budget > 0 else 0
    )
    admin_execution_rate = (
        round(admin_total_spent / admin_total_budget * 100) if admin_total_budget > 0 else 0
    )

    combined_budget = personnel_total_budget + admin_total_budget
    combined_spent = personnel_total_spent + admin_total_spent
    combined_execution_rate = (
        round(combined_spent / combined_budget * 100) if combined_budget > 0 else 0
    )

    total_budget = (
        await session.execute(
            select(func.sum(BudgetDetailYear.budgetAmount)).where(
                BudgetDetailYear.tenantId == tenant_id,
                BudgetDetailYear.year == yr,
                BudgetDetailYear.isActive.is_(True),
            )
        )
    ).scalar_one() or 0

    hr_admin_ratio = round(combined_budget / total_budget * 100) if total_budget > 0 else 0

    return {
        "year": yr,
        "personnel": {
            "items": personnel_items,
            "total": {
                "budget": personnel_total_budget,
                "spent": personnel_total_spent,
                "executionRate": personnel_execution_rate,
            },
        },
        "admin": {
            "groups": admin_groups,
            "total": {
                "budget": admin_total_budget,
                "spent": admin_total_spent,
                "executionRate": admin_execution_rate,
            },
        },
        "combined": {
            "budget": combined_budget,
            "spent": combined_spent,
            "executionRate": combined_execution_rate,
        },
        "overall": {
            "totalBudget": total_budget,
            "hrAdminBudget": combined_budget,
            "hrAdminRatio": hr_admin_ratio,
        },
    }


@router.get("/manager-exceptions")
async def get_manager_exceptions(
    year: int | None = None,
    user: CurrentUser = Depends(require_permission(PERMISSIONS.BUDGET_MANAGER_MANAGE)),
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
) -> dict:
    yr = year or _current_year()

    leader_rows = (
        await session.execute(
            select(UserYearRole.departmentId, UserYearRole.userId, User.username)
            .join(User, User.id == UserYearRole.userId)
            .where(
                UserYearRole.tenantId == tenant_id,
                UserYearRole.year == yr,
                UserYearRole.role == "team_leader",
                UserYearRole.departmentId.is_not(None),
            )
        )
    ).all()
    dept_leader_map = {
        dept_id: {"id": user_id, "name": username} for dept_id, user_id, username in leader_rows
    }

    manager_user = aliased(User)
    detail_rows = (
        await session.execute(
            select(
                BudgetDetailYear.id,
                BudgetDetailYear.budgetDetailId,
                BudgetDetailYear.managerId,
                manager_user.username,
                BudgetDetail.name,
                BudgetSubcategory.name,
                BudgetCategory.name,
                DepartmentBudgetDetail.departmentId,
                Department.name,
                Committee.name,
            )
            .select_from(BudgetDetailYear)
            .join(manager_user, manager_user.id == BudgetDetailYear.managerId)
            .join(BudgetDetail, BudgetDetail.id == BudgetDetailYear.budgetDetailId)
            .join(BudgetSubcategory, BudgetSubcategory.id == BudgetDetail.subcategoryId)
            .join(BudgetCategory, BudgetCategory.id == BudgetSubcategory.categoryId)
            .join(DepartmentBudgetDetail, DepartmentBudgetDetail.budgetDetailId == BudgetDetail.id)
            .join(Department, Department.id == DepartmentBudgetDetail.departmentId)
            .join(Committee, Committee.id == Department.committeeId)
            .where(
                BudgetDetailYear.tenantId == tenant_id,
                BudgetDetailYear.year == yr,
                BudgetDetailYear.isActive.is_(True),
                BudgetDetailYear.managerId.is_not(None),
            )
        )
    ).all()

    exceptions: list[dict] = []
    total_details = 0
    for (
        budget_detail_year_id,
        budget_detail_id,
        manager_id,
        manager_name,
        detail_name,
        subcat_name,
        cat_name,
        department_id,
        department_name,
        committee_name,
    ) in detail_rows:
        total_details += 1
        team_leader = dept_leader_map.get(department_id)
        if not team_leader or team_leader["id"] != manager_id:
            exceptions.append(
                {
                    "budgetDetailId": budget_detail_id,
                    "budgetDetailYearId": budget_detail_year_id,
                    "committee": committee_name,
                    "department": department_name,
                    "category": cat_name,
                    "subcategory": subcat_name,
                    "detail": detail_name,
                    "teamLeader": team_leader,
                    "manager": {"id": manager_id, "name": manager_name},
                }
            )

    exceptions.sort(key=lambda e: (e["committee"], e["department"], e["detail"]))

    exception_count = len(exceptions)
    exception_rate = (
        round(exception_count / total_details * 1000) / 10 if total_details > 0 else 0
    )

    return {
        "year": yr,
        "summary": {
            "totalDetails": total_details,
            "exceptionCount": exception_count,
            "exceptionRate": exception_rate,
        },
        "exceptions": exceptions,
    }


@router.get("/change-history")
async def get_change_history(
    history_type: str = Query(default="all", alias="type"),
    year: int | None = None,
    limit: int = 50,
    offset: int = 0,
    user: CurrentUser = Depends(require_permission(PERMISSIONS.ADMIN_DASHBOARD_READ)),
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
) -> dict:
    result: dict = {"total": {}}

    if history_type in ("all", "roles"):
        role_conditions = [UserYearRoleHistory.tenantId == tenant_id]
        if year is not None:
            role_conditions.append(UserYearRoleHistory.year == year)

        role_history_rows = (
            (
                await session.execute(
                    select(UserYearRoleHistory)
                    .where(*role_conditions)
                    .order_by(UserYearRoleHistory.changedAt.desc())
                    .limit(limit)
                    .offset(offset)
                )
            )
            .scalars()
            .all()
        )
        role_count = (
            await session.execute(select(func.count(UserYearRoleHistory.id)).where(*role_conditions))
        ).scalar_one()

        user_ids = list({h.userId for h in role_history_rows})
        user_rows = (
            (
                await session.execute(
                    select(User.id, User.username, User.userid).where(User.id.in_(user_ids))
                )
            ).all()
            if user_ids
            else []
        )
        user_map = {uid: (username, userid) for uid, username, userid in user_rows}

        role_history = []
        for h in role_history_rows:
            username, userid = user_map.get(h.userId, (None, None))
            entry = {**h.model_dump(), "userName": username if username is not None else h.userId}
            if userid is not None:
                entry["userLoginId"] = userid
            role_history.append(entry)

        result["roleHistory"] = role_history
        result["total"]["roles"] = role_count

    if history_type in ("all", "budgets"):
        budget_conditions = [BudgetDetailYearHistory.tenantId == tenant_id]
        if year is not None:
            budget_conditions.append(BudgetDetailYearHistory.year == year)

        budget_history_rows = (
            (
                await session.execute(
                    select(BudgetDetailYearHistory)
                    .where(*budget_conditions)
                    .order_by(BudgetDetailYearHistory.changedAt.desc())
                    .limit(limit)
                    .offset(offset)
                )
            )
            .scalars()
            .all()
        )
        budget_count = (
            await session.execute(
                select(func.count(BudgetDetailYearHistory.id)).where(*budget_conditions)
            )
        ).scalar_one()

        result["budgetHistory"] = [h.model_dump() for h in budget_history_rows]
        result["total"]["budgets"] = budget_count

    return result


# ── D4: 역할 관리 (app/api/admin/roles, [id]) ──────────────────────────


async def _role_counts(session: AsyncSession, role_ids: list[str]) -> tuple[dict, dict]:
    """roleId 기준 User·UserYearRole 카운트 맵 (prisma _count.{users,userYearRoles} 대응)."""
    if not role_ids:
        return {}, {}
    user_rows = await session.execute(
        select(User.roleId, func.count(User.id))
        .where(User.roleId.in_(role_ids))
        .group_by(User.roleId)
    )
    year_role_rows = await session.execute(
        select(UserYearRole.roleId, func.count(UserYearRole.id))
        .where(UserYearRole.roleId.in_(role_ids))
        .group_by(UserYearRole.roleId)
    )
    return dict(user_rows.all()), dict(year_role_rows.all())


def _role_with_count(role: Role, user_counts: dict, year_role_counts: dict) -> dict:
    return {
        **role.model_dump(),
        "_count": {
            "users": user_counts.get(role.id, 0),
            "userYearRoles": year_role_counts.get(role.id, 0),
        },
    }


@router.get("/roles")
async def list_roles(
    includeInactive: bool = False,
    user: CurrentUser = Depends(require_permission(PERMISSIONS.ROLE_MANAGE)),
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
) -> list[dict]:
    stmt = select(Role).where(Role.tenantId == tenant_id)
    if not includeInactive:
        stmt = stmt.where(Role.isActive == True)  # noqa: E712
    stmt = stmt.order_by(Role.sortOrder)
    roles = (await session.execute(stmt)).scalars().all()
    user_counts, year_role_counts = await _role_counts(session, [r.id for r in roles])
    return [_role_with_count(r, user_counts, year_role_counts) for r in roles]


class RoleCreateBody(BaseModel):
    code: str | None = None
    name: str | None = None
    description: str | None = None
    stepNumber: int | None = None
    sortOrder: int | None = None
    permissions: list[str] | None = None


@router.post("/roles", status_code=status.HTTP_201_CREATED)
async def create_role(
    body: RoleCreateBody,
    user: CurrentUser = Depends(require_permission(PERMISSIONS.ROLE_MANAGE)),
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
) -> dict:
    if not body.code or not body.name:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "역할 코드와 이름은 필수입니다.")

    existing = (
        await session.execute(
            select(Role).where(Role.tenantId == tenant_id, Role.code == body.code)
        )
    ).scalars().first()
    if existing:
        raise HTTPException(status.HTTP_409_CONFLICT, "이미 존재하는 역할 코드입니다.")

    role = Role(
        tenantId=tenant_id,
        code=body.code,
        name=body.name,
        description=body.description,
        stepNumber=body.stepNumber,
        sortOrder=body.sortOrder or 0,
        permissions=sanitize_permissions(body.permissions),
    )
    session.add(role)
    await session.commit()
    await session.refresh(role)
    return role.model_dump()


@router.get("/roles/{role_id}")
async def get_role(
    role_id: str,
    user: CurrentUser = Depends(require_permission(PERMISSIONS.ROLE_MANAGE)),
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
) -> dict:
    role = await session.get(Role, role_id)
    if role is None or role.tenantId != tenant_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "역할을 찾을 수 없습니다.")
    user_counts, year_role_counts = await _role_counts(session, [role.id])
    return _role_with_count(role, user_counts, year_role_counts)


class RoleUpdateBody(BaseModel):
    code: str | None = None
    name: str | None = None
    description: str | None = None
    stepNumber: int | None = None
    sortOrder: int | None = None
    isActive: bool | None = None
    permissions: list[str] | None = None


@router.put("/roles/{role_id}")
async def update_role(
    role_id: str,
    body: RoleUpdateBody,
    user: CurrentUser = Depends(require_permission(PERMISSIONS.ROLE_MANAGE)),
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
) -> dict:
    role = await session.get(Role, role_id)
    if role is None or role.tenantId != tenant_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "역할을 찾을 수 없습니다.")

    fields = body.model_fields_set

    if "code" in fields and body.code and body.code != role.code:
        duplicate = (
            await session.execute(
                select(Role).where(Role.tenantId == tenant_id, Role.code == body.code)
            )
        ).scalars().first()
        if duplicate:
            raise HTTPException(status.HTTP_409_CONFLICT, "이미 존재하는 역할 코드입니다.")

    if "code" in fields and body.code is not None:
        role.code = body.code
    if "name" in fields and body.name is not None:
        role.name = body.name
    if "description" in fields:
        role.description = body.description
    if "stepNumber" in fields:
        role.stepNumber = body.stepNumber
    if "sortOrder" in fields and body.sortOrder is not None:
        role.sortOrder = body.sortOrder
    if "isActive" in fields and body.isActive is not None:
        role.isActive = body.isActive
    if "permissions" in fields:
        role.permissions = sanitize_permissions(body.permissions)

    await session.commit()
    await session.refresh(role)
    return role.model_dump()


@router.delete("/roles/{role_id}")
async def deactivate_role(
    role_id: str,
    user: CurrentUser = Depends(require_permission(PERMISSIONS.ROLE_MANAGE)),
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
) -> dict:
    role = await session.get(Role, role_id)
    if role is None or role.tenantId != tenant_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "역할을 찾을 수 없습니다.")

    if is_protected_system_role(role.code):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "기본 역할은 삭제할 수 없습니다.")

    user_counts, year_role_counts = await _role_counts(session, [role.id])
    users_count = user_counts.get(role.id, 0)
    year_roles_count = year_role_counts.get(role.id, 0)
    if users_count > 0 or year_roles_count > 0:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            {
                "error": "사용 중인 역할은 삭제할 수 없습니다.",
                "usedBy": {"users": users_count, "yearRoles": year_roles_count},
            },
        )

    role.isActive = False
    await session.commit()
    await session.refresh(role)
    return {"message": "역할이 비활성화되었습니다.", "role": role.model_dump()}


# ── D4: 초대 관리 (app/api/admin/invitations) ──────────────────────────


@router.get("/invitations")
async def list_invitations(
    user: CurrentUser = Depends(require_permission(PERMISSIONS.USER_REGISTER)),
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
) -> list[dict]:
    rows = (
        await session.execute(
            select(Invitation)
            .where(Invitation.tenantId == tenant_id)
            .order_by(Invitation.createdAt.desc())
        )
    ).scalars().all()
    return [r.model_dump() for r in rows]


class InvitationCreateBody(BaseModel):
    email: str | None = None
    role: str | None = None


@router.post("/invitations", status_code=status.HTTP_201_CREATED)
async def create_invitation(
    body: InvitationCreateBody,
    user: CurrentUser = Depends(require_permission(PERMISSIONS.USER_REGISTER)),
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
) -> dict:
    invalid_email = body.email is not None and not _EMAIL_RE.match(body.email)
    invalid_role = body.role is not None and body.role not in ("TENANT_ADMIN", "MEMBER")
    if invalid_email or invalid_role:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "초대 정보가 올바르지 않습니다. 이메일과 역할(TENANT_ADMIN/MEMBER)을 확인해주세요.",
        )

    invitation = Invitation(
        tenantId=tenant_id,
        email=body.email,
        role=body.role or "MEMBER",
        token=secrets.token_hex(32),
        expiresAt=datetime.now(timezone.utc) + timedelta(days=INVITATION_EXPIRES_DAYS),
        invitedById=user.id,
    )
    session.add(invitation)
    await session.commit()
    await session.refresh(invitation)
    return invitation.model_dump()
