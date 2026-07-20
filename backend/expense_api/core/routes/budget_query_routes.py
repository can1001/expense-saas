"""예산 조회 라우터 — hierarchy / search / simple / usage-details / memo-examples.
(app/api/budget/* 하위 조회 라우트 이전 — 응답 형태는 Next.js 라우트와 동일 유지)
"""

import re
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, HTTPException
from openpyxl.styles import Alignment, Font
from openpyxl.workbook import Workbook
from pydantic import BaseModel
from sqlalchemy import and_, exists, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from expense_api.core.db.session import get_session
from expense_api.core.dependencies.tenant import require_tenant_id
from expense_api.core.excel import (
    THIN_BORDER,
    set_column_widths,
    style_header_row,
    workbook_to_xlsx_response,
)
from expense_api.core.models.budget import (
    BudgetCategory,
    BudgetDetail,
    BudgetDetailYear,
    BudgetSubcategory,
    Committee,
    Department,
    DepartmentBudgetDetail,
)
from expense_api.core.models.expense import Expense, ExpenseItem
from expense_api.core.models.user import User, UserYearRole

router = APIRouter()

_APPROVED_STATUSES = ["APPROVED_STEP_1", "APPROVED_STEP_2", "APPROVED_FINAL"]
_ID_RE = re.compile(r"^[a-z0-9]{20,}$")


def _current_year() -> int:
    return datetime.now(timezone.utc).year


def _year_range(year: int) -> tuple[datetime, datetime]:
    return datetime(year, 1, 1), datetime(year + 1, 1, 1)


async def _load_committees_and_links(
    session: AsyncSession, tenant_id: str, year: int, committee_id: str = ""
) -> tuple[list[Committee], list[Department], dict[str, list]]:
    """위원회/사역팀 + 부서-세목 연결(세목/목/항/연도설정/담당자 조인) 로드.

    `budget_hierarchy` 와 `budget_hierarchy_export` 가 공유하는 데이터 소스.
    """
    comm_stmt = (
        select(Committee)
        .where(Committee.tenantId == tenant_id, Committee.isActive.is_(True))
        .order_by(Committee.sortOrder)
    )
    if committee_id:
        comm_stmt = comm_stmt.where(Committee.id == committee_id)
    committees = (await session.execute(comm_stmt)).scalars().all()

    departments = (
        (
            await session.execute(
                select(Department)
                .where(Department.tenantId == tenant_id, Department.isActive.is_(True))
                .order_by(Department.sortOrder)
            )
        )
        .scalars()
        .all()
    )

    # 부서-세목 연결 + 세목/목/항 + 연도 설정 + 담당자 (일괄 조인)
    link_rows = (
        await session.execute(
            select(DepartmentBudgetDetail, BudgetDetail, BudgetSubcategory, BudgetCategory,
                   BudgetDetailYear, User)
            .join(BudgetDetail, BudgetDetail.id == DepartmentBudgetDetail.budgetDetailId)
            .join(BudgetSubcategory, BudgetSubcategory.id == BudgetDetail.subcategoryId,
                  isouter=True)
            .join(BudgetCategory, BudgetCategory.id == BudgetSubcategory.categoryId, isouter=True)
            .join(
                BudgetDetailYear,
                and_(
                    BudgetDetailYear.budgetDetailId == BudgetDetail.id,
                    BudgetDetailYear.year == year,
                    BudgetDetailYear.isActive.is_(True),
                ),
                isouter=True,
            )
            .join(User, User.id == BudgetDetailYear.managerId, isouter=True)
            .where(
                DepartmentBudgetDetail.tenantId == tenant_id,
                DepartmentBudgetDetail.isActive.is_(True),
                BudgetDetail.isActive.is_(True),
            )
        )
    ).all()

    links_by_dept: dict[str, list] = {}
    for row in link_rows:
        links_by_dept.setdefault(row[0].departmentId, []).append(row)

    return committees, departments, links_by_dept


# ── GET /api/budget/hierarchy — 조직별 예산 계층 조회 ─────────────────
@router.get("/hierarchy")
async def budget_hierarchy(
    year: int | None = None,
    search: str = "",
    committeeId: str = "",
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
) -> dict:
    yr = year or _current_year()
    start, end = _year_range(yr)

    # 지급완료 지출의 세목별 사용금액 집계
    used_rows = (
        await session.execute(
            select(ExpenseItem.budgetDetail, func.coalesce(func.sum(ExpenseItem.amount), 0))
            .join(Expense, Expense.id == ExpenseItem.expenseId)
            .where(
                Expense.tenantId == tenant_id,
                Expense.paymentStatus == "COMPLETED",
                Expense.expenseDate >= start,
                Expense.expenseDate < end,
            )
            .group_by(ExpenseItem.budgetDetail)
        )
    ).all()
    used_map = {name: total or 0 for name, total in used_rows}

    committees, departments, links_by_dept = await _load_committees_and_links(
        session, tenant_id, yr, committeeId
    )

    term = search.lower()
    total_details = 0
    total_budget_amount = 0
    unassigned_count = 0

    formatted_committees = []
    for committee in committees:
        depts_out = []
        for dept in departments:
            if dept.committeeId != committee.id:
                continue
            details = []
            for link, detail, sub, cat, year_setting, manager in links_by_dept.get(dept.id, []):
                category = cat.name if cat else ""
                subcategory = sub.name if sub else ""
                manager_name = manager.username if manager else None
                d = {
                    "id": link.id,
                    "detailId": detail.id,
                    "detailName": detail.name,
                    "category": category,
                    "subcategory": subcategory,
                    "fullPath": f"{category} > {subcategory} > {detail.name}",
                    "managerId": year_setting.managerId if year_setting else None,
                    "managerName": manager_name,
                    "budgetAmount": year_setting.budgetAmount if year_setting else 0,
                    "usedAmount": used_map.get(detail.name, 0),
                }
                if term and not (
                    term in d["detailName"].lower()
                    or term in category.lower()
                    or term in subcategory.lower()
                    or (manager_name and term in manager_name.lower())
                    or term in committee.name.lower()
                    or term in dept.name.lower()
                ):
                    continue
                details.append(d)

            details.sort(key=lambda d: (d["category"], d["subcategory"], d["detailName"]))
            for d in details:
                total_details += 1
                total_budget_amount += d["budgetAmount"]
                if not d["managerId"]:
                    unassigned_count += 1

            if details:
                depts_out.append(
                    {"id": dept.id, "name": dept.name, "detailCount": len(details),
                     "details": details}
                )

        if depts_out:
            formatted_committees.append(
                {
                    "id": committee.id,
                    "name": committee.name,
                    "departmentCount": len(depts_out),
                    "detailCount": sum(d["detailCount"] for d in depts_out),
                    "departments": depts_out,
                }
            )

    all_committees = (
        await session.execute(
            select(Committee.id, Committee.name)
            .where(Committee.tenantId == tenant_id, Committee.isActive.is_(True))
            .order_by(Committee.sortOrder)
        )
    ).all()

    return {
        "summary": {
            "totalCommittees": len(formatted_committees),
            "totalDepartments": sum(c["departmentCount"] for c in formatted_committees),
            "totalDetails": total_details,
            "totalBudgetAmount": total_budget_amount,
            "unassignedCount": unassigned_count,
        },
        "committees": formatted_committees,
        "allCommittees": [{"id": cid, "name": name} for cid, name in all_committees],
    }


_EXPORT_COLUMN_WIDTHS = [15, 15, 15, 15, 20, 12, 15]
_EXPORT_HEADERS = ["위원회", "사역팀", "예산(항)", "예산(목)", "예산(세목)", "담당자", "예산금액"]


# ── GET /api/budget/hierarchy/export — 예산 현황 Excel 내보내기 ────────
@router.get("/hierarchy/export")
async def budget_hierarchy_export(
    year: int | None = None,
    committeeId: str = "",
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
):
    yr = year or _current_year()
    committees, departments, links_by_dept = await _load_committees_and_links(
        session, tenant_id, yr, committeeId
    )

    rows: list[list] = []
    for committee in committees:
        for dept in departments:
            if dept.committeeId != committee.id:
                continue
            for link, detail, sub, cat, year_setting, manager in links_by_dept.get(dept.id, []):
                rows.append(
                    [
                        committee.name,
                        dept.name,
                        cat.name if cat else "",
                        sub.name if sub else "",
                        detail.name,
                        manager.username if manager else "",
                        year_setting.budgetAmount if year_setting else 0,
                    ]
                )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = f"{yr}년 예산현황"

    worksheet.append(_EXPORT_HEADERS)
    style_header_row(worksheet)
    set_column_widths(worksheet, _EXPORT_COLUMN_WIDTHS)

    for row in rows:
        worksheet.append(row)

    last_data_row = worksheet.max_row
    for excel_row in worksheet.iter_rows(min_row=2, max_row=last_data_row):
        for cell in excel_row:
            cell.border = THIN_BORDER
        amount_cell = excel_row[6]
        amount_cell.number_format = "#,##0"
        amount_cell.alignment = Alignment(horizontal="right")

    if rows:
        worksheet.append([])
        total_row_idx = worksheet.max_row + 1
        worksheet.append(["", "", "", "", "합계", "", f"=SUM(G2:G{last_data_row})"])
        for cell in worksheet[total_row_idx]:
            cell.font = Font(bold=True)
        total_amount_cell = worksheet.cell(row=total_row_idx, column=7)
        total_amount_cell.number_format = "#,##0"
        total_amount_cell.alignment = Alignment(horizontal="right")

    worksheet.auto_filter.ref = "A1:G1"

    return workbook_to_xlsx_response(workbook, f"budget_{yr}.xlsx")


# ── GET /api/budget/search — 계정과목 검색 ────────────────────────────
@router.get("/search")
async def budget_search(
    q: str = "",
    departmentId: str | None = None,
    year: int | None = None,
    limit: int = 20,
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
) -> dict:
    yr = year or _current_year()
    limit = min(limit, 50)

    if len(q) < 1:
        return {"results": [], "total": 0}

    pattern = f"%{q.lower()}%"
    year_exists = exists(
        select(BudgetDetailYear.id).where(
            BudgetDetailYear.budgetDetailId == BudgetDetail.id,
            BudgetDetailYear.year == yr,
            BudgetDetailYear.isActive.is_(True),
        )
    )
    conditions = [
        BudgetDetail.tenantId == tenant_id,
        BudgetDetail.isActive.is_(True),
        year_exists,
        (
            func.lower(BudgetDetail.name).like(pattern)
            | func.lower(BudgetSubcategory.name).like(pattern)
            | func.lower(BudgetCategory.name).like(pattern)
        ),
    ]
    if departmentId:
        conditions.append(
            exists(
                select(DepartmentBudgetDetail.id).where(
                    DepartmentBudgetDetail.budgetDetailId == BudgetDetail.id,
                    DepartmentBudgetDetail.departmentId == departmentId,
                    DepartmentBudgetDetail.isActive.is_(True),
                )
            )
        )

    base = (
        select(BudgetDetail, BudgetSubcategory, BudgetCategory)
        .join(BudgetSubcategory, BudgetSubcategory.id == BudgetDetail.subcategoryId)
        .join(BudgetCategory, BudgetCategory.id == BudgetSubcategory.categoryId)
        .where(*conditions)
    )
    rows = (
        await session.execute(
            base.order_by(
                BudgetCategory.sortOrder, BudgetSubcategory.sortOrder, BudgetDetail.sortOrder
            ).limit(limit)
        )
    ).all()
    total_count = (
        await session.execute(
            select(func.count())
            .select_from(BudgetDetail)
            .join(BudgetSubcategory, BudgetSubcategory.id == BudgetDetail.subcategoryId)
            .join(BudgetCategory, BudgetCategory.id == BudgetSubcategory.categoryId)
            .where(*conditions)
        )
    ).scalar_one()

    detail_ids = [d.id for d, _, _ in rows]
    managers: dict[str, tuple[str, str]] = {}
    dept_links: dict[str, list[dict]] = {}
    if detail_ids:
        manager_rows = (
            await session.execute(
                select(BudgetDetailYear.budgetDetailId, User.id, User.username)
                .join(User, User.id == BudgetDetailYear.managerId)
                .where(
                    BudgetDetailYear.budgetDetailId.in_(detail_ids),
                    BudgetDetailYear.year == yr,
                )
            )
        ).all()
        managers = {r[0]: (r[1], r[2]) for r in manager_rows}

        link_rows = (
            await session.execute(
                select(DepartmentBudgetDetail.budgetDetailId, Department, Committee)
                .join(Department, Department.id == DepartmentBudgetDetail.departmentId)
                .join(Committee, Committee.id == Department.committeeId)
                .where(
                    DepartmentBudgetDetail.budgetDetailId.in_(detail_ids),
                    DepartmentBudgetDetail.isActive.is_(True),
                )
            )
        ).all()
        for did, dept, comm in link_rows:
            dept_links.setdefault(did, []).append(
                {"id": dept.id, "name": dept.name,
                 "committee": {"id": comm.id, "name": comm.name}}
            )

    results = []
    for detail, sub, cat in rows:
        manager = managers.get(detail.id)
        links = dept_links.get(detail.id, [])
        department = links[0] if links else None
        if departmentId:
            matched = next((link for link in links if link["id"] == departmentId), None)
            if matched:
                department = matched
        results.append(
            {
                "id": detail.id,
                "detail": detail.name,
                "subcategory": sub.name,
                "category": cat.name,
                "fullPath": f"{cat.name} > {sub.name} > {detail.name}",
                "managerId": manager[0] if manager else None,
                "managerName": manager[1] if manager else None,
                "hierarchy": {
                    "committee": department["committee"]["name"] if department else "",
                    "department": department["name"] if department else "",
                    "category": cat.name,
                    "subcategory": sub.name,
                    "detail": detail.name,
                },
            }
        )

    return {"results": results, "total": total_count, "showing": len(results)}


# ── POST /api/budget/simple — 간편 예산 선택 (재정팀장 담당 세목만) ────
class SimpleCascadeRequest(BaseModel):
    category: str | None = None
    subcategory: str | None = None
    year: int | None = None


@router.post("/simple")
async def budget_simple(
    body: SimpleCascadeRequest | None = None,
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
) -> dict:
    b = body or SimpleCascadeRequest()
    yr = b.year or _current_year()

    finance_head_id = (
        await session.execute(
            select(UserYearRole.userId)
            .join(User, User.id == UserYearRole.userId)
            .where(
                UserYearRole.tenantId == tenant_id,
                UserYearRole.year == yr,
                UserYearRole.role == "finance_head",
                User.isActive.is_(True),
            )
            .limit(1)
        )
    ).scalar_one_or_none()

    options: list[str] = []

    def _detail_exists_for(subcategory_id_col):
        return exists(
            select(BudgetDetail.id).where(
                BudgetDetail.subcategoryId == subcategory_id_col,
                BudgetDetail.isActive.is_(True),
                exists(
                    select(BudgetDetailYear.id).where(
                        BudgetDetailYear.budgetDetailId == BudgetDetail.id,
                        BudgetDetailYear.year == yr,
                        BudgetDetailYear.managerId == finance_head_id,
                    )
                ),
            )
        )

    if not b.category:
        field = "categories"
        if finance_head_id:
            rows = (
                await session.execute(
                    select(BudgetCategory.name)
                    .where(
                        BudgetCategory.tenantId == tenant_id,
                        BudgetCategory.isActive.is_(True),
                        exists(
                            select(BudgetSubcategory.id).where(
                                BudgetSubcategory.categoryId == BudgetCategory.id,
                                BudgetSubcategory.isActive.is_(True),
                                _detail_exists_for(BudgetSubcategory.id),
                            )
                        ),
                    )
                    .order_by(BudgetCategory.sortOrder)
                )
            ).all()
            options = [r[0] for r in rows]
    elif not b.subcategory:
        field = "subcategories"
        if finance_head_id:
            category = (
                await session.execute(
                    select(BudgetCategory).where(
                        BudgetCategory.tenantId == tenant_id,
                        BudgetCategory.name == b.category,
                        BudgetCategory.isActive.is_(True),
                    )
                )
            ).scalars().first()
            if category:
                rows = (
                    await session.execute(
                        select(BudgetSubcategory.name)
                        .where(
                            BudgetSubcategory.categoryId == category.id,
                            BudgetSubcategory.isActive.is_(True),
                            _detail_exists_for(BudgetSubcategory.id),
                        )
                        .order_by(BudgetSubcategory.sortOrder)
                    )
                ).all()
                options = [r[0] for r in rows]
    else:
        field = "details"
        if finance_head_id:
            subcategory = (
                await session.execute(
                    select(BudgetSubcategory)
                    .join(BudgetCategory, BudgetCategory.id == BudgetSubcategory.categoryId)
                    .where(
                        BudgetSubcategory.tenantId == tenant_id,
                        BudgetSubcategory.name == b.subcategory,
                        BudgetCategory.name == b.category,
                        BudgetSubcategory.isActive.is_(True),
                    )
                )
            ).scalars().first()
            if subcategory:
                rows = (
                    await session.execute(
                        select(BudgetDetail.name)
                        .where(
                            BudgetDetail.subcategoryId == subcategory.id,
                            BudgetDetail.isActive.is_(True),
                            exists(
                                select(BudgetDetailYear.id).where(
                                    BudgetDetailYear.budgetDetailId == BudgetDetail.id,
                                    BudgetDetailYear.year == yr,
                                    BudgetDetailYear.managerId == finance_head_id,
                                )
                            ),
                        )
                        .order_by(BudgetDetail.sortOrder)
                    )
                ).all()
                options = [r[0] for r in rows]

    return {"field": field, "options": sorted(o for o in options if o)}


# ── GET /api/budget/simple/all-details — 전체 세목 + 부모/담당자 ───────
@router.get("/simple/all-details")
async def budget_all_details(
    year: int | None = None,
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
) -> dict:
    yr = year or _current_year()

    rows = (
        await session.execute(
            select(BudgetDetail, BudgetSubcategory, BudgetCategory, User)
            .join(BudgetSubcategory, BudgetSubcategory.id == BudgetDetail.subcategoryId)
            .join(BudgetCategory, BudgetCategory.id == BudgetSubcategory.categoryId)
            .join(
                BudgetDetailYear,
                and_(
                    BudgetDetailYear.budgetDetailId == BudgetDetail.id,
                    BudgetDetailYear.year == yr,
                ),
            )
            .join(User, User.id == BudgetDetailYear.managerId, isouter=True)
            .where(BudgetDetail.tenantId == tenant_id, BudgetDetail.isActive.is_(True))
            .order_by(BudgetDetail.name)
        )
    ).all()

    return {
        "details": [
            {
                "name": detail.name,
                "category": cat.name,
                "subcategory": sub.name,
                "managerId": manager.id if manager else None,
                "managerName": manager.username if manager else None,
            }
            for detail, sub, cat, manager in rows
        ]
    }


# ── GET /api/budget/usage-details — 항/목/세목별 사용 내역 ─────────────
@router.get("/usage-details")
async def budget_usage_details(
    budgetCategory: str | None = None,
    budgetSubcategory: str | None = None,
    budgetDetail: str | None = None,
    year: int | None = None,
    excludeExpenseId: str | None = None,
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
) -> dict:
    if not budgetCategory:
        raise HTTPException(400, "항(budgetCategory) 파라미터가 필요합니다.")
    if not budgetSubcategory:
        raise HTTPException(400, "목(budgetSubcategory) 파라미터가 필요합니다.")
    if not budgetDetail:
        raise HTTPException(400, "세목(budgetDetail) 파라미터가 필요합니다.")
    if year is None:
        raise HTTPException(400, "연도(year) 파라미터가 필요합니다.")
    if excludeExpenseId and not _ID_RE.match(excludeExpenseId):
        raise HTTPException(400, "잘못된 지출결의서 ID 형식입니다.")

    start, end = _year_range(year)
    conditions = [
        Expense.tenantId == tenant_id,
        ExpenseItem.budgetCategory == budgetCategory,
        ExpenseItem.budgetSubcategory == budgetSubcategory,
        ExpenseItem.budgetDetail == budgetDetail,
        Expense.status.in_(_APPROVED_STATUSES),
        Expense.requestDate >= start,
        Expense.requestDate < end,
    ]
    if excludeExpenseId:
        conditions.append(Expense.id != excludeExpenseId)

    rows = (
        await session.execute(
            select(ExpenseItem, Expense)
            .join(Expense, Expense.id == ExpenseItem.expenseId)
            .where(*conditions)
            .order_by(Expense.requestDate.desc())
        )
    ).all()

    items = [
        {
            "id": item.id,
            "expenseId": expense.id,
            "requestDate": expense.requestDate.date().isoformat(),
            "applicantName": expense.applicantName,
            "description": item.description,
            "amount": item.amount,
            "status": expense.status,
        }
        for item, expense in rows
    ]

    return {
        "budgetCategory": budgetCategory,
        "budgetSubcategory": budgetSubcategory,
        "budgetDetail": budgetDetail,
        "year": year,
        "items": items,
        "totalAmount": sum(item["amount"] for item in items),
        "count": len(items),
    }


# ── GET /api/budget/memo-examples — 세목 적요 예제 ─────────────────────
@router.get("/memo-examples")
async def budget_memo_examples(
    budgetDetailId: str | None = None,
    budgetDetailName: str | None = None,
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
) -> dict:
    if not budgetDetailId and not budgetDetailName:
        raise HTTPException(400, "budgetDetailId 또는 budgetDetailName이 필요합니다.")

    detail = None
    if budgetDetailId:
        candidate = await session.get(BudgetDetail, budgetDetailId)
        if candidate and candidate.tenantId == tenant_id:
            detail = candidate
    if detail is None and budgetDetailName:
        detail = (
            await session.execute(
                select(BudgetDetail).where(
                    BudgetDetail.tenantId == tenant_id,
                    BudgetDetail.name == budgetDetailName,
                )
            )
        ).scalars().first()

    if detail is None:
        return {"examples": [], "budgetDetail": None}

    examples = (
        [s.strip() for s in detail.description.split(",") if s.strip()]
        if detail.description
        else []
    )
    return {"examples": examples, "budgetDetail": {"id": detail.id, "name": detail.name}}
