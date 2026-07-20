"""예산 마스터 CRUD 라우터 — 위원회/부서/항/목/세목 목록·생성·수정·삭제.
(app/api/committees, departments, budget-categories, budget-subcategories, budget-details 이전)
"""

import io
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile, status
from fastapi.responses import JSONResponse
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel
from sqlalchemy import delete, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from expense_api.core.auth.permissions import PERMISSIONS
from expense_api.core.db.session import get_session
from expense_api.core.dependencies.authz import require_permission
from expense_api.core.dependencies.tenant import require_tenant_id
from expense_api.core.excel import set_column_widths, style_header_row, workbook_to_xlsx_response
from expense_api.core.models.budget import (
    BudgetCategory,
    BudgetDetail,
    BudgetDetailYear,
    BudgetSubcategory,
    Committee,
    Department,
    DepartmentBudgetDetail,
)
from expense_api.core.models.expense import Expense, ExpenseItem
from expense_api.core.models.ids import utcnow
from expense_api.core.models.user import Role, User, UserYearRole

router = APIRouter()

_APPROVED_STATUSES = ["APPROVED_STEP_1", "APPROVED_STEP_2", "APPROVED_FINAL"]


def _upload_error(
    message: str,
    *,
    error_type: str = "VALIDATION_ERROR",
    status_code: int = 400,
    fields: list[dict] | None = None,
) -> JSONResponse:
    error: dict = {"type": error_type, "message": message}
    if fields:
        error["fields"] = fields
    return JSONResponse({"success": False, "error": error}, status_code=status_code)


def _upload_success(message: str, data: dict) -> JSONResponse:
    return JSONResponse({"success": True, "message": message, "data": data})


# ── 공통 헬퍼 ─────────────────────────────────────────────────────────
async def _next_sort_order(session: AsyncSession, model, tenant_id: str) -> int:
    stmt = select(func.max(model.sortOrder)).where(model.tenantId == tenant_id)
    current = (await session.execute(stmt)).scalar_one_or_none()
    return (current or 0) + 1


async def _exists_by_name(session: AsyncSession, model, tenant_id: str, name: str, **extra) -> bool:
    stmt = select(model.id).where(model.tenantId == tenant_id, model.name == name)
    for k, v in extra.items():
        stmt = stmt.where(getattr(model, k) == v)
    return (await session.execute(stmt)).first() is not None


async def _list(session: AsyncSession, model, tenant_id: str, **extra) -> list[dict]:
    stmt = select(model).where(model.tenantId == tenant_id)
    for k, v in extra.items():
        stmt = stmt.where(getattr(model, k) == v)
    stmt = stmt.order_by(model.sortOrder)
    rows = (await session.execute(stmt)).scalars().all()
    return [r.model_dump() for r in rows]


async def _get_or_404(session: AsyncSession, model, entity_id: str, tenant_id: str, message: str):
    entity = await session.get(model, entity_id)
    if entity is None or entity.tenantId != tenant_id:
        raise HTTPException(404, message)
    return entity


# ── 위원회 ────────────────────────────────────────────────────────────
class CommitteeCreate(BaseModel):
    name: str
    leaderId: str | None = None


@router.get("/committees")
async def list_committees(
    tenant_id: str = Depends(require_tenant_id), session: AsyncSession = Depends(get_session)
) -> dict:
    committees = await _list(session, Committee, tenant_id)
    if not committees:
        return {"committees": []}

    leader_ids = {c["leaderId"] for c in committees if c.get("leaderId")}
    leaders: dict[str, str] = {}
    if leader_ids:
        rows = (
            await session.execute(select(User.id, User.username).where(User.id.in_(leader_ids)))
        ).all()
        leaders = {row[0]: row[1] for row in rows}

    counts = dict(
        (
            await session.execute(
                select(Department.committeeId, func.count(Department.id))
                .where(
                    Department.tenantId == tenant_id,
                    Department.committeeId.in_([c["id"] for c in committees]),
                )
                .group_by(Department.committeeId)
            )
        ).all()
    )

    for c in committees:
        leader_id = c.get("leaderId")
        c["leader"] = (
            {"id": leader_id, "username": leaders[leader_id]}
            if leader_id and leader_id in leaders
            else None
        )
        c["_count"] = {"departments": counts.get(c["id"], 0)}

    return {"committees": committees}


@router.post("/committees", status_code=status.HTTP_201_CREATED)
async def create_committee(
    body: CommitteeCreate,
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
    _=Depends(require_permission(PERMISSIONS.COMMITTEE_MANAGE)),
) -> dict:
    name = body.name.strip()
    if not name:
        raise HTTPException(400, "위원회명을 입력해주세요.")
    if await _exists_by_name(session, Committee, tenant_id, name):
        raise HTTPException(409, "이미 존재하는 위원회명입니다.")
    entity = Committee(
        tenantId=tenant_id,
        name=name,
        sortOrder=await _next_sort_order(session, Committee, tenant_id),
        leaderId=body.leaderId or None,
    )
    session.add(entity)
    await session.commit()
    await session.refresh(entity)
    return entity.model_dump()


class CommitteeUpdate(BaseModel):
    name: str | None = None
    isActive: bool | None = None
    sortOrder: int | None = None
    leaderId: str | None = None


@router.patch("/committees/{committee_id}")
async def update_committee(
    committee_id: str,
    body: CommitteeUpdate,
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
    _=Depends(require_permission(PERMISSIONS.COMMITTEE_MANAGE)),
) -> dict:
    entity = await _get_or_404(session, Committee, committee_id, tenant_id, "위원회를 찾을 수 없습니다.")
    fields = body.model_fields_set

    if "name" in fields:
        new_name = (body.name or "").strip()
        if body.name and new_name != entity.name:
            if await _exists_by_name(session, Committee, tenant_id, new_name):
                raise HTTPException(409, "이미 존재하는 위원회명입니다.")
        entity.name = new_name
    if "isActive" in fields:
        entity.isActive = body.isActive
    if "sortOrder" in fields:
        entity.sortOrder = body.sortOrder
    if "leaderId" in fields:
        entity.leaderId = body.leaderId or None

    await session.commit()
    await session.refresh(entity)
    return entity.model_dump()


@router.delete("/committees/{committee_id}")
async def delete_committee(
    committee_id: str,
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
    _=Depends(require_permission(PERMISSIONS.COMMITTEE_MANAGE)),
) -> dict:
    entity = await _get_or_404(
        session, Committee, committee_id, tenant_id, "위원회를 찾을 수 없습니다."
    )

    dept_count = (
        await session.execute(
            select(func.count(Department.id)).where(Department.committeeId == committee_id)
        )
    ).scalar_one()
    if dept_count > 0:
        raise HTTPException(400, "하위 사역팀이 있는 위원회는 삭제할 수 없습니다.")

    await session.delete(entity)
    await session.commit()
    return {"success": True}


# ── 부서 ──────────────────────────────────────────────────────────────
class DepartmentCreate(BaseModel):
    committeeId: str
    name: str


@router.get("/departments")
async def list_departments(
    committeeId: str | None = None,
    year: int | None = None,
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
) -> dict:
    extra = {"committeeId": committeeId} if committeeId else {}
    departments = await _list(session, Department, tenant_id, **extra)
    if not departments:
        return {"departments": []}

    committee_ids = {d["committeeId"] for d in departments}
    committee_names = dict(
        (
            await session.execute(select(Committee.id, Committee.name).where(Committee.id.in_(committee_ids)))
        ).all()
    )

    target_year = year or utcnow().year
    leader_rows = (
        await session.execute(
            select(UserYearRole.departmentId, User.id, User.username)
            .join(User, User.id == UserYearRole.userId)
            .where(
                UserYearRole.departmentId.in_([d["id"] for d in departments]),
                UserYearRole.year == target_year,
                UserYearRole.role == "team_leader",
            )
        )
    ).all()
    leaders = {row[0]: (row[1], row[2]) for row in leader_rows}

    return {
        "departments": [
            {
                "id": d["id"],
                "name": d["name"],
                "committeeId": d["committeeId"],
                "committeeName": committee_names.get(d["committeeId"], ""),
                "sortOrder": d["sortOrder"],
                "isActive": d["isActive"],
                "leaderId": leaders[d["id"]][0] if d["id"] in leaders else None,
                "leaderName": leaders[d["id"]][1] if d["id"] in leaders else None,
            }
            for d in departments
        ]
    }


@router.post("/departments", status_code=status.HTTP_201_CREATED)
async def create_department(
    body: DepartmentCreate,
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
    _=Depends(require_permission(PERMISSIONS.DEPARTMENT_MANAGE)),
) -> dict:
    name = body.name.strip()
    if not name:
        raise HTTPException(400, "부서명을 입력해주세요.")
    # @@unique([committeeId, name])
    if await _exists_by_name(session, Department, tenant_id, name, committeeId=body.committeeId):
        raise HTTPException(409, "해당 위원회에 이미 존재하는 부서명입니다.")
    entity = Department(
        tenantId=tenant_id,
        committeeId=body.committeeId,
        name=name,
        sortOrder=await _next_sort_order(session, Department, tenant_id),
    )
    session.add(entity)
    await session.commit()
    await session.refresh(entity)
    return entity.model_dump()


class DepartmentUpdate(BaseModel):
    name: str | None = None
    isActive: bool | None = None
    sortOrder: int | None = None
    committeeId: str | None = None


@router.patch("/departments/{department_id}")
async def update_department(
    department_id: str,
    body: DepartmentUpdate,
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
    _=Depends(require_permission(PERMISSIONS.DEPARTMENT_MANAGE)),
) -> dict:
    entity = await _get_or_404(session, Department, department_id, tenant_id, "사역팀을 찾을 수 없습니다.")
    fields = body.model_fields_set
    target_committee_id = body.committeeId if "committeeId" in fields else entity.committeeId

    if "name" in fields:
        new_name = (body.name or "").strip()
        if body.name and new_name != entity.name:
            if await _exists_by_name(
                session, Department, tenant_id, new_name, committeeId=target_committee_id
            ):
                raise HTTPException(409, "같은 위원회 내에 이미 존재하는 사역팀명입니다.")
        entity.name = new_name
    if "isActive" in fields:
        entity.isActive = body.isActive
    if "sortOrder" in fields:
        entity.sortOrder = body.sortOrder
    if "committeeId" in fields:
        entity.committeeId = body.committeeId

    await session.commit()
    await session.refresh(entity)
    return entity.model_dump()


@router.delete("/departments/{department_id}")
async def delete_department(
    department_id: str,
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
    _=Depends(require_permission(PERMISSIONS.DEPARTMENT_MANAGE)),
) -> dict:
    entity = await _get_or_404(session, Department, department_id, tenant_id, "사역팀을 찾을 수 없습니다.")

    detail_count = (
        await session.execute(
            select(func.count(DepartmentBudgetDetail.id)).where(
                DepartmentBudgetDetail.departmentId == department_id
            )
        )
    ).scalar_one()
    if detail_count > 0:
        raise HTTPException(
            400, "연결된 예산 세목이 있는 사역팀은 삭제할 수 없습니다. 비활성화를 사용해주세요."
        )

    year_role_count = (
        await session.execute(
            select(func.count(UserYearRole.id)).where(UserYearRole.departmentId == department_id)
        )
    ).scalar_one()
    if year_role_count > 0:
        raise HTTPException(
            400, "연결된 팀장 역할이 있는 사역팀은 삭제할 수 없습니다. 비활성화를 사용해주세요."
        )

    await session.delete(entity)
    await session.commit()
    return {"success": True}


# ── 사역팀장 엑셀 업로드/템플릿 ──────────────────────────────────────
# (app/api/departments/leaders-upload/route.ts 이전 — UserYearRole(team_leader) upsert/삭제)
_LEADERS_UPLOAD_GUIDE_ROWS_TAIL = [
    ("위원회", "위원회 이름 (필수)", "교육위원회"),
    ("사역팀", "사역팀 이름 (필수)", "유년부"),
    ("팀장", "팀장 이름 (사용자 이름과 일치해야 함)", "정혜종"),
    ("", "", ""),
    ("※ 참고사항", "", ""),
    ("- 팀장 비우기", "팀장 열을 비워두면 팀장이 해제됩니다", ""),
    ("- 사용자 매칭", "팀장 이름이 정확히 일치해야 합니다", ""),
    ("- 연도별 관리", "팀장은 연도별로 관리됩니다", ""),
]


@router.get("/departments/leaders-upload")
async def departments_leaders_upload_template(
    year: int | None = None,
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
    _=Depends(require_permission(PERMISSIONS.DEPARTMENT_MANAGE)),
):
    yr = year or utcnow().year

    departments = (
        await session.execute(
            select(Department, Committee)
            .join(Committee, Committee.id == Department.committeeId)
            .where(Department.tenantId == tenant_id, Department.isActive == True)  # noqa: E712
            .order_by(Committee.sortOrder, Department.sortOrder)
        )
    ).all()

    leaders: dict[str, str] = {}
    dept_ids = [d.id for d, _ in departments]
    if dept_ids:
        leader_rows = (
            await session.execute(
                select(UserYearRole.departmentId, User.username)
                .join(User, User.id == UserYearRole.userId)
                .where(
                    UserYearRole.departmentId.in_(dept_ids),
                    UserYearRole.year == yr,
                    UserYearRole.role == "team_leader",
                )
            )
        ).all()
        for did, username in leader_rows:
            leaders.setdefault(did, username)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "사역팀장목록"
    sheet.append(["위원회", "사역팀", "팀장"])
    style_header_row(sheet)
    set_column_widths(sheet, [20, 25, 15])

    if not departments:
        sheet.append(["", "", ""])
    else:
        for dept, comm in departments:
            sheet.append([comm.name, dept.name, leaders.get(dept.id, "")])

    guide = workbook.create_sheet("작성안내")
    guide.append(["항목", "설명", "예시"])
    style_header_row(guide)
    set_column_widths(guide, [20, 50, 20])
    guide.append(["적용 연도", str(yr), ""])
    guide.append(["", "", ""])
    for row in _LEADERS_UPLOAD_GUIDE_ROWS_TAIL:
        guide.append(list(row))

    date_str = datetime.now(timezone.utc).date().isoformat()
    return workbook_to_xlsx_response(workbook, f"leaders_template_{yr}_{date_str}.xlsx")


@router.post("/departments/leaders-upload")
async def departments_leaders_upload_post(
    file: UploadFile | None = File(None),
    dryRun: str = Form("false"),
    year: str = Form(""),
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
    _=Depends(require_permission(PERMISSIONS.DEPARTMENT_MANAGE)),
):
    if file is None:
        return _upload_error("파일이 필요합니다.")

    try:
        yr = int(year) if year else utcnow().year
    except ValueError:
        yr = utcnow().year

    try:
        data = await file.read()
        workbook = load_workbook(io.BytesIO(data), data_only=True)

        if not workbook.worksheets:
            return _upload_error("워크시트를 찾을 수 없습니다.")
        sheet = workbook.worksheets[0]

        headers: dict[int, str] = {}
        for cell in sheet[1]:
            headers[cell.column] = str(cell.value or "").lower()

        raw_rows: list[dict[str, object]] = []
        for row in sheet.iter_rows(min_row=2):
            row_data: dict[str, object] = {}
            for cell in row:
                header = headers.get(cell.column)
                if header:
                    row_data[header] = cell.value
            if any(v is not None and v != "" for v in row_data.values()):
                raw_rows.append(row_data)

        if not raw_rows:
            return _upload_error("데이터가 없습니다.")

        def _norm(value: object) -> str:
            return str(value).strip() if value is not None else ""

        rows: list[dict] = []
        for raw in raw_rows:
            normalized = {"committee": "", "department": "", "leader": ""}
            for key, value in raw.items():
                lowered = key.lower()
                if "위원회" in lowered or "committee" in lowered:
                    normalized["committee"] = _norm(value)
                elif "사역팀" in lowered or "department" in lowered or "부서" in lowered:
                    normalized["department"] = _norm(value)
                elif "팀장" in lowered or "leader" in lowered or "담당" in lowered:
                    normalized["leader"] = _norm(value)
            rows.append(normalized)

        committees = (
            await session.execute(
                select(Committee).where(
                    Committee.tenantId == tenant_id, Committee.isActive == True  # noqa: E712
                )
            )
        ).scalars().all()
        committee_map = {c.name: c.id for c in committees}
        committee_ids = [c.id for c in committees]

        dept_rows = []
        if committee_ids:
            dept_rows = (
                await session.execute(
                    select(Department).where(
                        Department.committeeId.in_(committee_ids),
                        Department.isActive == True,  # noqa: E712
                    )
                )
            ).scalars().all()
        department_map = {f"{d.committeeId}|{d.name}": d.id for d in dept_rows}

        user_rows = (
            await session.execute(
                select(User.id, User.username).where(
                    User.tenantId == tenant_id, User.isActive == True  # noqa: E712
                )
            )
        ).all()
        user_map = {username: uid for uid, username in user_rows}

        validation_errors: list[dict] = []
        to_update: list[dict] = []
        for index, row in enumerate(rows):
            row_num = index + 2
            if not row["committee"] and not row["department"]:
                continue
            if not row["committee"]:
                validation_errors.append(
                    {"fieldName": f"행 {row_num}", "message": "위원회가 비어있습니다."}
                )
                continue
            if not row["department"]:
                validation_errors.append(
                    {"fieldName": f"행 {row_num}", "message": "사역팀이 비어있습니다."}
                )
                continue

            committee_id = committee_map.get(row["committee"])
            if not committee_id:
                validation_errors.append(
                    {
                        "fieldName": f"행 {row_num}",
                        "message": f"위원회를 찾을 수 없습니다: {row['committee']}",
                    }
                )
                continue

            department_id = department_map.get(f"{committee_id}|{row['department']}")
            if not department_id:
                validation_errors.append(
                    {
                        "fieldName": f"행 {row_num}",
                        "message": f"사역팀을 찾을 수 없습니다: {row['committee']} - {row['department']}",
                    }
                )
                continue

            user_id = None
            if row["leader"]:
                user_id = user_map.get(row["leader"])
                if not user_id:
                    validation_errors.append(
                        {
                            "fieldName": f"행 {row_num}",
                            "message": f"사용자를 찾을 수 없습니다: {row['leader']}",
                        }
                    )
                    continue

            to_update.append(
                {
                    "departmentId": department_id,
                    "userId": user_id,
                    "departmentName": row["department"],
                    "leaderName": row["leader"] or "(없음)",
                }
            )

        if validation_errors:
            return _upload_error(
                f"검증 오류 {len(validation_errors)}건", status_code=200, fields=validation_errors
            )

        summary = {
            "totalRows": sum(1 for r in rows if r["committee"] or r["department"]),
            "updated": len(to_update),
            "skipped": 0,
            "errors": 0,
        }

        dry_run = dryRun == "true"
        if dry_run:
            return _upload_success(
                "검증 완료 (미리보기)",
                {
                    "summary": summary,
                    "dryRun": True,
                    "year": yr,
                    "preview": [
                        {"department": u["departmentName"], "leader": u["leaderName"]}
                        for u in to_update[:10]
                    ],
                },
            )

        team_leader_role = (
            await session.execute(
                select(Role).where(Role.tenantId == tenant_id, Role.code == "team_leader")
            )
        ).scalars().first()

        for item in to_update:
            if item["userId"]:
                existing = (
                    await session.execute(
                        select(UserYearRole).where(
                            UserYearRole.userId == item["userId"],
                            UserYearRole.year == yr,
                            UserYearRole.departmentId == item["departmentId"],
                            UserYearRole.role == "team_leader",
                        )
                    )
                ).scalars().first()
                if existing:
                    existing.roleId = team_leader_role.id if team_leader_role else None
                    session.add(existing)
                else:
                    session.add(
                        UserYearRole(
                            tenantId=tenant_id,
                            userId=item["userId"],
                            year=yr,
                            role="team_leader",
                            roleId=team_leader_role.id if team_leader_role else None,
                            departmentId=item["departmentId"],
                        )
                    )
            else:
                await session.execute(
                    delete(UserYearRole).where(
                        UserYearRole.year == yr,
                        UserYearRole.departmentId == item["departmentId"],
                        UserYearRole.role == "team_leader",
                    )
                )

        await session.commit()

        return _upload_success(
            f"업로드 완료: {yr}년도 {summary['updated']}개 사역팀 팀장 설정",
            {"summary": summary, "dryRun": False, "year": yr},
        )
    except Exception as err:  # noqa: BLE001 — Next 원본과 동일하게 예상 밖 오류도 500 으로 응답
        return _upload_error(
            str(err) or "업로드 처리 중 오류 발생", error_type="SERVER_ERROR", status_code=500
        )


# ── 예산(항) ──────────────────────────────────────────────────────────
class CategoryCreate(BaseModel):
    name: str


@router.get("/budget-categories")
async def list_categories(
    includeInactive: bool = False,
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
) -> dict:
    extra = {} if includeInactive else {"isActive": True}
    categories = await _list(session, BudgetCategory, tenant_id, **extra)
    if not categories:
        return {"categories": []}

    counts = dict(
        (
            await session.execute(
                select(BudgetSubcategory.categoryId, func.count(BudgetSubcategory.id))
                .where(
                    BudgetSubcategory.tenantId == tenant_id,
                    BudgetSubcategory.categoryId.in_([c["id"] for c in categories]),
                )
                .group_by(BudgetSubcategory.categoryId)
            )
        ).all()
    )
    for c in categories:
        c["_count"] = {"subcategories": counts.get(c["id"], 0)}

    return {"categories": categories}


@router.post("/budget-categories", status_code=status.HTTP_201_CREATED)
async def create_category(
    body: CategoryCreate,
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
    _=Depends(require_permission(PERMISSIONS.BUDGET_MASTER_MANAGE)),
) -> dict:
    name = body.name.strip()
    if not name:
        raise HTTPException(400, "예산(항)명을 입력해주세요.")
    if await _exists_by_name(session, BudgetCategory, tenant_id, name):
        raise HTTPException(409, "이미 존재하는 예산(항)명입니다.")
    entity = BudgetCategory(
        tenantId=tenant_id,
        name=name,
        sortOrder=await _next_sort_order(session, BudgetCategory, tenant_id),
    )
    session.add(entity)
    await session.commit()
    await session.refresh(entity)
    return entity.model_dump()


class CategoryUpdate(BaseModel):
    name: str | None = None
    isActive: bool | None = None
    sortOrder: int | None = None


@router.patch("/budget-categories/{category_id}")
async def update_category(
    category_id: str,
    body: CategoryUpdate,
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
    _=Depends(require_permission(PERMISSIONS.BUDGET_MASTER_MANAGE)),
) -> dict:
    entity = await _get_or_404(
        session, BudgetCategory, category_id, tenant_id, "예산(항)을 찾을 수 없습니다."
    )
    fields = body.model_fields_set

    if "name" in fields:
        new_name = (body.name or "").strip()
        if body.name and new_name != entity.name:
            if await _exists_by_name(session, BudgetCategory, tenant_id, new_name):
                raise HTTPException(409, "이미 존재하는 예산(항)입니다.")
        entity.name = new_name
    if "isActive" in fields:
        entity.isActive = body.isActive
    if "sortOrder" in fields:
        entity.sortOrder = body.sortOrder

    await session.commit()
    await session.refresh(entity)
    return entity.model_dump()


# ── 예산(목) ──────────────────────────────────────────────────────────
class SubcategoryCreate(BaseModel):
    categoryId: str
    name: str


@router.get("/budget-subcategories")
async def list_subcategories(
    categoryId: str | None = None,
    includeInactive: bool = False,
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
) -> dict:
    stmt = select(BudgetSubcategory).where(BudgetSubcategory.tenantId == tenant_id)
    if categoryId:
        stmt = stmt.where(BudgetSubcategory.categoryId == categoryId)
    if not includeInactive:
        stmt = (
            stmt.join(BudgetCategory, BudgetCategory.id == BudgetSubcategory.categoryId)
            .where(BudgetSubcategory.isActive == True)  # noqa: E712
            .where(BudgetCategory.isActive == True)  # noqa: E712
        )
    stmt = stmt.order_by(BudgetSubcategory.sortOrder)
    rows = (await session.execute(stmt)).scalars().all()
    subcategories = [r.model_dump() for r in rows]
    if not subcategories:
        return {"subcategories": []}

    counts = dict(
        (
            await session.execute(
                select(BudgetDetail.subcategoryId, func.count(BudgetDetail.id))
                .where(
                    BudgetDetail.tenantId == tenant_id,
                    BudgetDetail.subcategoryId.in_([s["id"] for s in subcategories]),
                )
                .group_by(BudgetDetail.subcategoryId)
            )
        ).all()
    )
    for s in subcategories:
        s["_count"] = {"details": counts.get(s["id"], 0)}

    return {"subcategories": subcategories}


@router.post("/budget-subcategories", status_code=status.HTTP_201_CREATED)
async def create_subcategory(
    body: SubcategoryCreate,
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
    _=Depends(require_permission(PERMISSIONS.BUDGET_MASTER_MANAGE)),
) -> dict:
    name = body.name.strip()
    if not name:
        raise HTTPException(400, "예산(목)명을 입력해주세요.")
    if await _exists_by_name(
        session, BudgetSubcategory, tenant_id, name, categoryId=body.categoryId
    ):
        raise HTTPException(409, "해당 항에 이미 존재하는 목명입니다.")
    entity = BudgetSubcategory(
        tenantId=tenant_id,
        categoryId=body.categoryId,
        name=name,
        sortOrder=await _next_sort_order(session, BudgetSubcategory, tenant_id),
    )
    session.add(entity)
    await session.commit()
    await session.refresh(entity)
    return entity.model_dump()


class SubcategoryUpdate(BaseModel):
    name: str | None = None
    isActive: bool | None = None
    sortOrder: int | None = None
    categoryId: str | None = None


@router.patch("/budget-subcategories/{subcategory_id}")
async def update_subcategory(
    subcategory_id: str,
    body: SubcategoryUpdate,
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
    _=Depends(require_permission(PERMISSIONS.BUDGET_MASTER_MANAGE)),
) -> dict:
    entity = await _get_or_404(
        session, BudgetSubcategory, subcategory_id, tenant_id, "예산(목)을 찾을 수 없습니다."
    )
    fields = body.model_fields_set
    target_category_id = body.categoryId if "categoryId" in fields else entity.categoryId

    if "name" in fields:
        new_name = (body.name or "").strip()
        if body.name and new_name != entity.name:
            if await _exists_by_name(
                session, BudgetSubcategory, tenant_id, new_name, categoryId=target_category_id
            ):
                raise HTTPException(409, "같은 예산(항) 내에 이미 존재하는 예산(목)입니다.")
        entity.name = new_name
    if "isActive" in fields:
        entity.isActive = body.isActive
    if "sortOrder" in fields:
        entity.sortOrder = body.sortOrder
    if "categoryId" in fields:
        entity.categoryId = body.categoryId

    await session.commit()
    await session.refresh(entity)
    return entity.model_dump()


# ── 예산(세목) ────────────────────────────────────────────────────────
class DetailCreate(BaseModel):
    subcategoryId: str
    name: str
    accountCode: str | None = None
    description: str | None = None


@router.get("/budget-details")
async def list_details(
    subcategoryId: str | None = None,
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
) -> dict:
    extra = {"subcategoryId": subcategoryId} if subcategoryId else {}
    return {"details": await _list(session, BudgetDetail, tenant_id, **extra)}


@router.post("/budget-details", status_code=status.HTTP_201_CREATED)
async def create_detail(
    body: DetailCreate,
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
    _=Depends(require_permission(PERMISSIONS.BUDGET_MASTER_MANAGE)),
) -> dict:
    name = body.name.strip()
    if not name:
        raise HTTPException(400, "예산(세목)명을 입력해주세요.")
    if await _exists_by_name(
        session, BudgetDetail, tenant_id, name, subcategoryId=body.subcategoryId
    ):
        raise HTTPException(409, "해당 목에 이미 존재하는 세목명입니다.")
    entity = BudgetDetail(
        tenantId=tenant_id,
        subcategoryId=body.subcategoryId,
        name=name,
        accountCode=body.accountCode,
        description=body.description,
        sortOrder=await _next_sort_order(session, BudgetDetail, tenant_id),
    )
    session.add(entity)
    await session.commit()
    await session.refresh(entity)
    return entity.model_dump()


class DetailUpdate(BaseModel):
    name: str | None = None
    isActive: bool | None = None
    sortOrder: int | None = None
    subcategoryId: str | None = None
    accountCode: str | None = None
    description: str | None = None


@router.patch("/budget-details/{detail_id}")
async def update_detail(
    detail_id: str,
    body: DetailUpdate,
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
    _=Depends(require_permission(PERMISSIONS.BUDGET_MASTER_MANAGE)),
) -> dict:
    entity = await _get_or_404(
        session, BudgetDetail, detail_id, tenant_id, "예산(세목)을 찾을 수 없습니다."
    )
    fields = body.model_fields_set
    target_subcategory_id = body.subcategoryId if "subcategoryId" in fields else entity.subcategoryId

    if "name" in fields:
        new_name = (body.name or "").strip()
        if body.name and new_name != entity.name:
            if await _exists_by_name(
                session, BudgetDetail, tenant_id, new_name, subcategoryId=target_subcategory_id
            ):
                raise HTTPException(409, "같은 예산(목) 내에 이미 존재하는 예산(세목)입니다.")
        entity.name = new_name
    if "isActive" in fields:
        entity.isActive = body.isActive
    if "sortOrder" in fields:
        entity.sortOrder = body.sortOrder
    if "subcategoryId" in fields:
        entity.subcategoryId = body.subcategoryId
    if "accountCode" in fields:
        entity.accountCode = (body.accountCode or "").strip() or None
    if "description" in fields:
        entity.description = (body.description or "").strip() or None

    await session.commit()
    await session.refresh(entity)
    return entity.model_dump()


class DetailDescriptionUpdate(BaseModel):
    description: str | None = None


@router.patch("/budget-details/{detail_id}/description")
async def update_detail_description(
    detail_id: str,
    body: DetailDescriptionUpdate,
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
    _=Depends(require_permission(PERMISSIONS.BUDGET_MASTER_MANAGE)),
) -> dict:
    if "description" not in body.model_fields_set:
        raise HTTPException(400, "description 필드가 필요합니다.")

    entity = await _get_or_404(session, BudgetDetail, detail_id, tenant_id, "세목을 찾을 수 없습니다.")
    entity.description = (body.description or "").strip() or None

    await session.commit()
    await session.refresh(entity)

    return {
        "success": True,
        "budgetDetail": {"id": entity.id, "name": entity.name, "description": entity.description},
    }


# ── 예산 세목 연도별 설정 ─────────────────────────────────────────────
# (app/api/budget-details/year, year/auto-assign 이전)
class BudgetDetailYearSetting(BaseModel):
    budgetDetailId: str
    managerId: str | None = None
    budgetAmount: int | None = None


class BudgetDetailYearBulkUpdate(BaseModel):
    year: int | None = None
    settings: list[BudgetDetailYearSetting] | None = None


@router.get("/budget-details/year")
async def budget_details_year(
    year: int | None = None,
    includeInactive: bool = False,
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
) -> dict:
    yr = year or utcnow().year
    start = datetime(yr, 1, 1, tzinfo=timezone.utc)
    end = datetime(yr + 1, 1, 1, tzinfo=timezone.utc)

    used_rows = (
        await session.execute(
            select(
                ExpenseItem.budgetCategory,
                ExpenseItem.budgetSubcategory,
                ExpenseItem.budgetDetail,
                func.coalesce(func.sum(ExpenseItem.amount), 0),
            )
            .join(Expense, Expense.id == ExpenseItem.expenseId)
            .where(
                Expense.tenantId == tenant_id,
                Expense.status.in_(_APPROVED_STATUSES),
                Expense.requestDate >= start,
                Expense.requestDate < end,
            )
            .group_by(ExpenseItem.budgetCategory, ExpenseItem.budgetSubcategory, ExpenseItem.budgetDetail)
        )
    ).all()
    used_map = {f"{cat}|{sub}|{det}": total or 0 for cat, sub, det, total in used_rows}

    stmt = (
        select(BudgetDetail, BudgetSubcategory, BudgetCategory)
        .join(BudgetSubcategory, BudgetSubcategory.id == BudgetDetail.subcategoryId)
        .join(BudgetCategory, BudgetCategory.id == BudgetSubcategory.categoryId)
        .where(BudgetDetail.tenantId == tenant_id)
    )
    if not includeInactive:
        stmt = stmt.where(
            BudgetDetail.isActive == True,  # noqa: E712
            BudgetSubcategory.isActive == True,  # noqa: E712
            BudgetCategory.isActive == True,  # noqa: E712
        )
    stmt = stmt.order_by(
        BudgetCategory.sortOrder, BudgetSubcategory.sortOrder, BudgetDetail.sortOrder
    )
    rows = (await session.execute(stmt)).all()

    detail_ids = [d.id for d, _, _ in rows]
    year_settings: dict[str, tuple] = {}
    dept_links: dict[str, list[dict]] = {}
    if detail_ids:
        ys_rows = (
            await session.execute(
                select(BudgetDetailYear, User)
                .join(User, User.id == BudgetDetailYear.managerId, isouter=True)
                .where(
                    BudgetDetailYear.budgetDetailId.in_(detail_ids),
                    BudgetDetailYear.year == yr,
                )
            )
        ).all()
        for ys, manager in ys_rows:
            year_settings[ys.budgetDetailId] = (ys, manager)

        link_rows = (
            await session.execute(
                select(DepartmentBudgetDetail.budgetDetailId, Department, Committee)
                .join(Department, Department.id == DepartmentBudgetDetail.departmentId)
                .join(Committee, Committee.id == Department.committeeId)
                .where(
                    DepartmentBudgetDetail.budgetDetailId.in_(detail_ids),
                    DepartmentBudgetDetail.isActive == True,  # noqa: E712
                )
            )
        ).all()
        for did, dept, comm in link_rows:
            dept_links.setdefault(did, []).append(
                {"id": dept.id, "name": dept.name, "committee": comm.name}
            )

    result = []
    for detail, sub, cat in rows:
        ys, manager = year_settings.get(detail.id, (None, None))
        key = f"{cat.name}|{sub.name}|{detail.name}"
        result.append(
            {
                "id": detail.id,
                "name": detail.name,
                "accountCode": detail.accountCode,
                "description": detail.description,
                "isActive": detail.isActive,
                "category": cat.name,
                "categoryId": cat.id,
                "categoryIsActive": cat.isActive,
                "subcategory": sub.name,
                "subcategoryId": sub.id,
                "subcategoryIsActive": sub.isActive,
                "departments": dept_links.get(detail.id, []),
                "yearSetting": (
                    {
                        "id": ys.id,
                        "year": ys.year,
                        "managerId": ys.managerId,
                        "managerName": manager.username if manager else None,
                        "budgetAmount": ys.budgetAmount,
                        "usedAmount": used_map.get(key, 0),
                    }
                    if ys
                    else None
                ),
            }
        )

    return {"year": yr, "details": result, "total": len(result)}


@router.post("/budget-details/year")
async def budget_details_year_bulk_update(
    body: BudgetDetailYearBulkUpdate,
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
    _=Depends(require_permission(PERMISSIONS.BUDGET_MASTER_MANAGE)),
) -> dict:
    if not body.year or not body.settings:
        raise HTTPException(400, "필수 파라미터 누락")

    results = []
    for setting in body.settings:
        existing = (
            await session.execute(
                select(BudgetDetailYear).where(
                    BudgetDetailYear.budgetDetailId == setting.budgetDetailId,
                    BudgetDetailYear.year == body.year,
                )
            )
        ).scalars().first()
        if existing:
            existing.managerId = setting.managerId or None
            existing.budgetAmount = setting.budgetAmount if setting.budgetAmount is not None else 0
            session.add(existing)
            entity = existing
        else:
            entity = BudgetDetailYear(
                tenantId=tenant_id,
                budgetDetailId=setting.budgetDetailId,
                year=body.year,
                managerId=setting.managerId or None,
                budgetAmount=setting.budgetAmount if setting.budgetAmount is not None else 0,
                usedAmount=0,
                isActive=True,
            )
            session.add(entity)
        await session.flush()

        manager = await session.get(User, entity.managerId) if entity.managerId else None
        row = entity.model_dump()
        row["manager"] = {"id": manager.id, "username": manager.username} if manager else None
        results.append(row)

    await session.commit()

    return {"message": f"{len(results)}건 저장 완료", "results": results}


class AutoAssignRequest(BaseModel):
    year: int | None = None
    overwrite: bool = False


@router.post("/budget-details/year/auto-assign")
async def budget_details_year_auto_assign(
    body: AutoAssignRequest,
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
    _=Depends(require_permission(PERMISSIONS.BUDGET_MASTER_MANAGE)),
) -> dict:
    if not body.year:
        raise HTTPException(400, "연도를 지정해주세요")

    team_leaders = (
        await session.execute(
            select(UserYearRole, User)
            .join(User, User.id == UserYearRole.userId)
            .where(
                UserYearRole.tenantId == tenant_id,
                UserYearRole.year == body.year,
                UserYearRole.role == "team_leader",
                UserYearRole.departmentId.is_not(None),
            )
        )
    ).all()

    department_to_leader: dict[str, tuple[str, str]] = {}
    for tl, user in team_leaders:
        if tl.departmentId:
            department_to_leader[tl.departmentId] = (user.id, user.username)

    dept_details = (
        await session.execute(
            select(DepartmentBudgetDetail, Department, Committee, BudgetDetail)
            .join(Department, Department.id == DepartmentBudgetDetail.departmentId)
            .join(Committee, Committee.id == Department.committeeId)
            .join(BudgetDetail, BudgetDetail.id == DepartmentBudgetDetail.budgetDetailId)
            .where(
                DepartmentBudgetDetail.tenantId == tenant_id,
                DepartmentBudgetDetail.isActive == True,  # noqa: E712
            )
        )
    ).all()

    created = 0
    updated = 0
    skipped = 0
    results = []

    for dd, dept, comm, detail in dept_details:
        committee_dept = f"{comm.name}/{dept.name}"
        leader = department_to_leader.get(dd.departmentId)

        if not leader:
            skipped += 1
            results.append(
                {
                    "budgetDetailId": dd.budgetDetailId,
                    "budgetDetailName": detail.name,
                    "department": committee_dept,
                    "managerId": None,
                    "managerName": None,
                    "action": "skipped",
                }
            )
            continue

        leader_id, leader_name = leader

        existing = (
            await session.execute(
                select(BudgetDetailYear).where(
                    BudgetDetailYear.budgetDetailId == dd.budgetDetailId,
                    BudgetDetailYear.year == body.year,
                )
            )
        ).scalars().first()

        if existing and existing.managerId and not body.overwrite:
            skipped += 1
            results.append(
                {
                    "budgetDetailId": dd.budgetDetailId,
                    "budgetDetailName": detail.name,
                    "department": committee_dept,
                    "managerId": existing.managerId,
                    "managerName": None,
                    "action": "skipped",
                }
            )
            continue

        if existing:
            existing.managerId = leader_id
            session.add(existing)
            updated += 1
            action = "updated"
        else:
            session.add(
                BudgetDetailYear(
                    tenantId=tenant_id,
                    budgetDetailId=dd.budgetDetailId,
                    year=body.year,
                    managerId=leader_id,
                    budgetAmount=0,
                    usedAmount=0,
                    isActive=True,
                )
            )
            created += 1
            action = "created"

        results.append(
            {
                "budgetDetailId": dd.budgetDetailId,
                "budgetDetailName": detail.name,
                "department": committee_dept,
                "managerId": leader_id,
                "managerName": leader_name,
                "action": action,
            }
        )

    await session.commit()

    return {
        "message": "담당자 자동 설정 완료",
        "year": body.year,
        "summary": {
            "total": len(dept_details),
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "teamLeadersFound": len(team_leaders),
        },
        "results": results[:20],
    }
