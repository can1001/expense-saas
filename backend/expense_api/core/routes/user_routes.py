"""사용자 목록·생성·상세·수정·비활성화 라우터.
(app/api/users/route.ts, app/api/users/[id]/route.ts 이전)
"""

import io
import math
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import JSONResponse
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel
from sqlalchemy import func, or_, select, update
from sqlalchemy.ext.asyncio import AsyncSession

from expense_api.core.auth.permissions import ROLE_CODES, ROLE_NAMES, YEAR_ROLE_CODES, PERMISSIONS
from expense_api.core.db.session import get_session
from expense_api.core.dependencies.auth import CurrentUser, get_current_user
from expense_api.core.dependencies.authz import effective_permissions, require_permission
from expense_api.core.dependencies.tenant import require_tenant_id
from expense_api.core.excel import set_column_widths, style_header_row, workbook_to_xlsx_response
from expense_api.core.models.budget import Department
from expense_api.core.models.user import Membership, Role, User, UserSignature, UserYearRole
from expense_api.core.security.jwt import hash_password

router = APIRouter()

DEFAULT_PASSWORD = "chc2026"


def _membership_role(role_code: str) -> str:
    return "TENANT_ADMIN" if role_code == "admin" else "MEMBER"


async def _get_role_by_code(session: AsyncSession, tenant_id: str, code: str) -> Role | None:
    stmt = (
        select(Role)
        .where(Role.tenantId == tenant_id, Role.code == code, Role.isActive == True)  # noqa: E712
        .order_by(Role.sortOrder)
    )
    return (await session.execute(stmt)).scalars().first()


async def _resolve_role_on_write(
    session: AsyncSession, tenant_id: str, role: str | None, role_id: str | None
) -> tuple[str | None, str | None]:
    """route 단 role/roleId 해석. roleId 만 있고 role 이 없으면 roleId 를 code 로 취급해 조회한다.

    (Next lib/services/user-service.ts 의 동일 동작을 그대로 재현 — roleId 인자를
    getRoleByCode 에 넘기는 부분은 원본의 의도된 동작이며 여기서 고치지 않는다.)
    """
    resolved_role = role
    resolved_role_id = role_id
    if role_id and not role:
        role_ref = await _get_role_by_code(session, tenant_id, role_id)
        if role_ref:
            resolved_role = role_ref.code
            resolved_role_id = role_ref.id
    return resolved_role, resolved_role_id


# ── 목록 ──────────────────────────────────────────────────────────────
@router.get("/users")
async def list_users(
    page: int = 1,
    pageSize: int = 20,
    role: str | None = None,
    isActive: str | None = None,
    search: str | None = None,
    includeRoleRef: bool = False,
    includeYearRoles: bool = False,
    year: int | None = None,
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
) -> dict:
    target_year = year or datetime.now().year

    where = [User.tenantId == tenant_id]

    if role:
        if role == "admin":
            where.append(User.role == "admin")
        else:
            year_role_users = (
                await session.execute(
                    select(UserYearRole.userId).where(
                        UserYearRole.tenantId == tenant_id,
                        UserYearRole.role == role,
                        UserYearRole.year == target_year,
                    )
                )
            ).scalars().all()
            where.append(User.id.in_(list(year_role_users)))

    if isActive == "true":
        where.append(User.isActive == True)  # noqa: E712
    elif isActive == "false":
        where.append(User.isActive == False)  # noqa: E712

    if search:
        pattern = f"%{search}%"
        where.append(
            or_(
                User.userid.ilike(pattern),
                User.username.ilike(pattern),
                User.department.ilike(pattern),
            )
        )

    total = (
        await session.execute(select(func.count(User.id)).where(*where))
    ).scalar_one()

    skip = (page - 1) * pageSize
    stmt = (
        select(User)
        .where(*where)
        .order_by(User.role, User.username)
        .offset(skip)
        .limit(pageSize)
    )
    users = (await session.execute(stmt)).scalars().all()

    results = [u.model_dump() for u in users]

    if includeRoleRef:
        role_ids = {u.roleId for u in users if u.roleId}
        role_refs: dict[str, dict] = {}
        if role_ids:
            rows = (
                await session.execute(select(Role).where(Role.id.in_(role_ids)))
            ).scalars().all()
            role_refs = {r.id: r.model_dump() for r in rows}
        for u, out in zip(users, results):
            out["roleRef"] = role_refs.get(u.roleId) if u.roleId else None

    if includeYearRoles:
        user_ids = [u.id for u in users]
        year_roles_by_user: dict[str, list[dict]] = {uid: [] for uid in user_ids}
        if user_ids:
            rows = (
                await session.execute(
                    select(UserYearRole)
                    .where(UserYearRole.userId.in_(user_ids), UserYearRole.year == target_year)
                    .order_by(UserYearRole.role)
                )
            ).scalars().all()
            for yr in rows:
                year_roles_by_user.setdefault(yr.userId, []).append(yr.model_dump())
        for out in results:
            out["yearRoles"] = year_roles_by_user.get(out["id"], [])

    return {
        "users": results,
        "pagination": {
            "page": page,
            "pageSize": pageSize,
            "total": total,
            "totalPages": math.ceil(total / pageSize) if pageSize else 0,
        },
    }


# ── 생성 ──────────────────────────────────────────────────────────────
class UserCreate(BaseModel):
    userid: str | None = None
    username: str | None = None
    role: str | None = None
    roleId: str | None = None
    department: str | None = None
    password: str | None = None
    phoneNumber: str | None = None


@router.post("/users", status_code=201)
async def create_user(
    body: UserCreate,
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
    _=Depends(require_permission(PERMISSIONS.USER_MANAGE)),
) -> dict:
    if not body.userid or not body.username:
        raise HTTPException(400, "userid and username are required")

    existing_by_userid = (
        await session.execute(
            select(User).where(User.tenantId == tenant_id, User.userid == body.userid)
        )
    ).scalars().first()
    if existing_by_userid:
        raise HTTPException(409, "User with this userid already exists")

    existing_by_username = (
        await session.execute(
            select(User).where(
                User.tenantId == tenant_id,
                User.username == body.username,
                User.isActive == True,  # noqa: E712
            )
        )
    ).scalars().first()
    if existing_by_username:
        raise HTTPException(409, "User with this username already exists")

    if body.role and body.role not in ROLE_CODES:
        raise HTTPException(400, "Invalid role")

    resolved_role, resolved_role_id = await _resolve_role_on_write(
        session, tenant_id, body.role, body.roleId
    )
    resolved_role = resolved_role or "user"

    if resolved_role and not resolved_role_id:
        role_ref = await _get_role_by_code(session, tenant_id, resolved_role)
        if role_ref:
            resolved_role_id = role_ref.id

    plain_password = body.password or DEFAULT_PASSWORD
    user = User(
        tenantId=tenant_id,
        userid=body.userid,
        username=body.username,
        role=resolved_role,
        roleId=resolved_role_id,
        department=body.department,
        phoneNumber=body.phoneNumber,
        password=hash_password(plain_password),
        mustChangePassword=True,
    )
    session.add(user)
    await session.flush()

    session.add(
        Membership(
            userId=user.id,
            tenantId=tenant_id,
            role=_membership_role(resolved_role),
            isDefault=True,
        )
    )
    await session.commit()
    await session.refresh(user)
    return user.model_dump()


# ── 엑셀 업로드/템플릿 ────────────────────────────────────────────────
# (app/api/users/upload/route.ts 이전 — lib/api/response-handler.ts 미사용,
# {success, error:{type,message,fields}}/{success, message, data} 원형 그대로 유지)
_USER_UPLOAD_HEADERS = [
    "userid (아이디)", "username (이름)", "role (역할)", "department (부서)", "isActive (활성화)",
]
_USER_UPLOAD_WIDTHS = [25, 15, 15, 20, 12]
_USER_UPLOAD_GUIDE_ROWS = [
    ("userid (아이디)", "로그인 아이디 (필수, 중복불가)", "청연정혜종"),
    ("username (이름)", "표시 이름 (필수)", "정혜종"),
    ("role (역할)", "역할 (관리자/재정팀장/회계/팀장/행정간사/사용자)", "사용자"),
    ("department (부서)", "소속 부서 (선택)", "재정팀"),
    ("isActive (활성화)", "활성화 여부 (Y/N)", "Y"),
    ("", "", ""),
    ("※ 참고사항", "", ""),
    ("- 새 사용자", "기본 비밀번호 chc2026 으로 생성됩니다", ""),
    ("- 기존 사용자", "병합 모드에서 이름/역할/부서/활성화 상태가 업데이트됩니다", ""),
    ("- 비밀번호", "엑셀로 변경할 수 없습니다 (보안)", ""),
]
# Next 원본 ROLE_MAP — 한글/영문 라벨 → 역할 코드
_ROLE_UPLOAD_MAP: dict[str, str] = {name: code for code, name in ROLE_NAMES.items()}
_ROLE_UPLOAD_MAP.update({code: code for code in ROLE_CODES})
_ROLE_UPLOAD_MAP[""] = "user"


def _upload_error(
    message: str,
    *,
    error_type: str = "VALIDATION_ERROR",
    status_code: int = 400,
    fields: list[dict] | None = None,
) -> JSONResponse:
    error: dict = {"type": error_type, "message": message}
    if fields:
        error["fields"] = fields
    return JSONResponse({"success": False, "error": error}, status_code=status_code)


def _upload_success(message: str, data: dict) -> JSONResponse:
    return JSONResponse({"success": True, "message": message, "data": data})


@router.get("/users/upload")
async def users_upload_template(
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
    _=Depends(require_permission(PERMISSIONS.USER_MANAGE)),
):
    users = (
        await session.execute(
            select(User).where(User.tenantId == tenant_id).order_by(User.role, User.username)
        )
    ).scalars().all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "사용자목록"
    sheet.append(_USER_UPLOAD_HEADERS)
    style_header_row(sheet)
    set_column_widths(sheet, _USER_UPLOAD_WIDTHS)

    if not users:
        sheet.append(["", "", "사용자", "", "Y"])
    else:
        for u in users:
            sheet.append(
                [u.userid, u.username, ROLE_NAMES.get(u.role, u.role), u.department or "",
                 "Y" if u.isActive else "N"]
            )

    guide = workbook.create_sheet("작성안내")
    guide.append(["항목", "설명", "예시"])
    style_header_row(guide)
    set_column_widths(guide, [20, 50, 20])
    for row in _USER_UPLOAD_GUIDE_ROWS:
        guide.append(list(row))

    date_str = datetime.now(timezone.utc).date().isoformat()
    return workbook_to_xlsx_response(workbook, f"users_template_{date_str}.xlsx")


@router.post("/users/upload")
async def users_upload_post(
    file: UploadFile | None = File(None),
    mode: str = Form("merge"),
    dryRun: str = Form("false"),
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
    _=Depends(require_permission(PERMISSIONS.USER_MANAGE)),
):
    if file is None:
        return _upload_error("파일이 필요합니다.")

    try:
        data = await file.read()
        workbook = load_workbook(io.BytesIO(data), data_only=True)

        if not workbook.worksheets:
            return _upload_error("워크시트를 찾을 수 없습니다.")
        sheet = workbook.worksheets[0]

        headers: dict[int, str] = {}
        for cell in sheet[1]:
            headers[cell.column] = str(cell.value or "").lower()

        raw_rows: list[dict[str, object]] = []
        for row in sheet.iter_rows(min_row=2):
            row_data: dict[str, object] = {}
            for cell in row:
                header = headers.get(cell.column)
                if header:
                    row_data[header] = cell.value
            if any(v is not None and v != "" for v in row_data.values()):
                raw_rows.append(row_data)

        if not raw_rows:
            return _upload_error("데이터가 없습니다.")

        def _norm(value: object) -> str:
            return str(value).strip() if value is not None else ""

        rows: list[dict] = []
        for raw in raw_rows:
            normalized: dict = {
                "userid": "", "username": "", "role": None, "department": None, "isActive": None,
            }
            for key, value in raw.items():
                lowered = key.lower()
                if "userid" in lowered or "아이디" in lowered:
                    normalized["userid"] = _norm(value)
                elif "username" in lowered or "이름" in lowered:
                    normalized["username"] = _norm(value)
                elif "role" in lowered or "역할" in lowered:
                    normalized["role"] = _norm(value)
                elif "department" in lowered or "부서" in lowered:
                    normalized["department"] = _norm(value) or None
                elif "isactive" in lowered or "활성" in lowered:
                    normalized["isActive"] = _norm(value).upper() in ("Y", "TRUE", "1", "YES")
            rows.append(normalized)

        validation_errors: list[dict] = []
        valid_rows: list[dict] = []
        for index, row in enumerate(rows):
            row_num = index + 2
            if not row["userid"]:
                validation_errors.append(
                    {"fieldName": f"행 {row_num}", "message": "userid(아이디)가 비어있습니다."}
                )
                continue
            if not row["username"]:
                validation_errors.append(
                    {"fieldName": f"행 {row_num}", "message": "username(이름)이 비어있습니다."}
                )
                continue
            if row["role"]:
                mapped = _ROLE_UPLOAD_MAP.get(row["role"])
                if not mapped:
                    validation_errors.append(
                        {"fieldName": f"행 {row_num}", "message": f"유효하지 않은 역할: {row['role']}"}
                    )
                    continue
                row["role"] = mapped
            else:
                row["role"] = "user"
            if row["isActive"] is None:
                row["isActive"] = True
            valid_rows.append(row)

        userid_counts: dict[str, int] = {}
        for row in valid_rows:
            userid_counts[row["userid"]] = userid_counts.get(row["userid"], 0) + 1
        for userid, count in userid_counts.items():
            if count > 1:
                validation_errors.append(
                    {"fieldName": userid, "message": f"파일 내 중복된 userid: {userid} ({count}회)"}
                )

        if validation_errors:
            return _upload_error(
                f"검증 오류 {len(validation_errors)}건", status_code=200, fields=validation_errors
            )

        existing_users = (
            await session.execute(select(User.id, User.userid).where(User.tenantId == tenant_id))
        ).all()
        existing_map = {userid: uid for uid, userid in existing_users}

        mode_value = mode or "merge"
        to_create: list[dict] = []
        to_update: list[dict] = []
        for row in valid_rows:
            existing_id = existing_map.get(row["userid"])
            if existing_id:
                if mode_value == "merge":
                    to_update.append({**row, "existingId": existing_id})
            else:
                to_create.append(row)

        summary = {
            "totalRows": len(valid_rows),
            "created": len(to_create),
            "updated": len(to_update),
            "skipped": len(valid_rows) - len(to_create) - len(to_update),
            "errors": 0,
        }

        dry_run = dryRun == "true"
        if dry_run:
            return _upload_success(
                "검증 완료 (미리보기)", {"summary": summary, "dryRun": True, "mode": mode_value}
            )

        hashed_password = hash_password(DEFAULT_PASSWORD)

        if to_create:
            role_rows = (
                await session.execute(
                    select(Role.code, Role.id).where(
                        Role.tenantId == tenant_id, Role.isActive == True  # noqa: E712
                    )
                )
            ).all()
            roles_by_code = dict(role_rows)

            for row in to_create:
                user = User(
                    tenantId=tenant_id,
                    userid=row["userid"],
                    username=row["username"],
                    role=row["role"],
                    roleId=roles_by_code.get(row["role"]),
                    department=row["department"],
                    isActive=bool(row["isActive"]),
                    password=hashed_password,
                    mustChangePassword=True,
                )
                session.add(user)
                await session.flush()
                session.add(
                    Membership(
                        userId=user.id,
                        tenantId=tenant_id,
                        role=_membership_role(row["role"]),
                        isDefault=True,
                    )
                )

        for row in to_update:
            existing_user = await session.get(User, row["existingId"])
            if existing_user is not None:
                existing_user.username = row["username"]
                existing_user.role = row["role"]
                existing_user.department = row["department"]
                existing_user.isActive = bool(row["isActive"])
                session.add(existing_user)

        await session.commit()

        return _upload_success(
            f"업로드 완료: {summary['created']}명 생성, {summary['updated']}명 업데이트",
            {"summary": summary, "dryRun": False, "mode": mode_value},
        )
    except Exception as err:  # noqa: BLE001 — Next 원본과 동일하게 예상 밖 오류도 500 으로 응답
        return _upload_error(
            str(err) or "업로드 처리 중 오류 발생", error_type="SERVER_ERROR", status_code=500
        )


# ── 역할별 목록 ───────────────────────────────────────────────────────
# (app/api/users/by-role/[role]/route.ts 이전 — /users/{user_id} 보다 먼저 등록해
# FastAPI 라우팅이 "by-role" 을 user_id 로 잘못 매칭하지 않도록 한다.)
@router.get("/users/by-role/{role}")
async def list_users_by_role(
    role: str,
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
) -> dict:
    if role not in ROLE_CODES:
        raise HTTPException(
            400,
            "Invalid role. Valid roles: admin, finance_head, accountant, finance_member, "
            "team_leader, admin_assistant, user",
        )

    stmt = (
        select(User)
        .where(User.tenantId == tenant_id, User.role == role, User.isActive == True)  # noqa: E712
        .order_by(User.username)
    )
    users = (await session.execute(stmt)).scalars().all()
    return {"users": [u.model_dump() for u in users]}


# ── 간편 등록 ─────────────────────────────────────────────────────────
# (app/api/users/quick-register/route.ts 이전)
USERID_PREFIX = "청연"


class QuickRegisterBody(BaseModel):
    name: str | None = None


@router.post("/users/quick-register")
async def quick_register_user(
    body: QuickRegisterBody,
    user: CurrentUser = Depends(get_current_user),
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
) -> dict:
    perms = await effective_permissions(user, session)
    if PERMISSIONS.USER_REGISTER not in perms:
        raise HTTPException(403, "사용자 등록 권한이 없습니다.")

    name = (body.name or "").strip()
    if not name:
        raise HTTPException(400, "이름을 입력해주세요.")

    userid = f"{USERID_PREFIX}{name}"

    existing = (
        await session.execute(
            select(User).where(User.tenantId == tenant_id, User.userid == userid)
        )
    ).scalars().first()
    if existing:
        raise HTTPException(409, f"이미 등록된 사용자입니다: {userid}")

    new_user = User(
        tenantId=tenant_id,
        userid=userid,
        username=name,
        role="user",
        password=hash_password(DEFAULT_PASSWORD),
        mustChangePassword=True,
    )
    session.add(new_user)
    await session.flush()

    session.add(
        Membership(
            userId=new_user.id,
            tenantId=tenant_id,
            role=_membership_role("user"),
            isDefault=True,
        )
    )
    await session.commit()
    await session.refresh(new_user)

    return {
        "success": True,
        "user": {
            "id": new_user.id,
            "userid": new_user.userid,
            "username": new_user.username,
            "role": new_user.role,
        },
        "message": f"사용자가 등록되었습니다.\n아이디: {userid}\n기본 비밀번호: {DEFAULT_PASSWORD}",
    }


# ── 연도별 역할 ───────────────────────────────────────────────────────
# (app/api/users/year-roles/route.ts 이전)
class YearRoleSet(BaseModel):
    userId: str | None = None
    userid: str | None = None
    year: int | None = None
    role: str | None = None
    departmentId: str | None = None


class YearRoleDelete(BaseModel):
    userId: str | None = None
    year: int | None = None
    departmentId: str | None = None


@router.get("/users/year-roles")
async def list_year_roles(
    year: int | None = None,
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
) -> dict:
    target_year = year or datetime.now().year

    rows = (
        await session.execute(
            select(UserYearRole).where(
                UserYearRole.tenantId == tenant_id, UserYearRole.year == target_year
            )
        )
    ).scalars().all()

    user_ids = {r.userId for r in rows}
    users_by_id: dict[str, User] = {}
    if user_ids:
        users = (
            await session.execute(select(User).where(User.id.in_(user_ids)))
        ).scalars().all()
        users_by_id = {u.id: u for u in users}

    dept_ids = {r.departmentId for r in rows if r.departmentId}
    depts_by_id: dict[str, Department] = {}
    if dept_ids:
        depts = (
            await session.execute(select(Department).where(Department.id.in_(dept_ids)))
        ).scalars().all()
        depts_by_id = {d.id: d for d in depts}

    rows_sorted = sorted(
        rows,
        key=lambda r: (r.role, users_by_id[r.userId].username if r.userId in users_by_id else ""),
    )

    year_roles = []
    for yr in rows_sorted:
        u = users_by_id.get(yr.userId)
        year_roles.append(
            {
                "id": yr.id,
                "userId": yr.userId,
                "year": yr.year,
                "role": yr.role,
                "departmentId": yr.departmentId,
                "department": depts_by_id[yr.departmentId].name if yr.departmentId in depts_by_id else None,
                "user": {"id": u.id, "username": u.username} if u else None,
            }
        )

    return {"year": target_year, "yearRoles": year_roles}


@router.post("/users/year-roles", status_code=201)
async def set_year_role(
    body: YearRoleSet,
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
    _=Depends(require_permission(PERMISSIONS.USER_MANAGE)),
) -> dict:
    target_user_id = body.userId
    if not target_user_id and body.userid:
        found = (
            await session.execute(
                select(User).where(User.tenantId == tenant_id, User.userid == body.userid)
            )
        ).scalars().first()
        if not found:
            raise HTTPException(404, "User not found")
        target_user_id = found.id

    if not target_user_id or not body.year or not body.role:
        raise HTTPException(400, "userId (or userid), year, and role are required")

    target_user = await session.get(User, target_user_id)
    if target_user is None or target_user.tenantId != tenant_id:
        raise HTTPException(404, "User not found")

    if body.role not in YEAR_ROLE_CODES:
        raise HTTPException(
            400,
            "Invalid role for year role. Valid roles: finance_head, accountant, "
            "finance_member, team_leader, admin_assistant",
        )

    role_ref = await _get_role_by_code(session, tenant_id, body.role)
    role_id = role_ref.id if role_ref else None

    if not body.departmentId:
        existing_yr = (
            await session.execute(
                select(UserYearRole).where(
                    UserYearRole.userId == target_user_id,
                    UserYearRole.year == body.year,
                    UserYearRole.departmentId.is_(None),
                )
            )
        ).scalars().first()
        if existing_yr:
            existing_yr.role = body.role
            existing_yr.roleId = role_id
            year_role = existing_yr
        else:
            year_role = UserYearRole(
                tenantId=tenant_id,
                userId=target_user_id,
                year=body.year,
                role=body.role,
                roleId=role_id,
                departmentId=None,
            )
            session.add(year_role)
    else:
        existing_yr = (
            await session.execute(
                select(UserYearRole).where(
                    UserYearRole.userId == target_user_id,
                    UserYearRole.year == body.year,
                    UserYearRole.departmentId == body.departmentId,
                    UserYearRole.role == body.role,
                )
            )
        ).scalars().first()
        if existing_yr:
            existing_yr.roleId = role_id
            year_role = existing_yr
        else:
            year_role = UserYearRole(
                tenantId=tenant_id,
                userId=target_user_id,
                year=body.year,
                role=body.role,
                roleId=role_id,
                departmentId=body.departmentId,
            )
            session.add(year_role)

    await session.commit()
    await session.refresh(year_role)
    return year_role.model_dump()


@router.delete("/users/year-roles")
async def delete_year_role(
    body: YearRoleDelete,
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
    _=Depends(require_permission(PERMISSIONS.USER_MANAGE)),
) -> dict:
    if not body.userId or not body.year:
        raise HTTPException(400, "userId and year are required")

    where = [
        UserYearRole.tenantId == tenant_id,
        UserYearRole.userId == body.userId,
        UserYearRole.year == body.year,
    ]
    if body.departmentId:
        where.append(UserYearRole.departmentId == body.departmentId)

    rows = (await session.execute(select(UserYearRole).where(*where))).scalars().all()
    for row in rows:
        await session.delete(row)
    await session.commit()

    return {"success": True}


# ── 상세 ──────────────────────────────────────────────────────────────
@router.get("/users/{user_id}")
async def get_user(
    user_id: str,
    includeRoleRef: bool = False,
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
) -> dict:
    user = await session.get(User, user_id)
    if user is None or user.tenantId != tenant_id:
        raise HTTPException(404, "User not found")

    result = user.model_dump()
    if includeRoleRef:
        role_ref = await session.get(Role, user.roleId) if user.roleId else None
        result["roleRef"] = role_ref.model_dump() if role_ref else None
    return result


# ── 수정 ──────────────────────────────────────────────────────────────
class UserUpdate(BaseModel):
    username: str | None = None
    role: str | None = None
    roleId: str | None = None
    department: str | None = None
    password: str | None = None
    phoneNumber: str | None = None
    isActive: bool | None = None
    canRegisterUsers: bool | None = None


@router.put("/users/{user_id}")
async def update_user(
    user_id: str,
    body: UserUpdate,
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
    _=Depends(require_permission(PERMISSIONS.USER_MANAGE)),
) -> dict:
    user = await session.get(User, user_id)
    if user is None or user.tenantId != tenant_id:
        raise HTTPException(404, "User not found")

    fields = body.model_fields_set

    if "role" in fields and body.role and body.role not in ROLE_CODES:
        raise HTTPException(400, "Invalid role")

    resolved_role, resolved_role_id = await _resolve_role_on_write(
        session, tenant_id, body.role if "role" in fields else None, body.roleId if "roleId" in fields else None
    )

    if resolved_role and not resolved_role_id:
        role_ref = await _get_role_by_code(session, tenant_id, resolved_role)
        if role_ref:
            resolved_role_id = role_ref.id

    if "username" in fields:
        user.username = body.username
    if resolved_role:
        user.role = resolved_role
    if resolved_role_id:
        user.roleId = resolved_role_id
    if "department" in fields:
        user.department = body.department
    if "password" in fields and body.password:
        user.password = hash_password(body.password)
    if "phoneNumber" in fields:
        user.phoneNumber = body.phoneNumber
    if "isActive" in fields:
        user.isActive = body.isActive
    if "canRegisterUsers" in fields:
        user.canRegisterUsers = body.canRegisterUsers

    await session.commit()
    await session.refresh(user)
    return user.model_dump()


# ── 비활성화 (soft delete) ───────────────────────────────────────────
@router.delete("/users/{user_id}")
async def deactivate_user(
    user_id: str,
    tenant_id: str = Depends(require_tenant_id),
    session: AsyncSession = Depends(get_session),
    _=Depends(require_permission(PERMISSIONS.USER_MANAGE)),
) -> dict:
    user = await session.get(User, user_id)
    if user is None or user.tenantId != tenant_id:
        raise HTTPException(404, "User not found")

    user.isActive = False
    await session.commit()
    await session.refresh(user)

    return {"message": "User deactivated successfully", "user": user.model_dump()}


# ── 서명/도장 ──────────────────────────────────────────────────────────
# (app/api/users/me/signatures/**, [id]/default 이전 — 본인 것만 조작 가능)
SIGNATURE_TYPES = {"signature", "stamp"}


class SignatureCreate(BaseModel):
    type: str | None = None
    name: str | None = None
    imageData: str | None = None
    isDefault: bool | None = None


class SignatureUpdate(BaseModel):
    name: str | None = None
    imageData: str | None = None
    isDefault: bool | None = None


def _signature_out(sig: UserSignature) -> dict:
    return {
        "id": sig.id,
        "type": sig.type,
        "name": sig.name,
        "imageData": sig.imageData,
        "isDefault": sig.isDefault,
        "createdAt": sig.createdAt,
    }


async def _get_my_signature(
    session: AsyncSession, user_id: str, signature_id: str
) -> UserSignature | None:
    return (
        await session.execute(
            select(UserSignature).where(
                UserSignature.id == signature_id, UserSignature.userId == user_id
            )
        )
    ).scalars().first()


@router.get("/users/me/signatures")
async def list_my_signatures(
    user: CurrentUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> dict:
    stmt = (
        select(UserSignature)
        .where(UserSignature.userId == user.id)
        .order_by(
            UserSignature.type,
            UserSignature.isDefault.desc(),
            UserSignature.createdAt.desc(),
        )
    )
    rows = (await session.execute(stmt)).scalars().all()
    return {"signatures": [_signature_out(s) for s in rows]}


@router.post("/users/me/signatures")
async def create_my_signature(
    body: SignatureCreate,
    user: CurrentUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> dict:
    if not body.type or body.type not in SIGNATURE_TYPES:
        raise HTTPException(400, "유효하지 않은 타입입니다. (signature 또는 stamp)")
    if not body.name or not body.name.strip():
        raise HTTPException(400, "이름을 입력해주세요.")
    if not body.imageData or not body.imageData.startswith("data:image/"):
        raise HTTPException(400, "유효하지 않은 이미지 데이터입니다.")

    if body.isDefault:
        await session.execute(
            update(UserSignature)
            .where(
                UserSignature.userId == user.id,
                UserSignature.type == body.type,
                UserSignature.isDefault == True,  # noqa: E712
            )
            .values(isDefault=False)
        )

    signature = UserSignature(
        userId=user.id,
        type=body.type,
        name=body.name.strip(),
        imageData=body.imageData,
        isDefault=bool(body.isDefault),
    )
    session.add(signature)
    await session.commit()
    await session.refresh(signature)

    return {
        "success": True,
        "message": f"{'서명' if body.type == 'signature' else '도장'}이 등록되었습니다.",
        "signature": _signature_out(signature),
    }


@router.get("/users/me/signatures/{signature_id}")
async def get_my_signature(
    signature_id: str,
    user: CurrentUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> dict:
    signature = await _get_my_signature(session, user.id, signature_id)
    if signature is None:
        raise HTTPException(404, "서명/도장을 찾을 수 없습니다.")
    return {"signature": signature.model_dump()}


@router.put("/users/me/signatures/{signature_id}")
async def update_my_signature(
    signature_id: str,
    body: SignatureUpdate,
    user: CurrentUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> dict:
    existing = await _get_my_signature(session, user.id, signature_id)
    if existing is None:
        raise HTTPException(404, "서명/도장을 찾을 수 없습니다.")

    fields = body.model_fields_set

    if "imageData" in fields and body.imageData and not body.imageData.startswith("data:image/"):
        raise HTTPException(400, "유효하지 않은 이미지 데이터입니다.")

    if "isDefault" in fields and body.isDefault is True:
        await session.execute(
            update(UserSignature)
            .where(
                UserSignature.userId == user.id,
                UserSignature.type == existing.type,
                UserSignature.isDefault == True,  # noqa: E712
                UserSignature.id != signature_id,
            )
            .values(isDefault=False)
        )

    if "name" in fields and body.name is not None:
        existing.name = body.name.strip()
    if "imageData" in fields and body.imageData is not None:
        existing.imageData = body.imageData
    if "isDefault" in fields and body.isDefault is not None:
        existing.isDefault = body.isDefault

    await session.commit()
    await session.refresh(existing)

    return {
        "success": True,
        "message": "수정되었습니다.",
        "signature": _signature_out(existing),
    }


@router.delete("/users/me/signatures/{signature_id}")
async def delete_my_signature(
    signature_id: str,
    user: CurrentUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> dict:
    existing = await _get_my_signature(session, user.id, signature_id)
    if existing is None:
        raise HTTPException(404, "서명/도장을 찾을 수 없습니다.")

    await session.delete(existing)
    await session.commit()

    return {
        "success": True,
        "message": f"{'서명' if existing.type == 'signature' else '도장'}이 삭제되었습니다.",
    }


@router.put("/users/me/signatures/{signature_id}/default")
async def set_default_signature(
    signature_id: str,
    user: CurrentUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
) -> dict:
    target = await _get_my_signature(session, user.id, signature_id)
    if target is None:
        raise HTTPException(404, "서명/도장을 찾을 수 없습니다.")

    await session.execute(
        update(UserSignature)
        .where(
            UserSignature.userId == user.id,
            UserSignature.type == target.type,
            UserSignature.isDefault == True,  # noqa: E712
        )
        .values(isDefault=False)
    )
    target.isDefault = True
    await session.commit()

    return {
        "success": True,
        "message": f"기본 {'서명' if target.type == 'signature' else '도장'}으로 설정되었습니다.",
    }
