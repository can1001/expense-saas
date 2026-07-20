"""예산 마스터 데이터 업로드/템플릿 내보내기 (lib/budget-upload.ts 712L 포팅).

정규화 테이블(BudgetCategory/Subcategory/Detail/DetailYear/DepartmentBudgetDetail) 대상.
행 단위 오류는 격리(해당 행만 저장점 롤백)하고 다음 행을 계속 처리한다 — Next 원본의
"행별 try/catch로 개별 실패를 흡수하고 트랜잭션 전체는 계속 진행" 동작과 동일한 결과.
"""

import io
import math
from dataclasses import dataclass, field
from typing import Literal

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession

from expense_api.core.excel import set_column_widths, style_header_row
from expense_api.core.models.budget import (
    BudgetCategory,
    BudgetDetail,
    BudgetDetailYear,
    Committee,
    Department,
    DepartmentBudgetDetail,
    BudgetSubcategory,
)
from expense_api.core.models.user import User
from expense_api.core.security.jwt import hash_password

UploadMode = Literal["replace", "merge", "append"]

_COLUMN_INDEX = {
    "committee": 1,
    "department": 2,
    "category": 3,
    "subcategory": 4,
    "detail": 5,
    "manager": 6,
    "accountCode": 7,
    "description": 8,
    "isActive": 9,
    "year": 10,
    "budgetAmount": 11,
}

_TEMPLATE_HEADERS = [
    "위원회", "사역팀(부)", "예산(항)", "예산(목)", "예산(세목)",
    "담당자", "계정코드", "항목 내역", "활성화", "연도", "예산금액",
]
_TEMPLATE_WIDTHS = [20, 20, 20, 20, 25, 15, 15, 40, 10, 10, 15]

_DEFAULT_MANAGER_PASSWORD = "chc2026"


@dataclass
class BudgetRow:
    committee: str
    department: str
    category: str
    subcategory: str
    detail: str
    manager: str | None = None
    accountCode: str | None = None
    description: str | None = None
    isActive: bool = True
    year: int | None = None
    budgetAmount: int | None = None


@dataclass
class ValidationError:
    row: int
    field: str
    message: str


@dataclass
class UploadSummary:
    totalRows: int
    created: int = 0
    updated: int = 0
    skipped: int = 0
    errors: int = 0


@dataclass
class UploadResult:
    success: bool
    summary: UploadSummary
    validationErrors: list[ValidationError] = field(default_factory=list)


def _as_str(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        return value.strip() or None
    return str(value).strip() or None


def _as_bool(value: object) -> bool:
    if value is None:
        return True
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    lowered = str(value).strip().lower()
    return lowered not in ("false", "0", "no", "n")


def _as_int(value: object) -> int | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return math.floor(value)
    text = str(value).strip().replace(",", "")
    if text == "":
        return None
    try:
        return int(text)
    except ValueError:
        try:
            return math.floor(float(text))
        except ValueError:
            return None


def parse_excel_file(
    data: bytes, data_start_row: int = 2
) -> tuple[list[BudgetRow], list[ValidationError]]:
    """Excel 파일 파싱 — 위원회/사역팀/항/목/세목(필수) + 담당자/계정코드/내역/활성화/연도/예산금액(선택)."""
    workbook = load_workbook(io.BytesIO(data), data_only=True)
    if not workbook.worksheets:
        raise ValueError("Excel 파일에 시트가 없습니다.")
    sheet = workbook.worksheets[0]

    rows: list[BudgetRow] = []
    errors: list[ValidationError] = []

    for row_num in range(data_start_row, sheet.max_row + 1):
        row_cells = sheet[row_num]

        def cell(col_key: str):
            idx = _COLUMN_INDEX[col_key] - 1
            return row_cells[idx].value if idx < len(row_cells) else None

        committee = _as_str(cell("committee"))
        if not committee:
            continue  # 빈 행 건너뛰기

        department = _as_str(cell("department"))
        category = _as_str(cell("category"))
        subcategory = _as_str(cell("subcategory"))
        detail = _as_str(cell("detail"))

        row_errors: list[ValidationError] = []
        if not department:
            row_errors.append(ValidationError(row_num, "department", "사역팀(부)은 필수입니다."))
        if not category:
            row_errors.append(ValidationError(row_num, "category", "예산(항)은 필수입니다."))
        if not subcategory:
            row_errors.append(ValidationError(row_num, "subcategory", "예산(목)은 필수입니다."))
        if not detail:
            row_errors.append(ValidationError(row_num, "detail", "예산(세목)은 필수입니다."))

        if row_errors:
            errors.extend(row_errors)
            continue

        rows.append(
            BudgetRow(
                committee=committee,
                department=department,  # type: ignore[arg-type]
                category=category,  # type: ignore[arg-type]
                subcategory=subcategory,  # type: ignore[arg-type]
                detail=detail,  # type: ignore[arg-type]
                manager=_as_str(cell("manager")),
                accountCode=_as_str(cell("accountCode")),
                description=_as_str(cell("description")),
                isActive=_as_bool(cell("isActive")),
                year=_as_int(cell("year")),
                budgetAmount=_as_int(cell("budgetAmount")),
            )
        )

    return rows, errors


async def _dry_run(
    session: AsyncSession, tenant_id: str, rows: list[BudgetRow], mode: UploadMode
) -> UploadResult:
    summary = UploadSummary(totalRows=len(rows))
    for row in rows:
        existing = (
            await session.execute(
                select(BudgetDetail.id)
                .join(BudgetSubcategory, BudgetSubcategory.id == BudgetDetail.subcategoryId)
                .join(BudgetCategory, BudgetCategory.id == BudgetSubcategory.categoryId)
                .where(
                    BudgetDetail.tenantId == tenant_id,
                    BudgetDetail.name == row.detail,
                    BudgetSubcategory.name == row.subcategory,
                    BudgetCategory.name == row.category,
                )
            )
        ).first()
        if existing:
            if mode == "append":
                summary.skipped += 1
            else:
                summary.updated += 1
        else:
            summary.created += 1

    return UploadResult(success=True, summary=summary, validationErrors=[])


async def _upload_row(
    session: AsyncSession,
    tenant_id: str,
    row: BudgetRow,
    mode: UploadMode,
    summary: UploadSummary,
    committee_cache: dict[str, str],
    department_cache: dict[str, str],
    category_cache: dict[str, str],
    subcategory_cache: dict[str, str],
    user_cache: dict[str, str],
    default_password_hash: str,
) -> None:
    # 1. Committee find or create
    committee_id = committee_cache.get(row.committee)
    if committee_id is None:
        committee = (
            await session.execute(
                select(Committee).where(
                    Committee.tenantId == tenant_id, Committee.name == row.committee
                )
            )
        ).scalars().first()
        if committee is None:
            committee = Committee(tenantId=tenant_id, name=row.committee, isActive=True)
            session.add(committee)
            await session.flush()
        committee_id = committee.id
        committee_cache[row.committee] = committee_id

    # 2. Department upsert
    dept_key = f"{row.committee}|{row.department}"
    department_id = department_cache.get(dept_key)
    if department_id is None:
        department = (
            await session.execute(
                select(Department).where(
                    Department.committeeId == committee_id, Department.name == row.department
                )
            )
        ).scalars().first()
        if department is None:
            department = Department(
                tenantId=tenant_id, committeeId=committee_id, name=row.department, isActive=True
            )
            session.add(department)
            await session.flush()
        department_id = department.id
        department_cache[dept_key] = department_id

    # 3. BudgetCategory find or create
    category_id = category_cache.get(row.category)
    if category_id is None:
        category = (
            await session.execute(
                select(BudgetCategory).where(
                    BudgetCategory.tenantId == tenant_id, BudgetCategory.name == row.category
                )
            )
        ).scalars().first()
        if category is None:
            category = BudgetCategory(tenantId=tenant_id, name=row.category, isActive=True)
            session.add(category)
            await session.flush()
        category_id = category.id
        category_cache[row.category] = category_id

    # 4. BudgetSubcategory upsert
    subcat_key = f"{row.category}|{row.subcategory}"
    subcategory_id = subcategory_cache.get(subcat_key)
    if subcategory_id is None:
        subcategory = (
            await session.execute(
                select(BudgetSubcategory).where(
                    BudgetSubcategory.categoryId == category_id,
                    BudgetSubcategory.name == row.subcategory,
                )
            )
        ).scalars().first()
        if subcategory is None:
            subcategory = BudgetSubcategory(
                tenantId=tenant_id, categoryId=category_id, name=row.subcategory, isActive=True
            )
            session.add(subcategory)
            await session.flush()
        subcategory_id = subcategory.id
        subcategory_cache[subcat_key] = subcategory_id

    # 5. BudgetDetail upsert
    existing_detail = (
        await session.execute(
            select(BudgetDetail).where(
                BudgetDetail.subcategoryId == subcategory_id, BudgetDetail.name == row.detail
            )
        )
    ).scalars().first()

    if existing_detail is not None:
        if mode == "append":
            summary.skipped += 1
            return
        existing_detail.accountCode = row.accountCode
        existing_detail.description = row.description
        existing_detail.isActive = row.isActive
        session.add(existing_detail)
        await session.flush()
        budget_detail_id = existing_detail.id
        summary.updated += 1
    else:
        new_detail = BudgetDetail(
            tenantId=tenant_id,
            subcategoryId=subcategory_id,
            name=row.detail,
            accountCode=row.accountCode,
            description=row.description,
            isActive=row.isActive,
        )
        session.add(new_detail)
        await session.flush()
        budget_detail_id = new_detail.id
        summary.created += 1

    # 6. DepartmentBudgetDetail 연결 (upsert)
    link = (
        await session.execute(
            select(DepartmentBudgetDetail).where(
                DepartmentBudgetDetail.departmentId == department_id,
                DepartmentBudgetDetail.budgetDetailId == budget_detail_id,
            )
        )
    ).scalars().first()
    if link is None:
        session.add(
            DepartmentBudgetDetail(
                tenantId=tenant_id,
                departmentId=department_id,
                budgetDetailId=budget_detail_id,
                isActive=True,
            )
        )
    else:
        link.isActive = True
        session.add(link)
    await session.flush()

    # 7. 담당자(User) 조회 또는 생성
    manager_id: str | None = None
    if row.manager:
        manager_id = user_cache.get(row.manager)
        if manager_id is None:
            existing_user = (
                await session.execute(
                    select(User).where(User.tenantId == tenant_id, User.username == row.manager)
                )
            ).scalars().first()
            if existing_user is not None:
                manager_id = existing_user.id
            else:
                userid = f"청연{row.manager}"
                existing_by_userid = (
                    await session.execute(
                        select(User).where(User.tenantId == tenant_id, User.userid == userid)
                    )
                ).scalars().first()
                if existing_by_userid is not None:
                    manager_id = existing_by_userid.id
                else:
                    new_user = User(
                        tenantId=tenant_id,
                        userid=userid,
                        username=row.manager,
                        password=default_password_hash,
                        role="user",
                        isActive=True,
                    )
                    session.add(new_user)
                    await session.flush()
                    manager_id = new_user.id
            user_cache[row.manager] = manager_id

    # 8. BudgetDetailYear upsert (연도별 예산금액 + 담당자)
    if row.year:
        year_setting = (
            await session.execute(
                select(BudgetDetailYear).where(
                    BudgetDetailYear.budgetDetailId == budget_detail_id,
                    BudgetDetailYear.year == row.year,
                )
            )
        ).scalars().first()
        if year_setting is None:
            session.add(
                BudgetDetailYear(
                    tenantId=tenant_id,
                    budgetDetailId=budget_detail_id,
                    year=row.year,
                    budgetAmount=row.budgetAmount or 0,
                    managerId=manager_id,
                    isActive=True,
                )
            )
        else:
            year_setting.budgetAmount = row.budgetAmount or 0
            year_setting.managerId = manager_id
            year_setting.isActive = True
            session.add(year_setting)
        await session.flush()


async def upload_budget_data(
    session: AsyncSession,
    tenant_id: str,
    rows: list[BudgetRow],
    mode: UploadMode,
    *,
    dry_run: bool = False,
) -> UploadResult:
    if dry_run:
        return await _dry_run(session, tenant_id, rows, mode)

    summary = UploadSummary(totalRows=len(rows))
    validation_errors: list[ValidationError] = []

    try:
        if mode == "replace":
            # Committee/Department 는 다른 곳에서 참조될 수 있어 유지한다.
            await session.execute(
                delete(BudgetDetailYear).where(BudgetDetailYear.tenantId == tenant_id)
            )
            await session.execute(
                delete(DepartmentBudgetDetail).where(
                    DepartmentBudgetDetail.tenantId == tenant_id
                )
            )
            await session.execute(delete(BudgetDetail).where(BudgetDetail.tenantId == tenant_id))
            await session.execute(
                delete(BudgetSubcategory).where(BudgetSubcategory.tenantId == tenant_id)
            )
            await session.execute(
                delete(BudgetCategory).where(BudgetCategory.tenantId == tenant_id)
            )

        committee_cache: dict[str, str] = {}
        department_cache: dict[str, str] = {}
        category_cache: dict[str, str] = {}
        subcategory_cache: dict[str, str] = {}
        user_cache: dict[str, str] = {}
        default_password_hash = hash_password(_DEFAULT_MANAGER_PASSWORD)

        for i, row in enumerate(rows):
            try:
                async with session.begin_nested():
                    await _upload_row(
                        session,
                        tenant_id,
                        row,
                        mode,
                        summary,
                        committee_cache,
                        department_cache,
                        category_cache,
                        subcategory_cache,
                        user_cache,
                        default_password_hash,
                    )
            except Exception as err:  # noqa: BLE001 — 행 단위 실패 격리, 다음 행 계속 처리
                summary.errors += 1
                validation_errors.append(
                    ValidationError(row=i + 2, field="database", message=str(err))
                )

        await session.commit()
    except Exception as err:  # noqa: BLE001 — Next 원본과 동일하게 트랜잭션 실패는 전체 실패로 보고
        await session.rollback()
        return UploadResult(
            success=False,
            summary=UploadSummary(totalRows=len(rows), errors=len(rows)),
            validationErrors=[ValidationError(row=0, field="transaction", message=str(err))],
        )

    return UploadResult(
        success=summary.errors == 0, summary=summary, validationErrors=validation_errors
    )


async def export_budget_template(session: AsyncSession, tenant_id: str) -> Workbook:
    """정규화 테이블 → Excel 템플릿(현재 데이터 포함)."""
    detail_rows = (
        await session.execute(
            select(BudgetDetail, BudgetSubcategory, BudgetCategory)
            .join(BudgetSubcategory, BudgetSubcategory.id == BudgetDetail.subcategoryId)
            .join(BudgetCategory, BudgetCategory.id == BudgetSubcategory.categoryId)
            .where(BudgetDetail.tenantId == tenant_id)
            .order_by(BudgetCategory.name, BudgetSubcategory.name, BudgetDetail.name)
        )
    ).all()

    detail_ids = [detail.id for detail, _, _ in detail_rows]

    dept_links: dict[str, list[tuple[Department, Committee]]] = {}
    year_settings: dict[str, list[BudgetDetailYear]] = {}
    if detail_ids:
        link_rows = (
            await session.execute(
                select(DepartmentBudgetDetail.budgetDetailId, Department, Committee)
                .join(Department, Department.id == DepartmentBudgetDetail.departmentId)
                .join(Committee, Committee.id == Department.committeeId)
                .where(DepartmentBudgetDetail.budgetDetailId.in_(detail_ids))
            )
        ).all()
        for detail_id, dept, comm in link_rows:
            dept_links.setdefault(detail_id, []).append((dept, comm))

        ys_rows = (
            await session.execute(
                select(BudgetDetailYear)
                .where(BudgetDetailYear.budgetDetailId.in_(detail_ids))
                .order_by(BudgetDetailYear.year.desc())
            )
        ).scalars().all()
        for ys in ys_rows:
            year_settings.setdefault(ys.budgetDetailId, []).append(ys)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "예산 데이터"
    sheet.append(_TEMPLATE_HEADERS)
    style_header_row(sheet)
    set_column_widths(sheet, _TEMPLATE_WIDTHS)

    for detail, sub, cat in detail_rows:
        settings = year_settings.get(detail.id) or [None]
        links = dept_links.get(detail.id) or [(None, None)]
        for dept, comm in links:
            for ys in settings:
                sheet.append(
                    [
                        comm.name if comm else "",
                        dept.name if dept else "",
                        cat.name,
                        sub.name,
                        detail.name,
                        "",
                        detail.accountCode or "",
                        detail.description or "",
                        "true" if detail.isActive else "false",
                        ys.year if ys else "",
                        ys.budgetAmount if ys else "",
                    ]
                )

    return workbook
