"""지출결의서 Excel 내보내기/일괄업로드 서비스.

(lib/excel-export.ts, lib/services/bulk-expense-upload-service.ts,
lib/services/bulk-expense-template.ts, lib/services/budget-lookup-service.ts 일부 포팅. C3)
"""

from __future__ import annotations

import io
import math
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone

from openpyxl import Workbook, load_workbook
from openpyxl.workbook import Workbook as WorkbookType
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from expense_api.core.excel import set_column_widths, style_header_row
from expense_api.core.models.budget import (
    BudgetCategory,
    BudgetDetail,
    BudgetSubcategory,
    Committee,
    Department,
    DepartmentBudgetDetail,
)
from expense_api.core.models.expense import Expense, ExpenseItem

# ============================================================
# 날짜 포맷
# ============================================================


def format_date_for_excel(value: datetime) -> str:
    return value.strftime("%Y-%m-%d")


# ============================================================
# 기본(웹교적 지출재정) 내보내기 양식
# ============================================================

_EXPORT_HEADERS = ["항", "목", "세목", "세세목", "지급방법", "예금주", "은행", "계좌번호", "금액", "날짜", "메모"]
_EXPORT_WIDTHS = [20, 20, 20, 10, 10, 12, 12, 18, 15, 12, 40]


@dataclass
class ExpenseItemForExcel:
    budgetCategory: str
    budgetSubcategory: str
    budgetDetail: str
    description: str
    amount: int


@dataclass
class ExpenseForExcel:
    accountHolder: str
    bankName: str
    accountNumber: str
    expenseDate: datetime | None
    requestDate: datetime
    items: list[ExpenseItemForExcel] = field(default_factory=list)


def expense_to_excel_rows(
    expense: ExpenseForExcel, override_date: datetime | None = None
) -> list[list]:
    date_value = override_date or expense.expenseDate or expense.requestDate
    date_str = format_date_for_excel(date_value)
    rows: list[list] = []
    for item in expense.items:
        rows.append(
            [
                item.budgetCategory or "",
                item.budgetSubcategory or "",
                item.budgetDetail,
                "",
                "이체",
                expense.accountHolder,
                expense.bankName,
                expense.accountNumber,
                item.amount,
                date_str,
                item.description,
            ]
        )
    return rows


def build_export_workbook(
    expenses: list[ExpenseForExcel], override_date: datetime | None = None
) -> WorkbookType:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "지출재정"
    sheet.append(_EXPORT_HEADERS)
    style_header_row(sheet)
    set_column_widths(sheet, _EXPORT_WIDTHS)
    for expense in expenses:
        for row in expense_to_excel_rows(expense, override_date):
            sheet.append(row)
    return workbook


def generate_export_filename(
    expenses: list[ExpenseForExcel],
    start_date: datetime | None = None,
    end_date: datetime | None = None,
) -> str:
    if len(expenses) == 1:
        expense = expenses[0]
        expense_date = expense.expenseDate or expense.requestDate
        return f"지출재정_{expense.accountHolder}_{format_date_for_excel(expense_date)}.xlsx"
    if start_date and end_date:
        return f"지출재정_{format_date_for_excel(start_date)}_{format_date_for_excel(end_date)}.xlsx"
    now = datetime.now(timezone.utc)
    return f"지출재정_{format_date_for_excel(now)}.xlsx"


# ============================================================
# 우리은행 대량이체 양식
# ============================================================


@dataclass
class ExpenseForWoori:
    bankName: str
    accountNumber: str
    accountHolder: str
    requestAmount: int


def build_woori_workbook(expenses: list[ExpenseForWoori]) -> WorkbookType:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    set_column_widths(sheet, [12, 18, 12, 12, 12, 12])
    for expense in expenses:
        sheet.append(
            [
                expense.bankName,
                expense.accountNumber.replace("-", ""),
                expense.requestAmount,
                expense.accountHolder,
                "청연교회",
                "",
            ]
        )
    return workbook


def generate_woori_filename() -> str:
    now = datetime.now(timezone.utc)
    return f"우리은행_대량이체_{format_date_for_excel(now)}.xlsx"


# ============================================================
# 예산 계층 조회 (테넌트 스코프) — lib/services/budget-lookup-service.ts 일부
# ============================================================


@dataclass
class BudgetHierarchyInfo:
    committee: str
    department: str


async def lookup_budget_hierarchy(
    session: AsyncSession, tenant_id: str, category: str, subcategory: str, detail: str
) -> BudgetHierarchyInfo | None:
    """세목 이름으로 (임의의 하나의) 위원회/사역팀을 역추출."""
    row = (
        await session.execute(
            select(Department.name, Committee.name)
            .select_from(DepartmentBudgetDetail)
            .join(BudgetDetail, BudgetDetail.id == DepartmentBudgetDetail.budgetDetailId)
            .join(BudgetSubcategory, BudgetSubcategory.id == BudgetDetail.subcategoryId)
            .join(BudgetCategory, BudgetCategory.id == BudgetSubcategory.categoryId)
            .join(Department, Department.id == DepartmentBudgetDetail.departmentId)
            .join(Committee, Committee.id == Department.committeeId)
            .where(
                DepartmentBudgetDetail.isActive.is_(True),
                BudgetDetail.tenantId == tenant_id,
                BudgetDetail.name == detail,
                BudgetSubcategory.name == subcategory,
                BudgetCategory.name == category,
            )
        )
    ).first()
    if row is None:
        return None
    department_name, committee_name = row
    return BudgetHierarchyInfo(committee=committee_name, department=department_name)


async def verify_budget_mapping(
    session: AsyncSession,
    tenant_id: str,
    committee: str,
    department: str,
    category: str,
    subcategory: str,
    detail: str,
) -> bool:
    """(위원회, 사역팀, 항, 목, 세목) 조합이 실제 활성 매핑인지 검증."""
    row = (
        await session.execute(
            select(DepartmentBudgetDetail.id)
            .join(Department, Department.id == DepartmentBudgetDetail.departmentId)
            .join(Committee, Committee.id == Department.committeeId)
            .join(BudgetDetail, BudgetDetail.id == DepartmentBudgetDetail.budgetDetailId)
            .join(BudgetSubcategory, BudgetSubcategory.id == BudgetDetail.subcategoryId)
            .join(BudgetCategory, BudgetCategory.id == BudgetSubcategory.categoryId)
            .where(
                DepartmentBudgetDetail.isActive.is_(True),
                BudgetDetail.tenantId == tenant_id,
                Department.name == department,
                Committee.name == committee,
                BudgetDetail.name == detail,
                BudgetSubcategory.name == subcategory,
                BudgetCategory.name == category,
            )
        )
    ).first()
    return row is not None


# ============================================================
# 일괄 업로드 — 파싱
# ============================================================

EXCEL_ROW_HEADERS = (
    "groupId",
    "committee",
    "department",
    "budgetCategory",
    "budgetSubcategory",
    "budgetDetail",
    "description",
    "unitPrice",
    "quantity",
    "requestDate",
    "expenseDate",
    "bankName",
    "accountNumber",
    "accountHolder",
)

MAX_ROWS = 500


@dataclass
class BulkApplicant:
    userId: str
    username: str


@dataclass
class ValidationError:
    rowNumber: int
    groupId: str | None = None
    field: str | None = None
    message: str = ""


@dataclass
class PreviewItem:
    groupId: str
    committee: str
    department: str
    applicantName: str
    itemsCount: int
    requestAmount: int


@dataclass
class BulkUploadResult:
    dryRun: bool
    totalRows: int
    totalExpenses: int
    errors: list[ValidationError] = field(default_factory=list)
    preview: list[PreviewItem] | None = None
    createdIds: list[str] | None = None


def parse_expense_excel_buffer(data: bytes) -> list[dict]:
    """Excel 버퍼 → 행 딕셔너리 목록 (헤더는 EXCEL_ROW_HEADERS 화이트리스트만 매칭)."""
    workbook = load_workbook(io.BytesIO(data), data_only=True)
    if not workbook.worksheets:
        raise ValueError("워크시트를 찾을 수 없습니다.")
    sheet = workbook.worksheets[0]

    allowed = set(EXCEL_ROW_HEADERS)
    header_row = next(sheet.iter_rows(min_row=1, max_row=1), ())
    headers: list[str | None] = []
    for cell in header_row:
        h = str(cell.value or "").strip()
        headers.append(h if h in allowed else None)

    rows: list[dict] = []
    for row in sheet.iter_rows(min_row=2):
        row_data: dict = {}
        for idx, cell in enumerate(row):
            if idx >= len(headers):
                break
            header = headers[idx]
            if header:
                row_data[header] = cell.value
        if any(v is not None and v != "" for v in row_data.values()):
            rows.append(row_data)

    return rows


# ============================================================
# 일괄 업로드 — 검증
# ============================================================

_REQUIRED_FIELDS: tuple[tuple[str, str], ...] = (
    ("committee", "위원회"),
    ("department", "사역팀(부)"),
    ("budgetCategory", "예산(항)"),
    ("budgetSubcategory", "예산(목)"),
    ("budgetDetail", "예산(세목)"),
    ("description", "적요"),
    ("requestDate", "청구일자"),
    ("bankName", "은행명"),
    ("accountNumber", "계좌번호"),
    ("accountHolder", "예금주"),
)


def _as_number(value: object) -> float:
    if value is None:
        return float("nan")
    try:
        return float(value)
    except (TypeError, ValueError):
        return float("nan")


def parse_date(value: object) -> datetime | None:
    """Excel 시리얼(숫자)·문자열·datetime 모두 처리. (parseDate 이전)"""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, (int, float)):
        excel_epoch = datetime(1899, 12, 30, tzinfo=timezone.utc)
        return excel_epoch + timedelta(days=float(value))
    try:
        text = str(value).strip()
        if not text:
            return None
        return datetime.fromisoformat(text.replace("Z", "+00:00"))
    except ValueError:
        return None


def validate_rows(rows: list[dict]) -> list[ValidationError]:
    errors: list[ValidationError] = []
    for index, row in enumerate(rows):
        row_number = index + 2
        group_id = str(row["groupId"]) if row.get("groupId") not in (None, "") else None

        for key, label in _REQUIRED_FIELDS:
            v = row.get(key)
            if v is None or v == "":
                errors.append(
                    ValidationError(rowNumber=row_number, groupId=group_id, field=key, message=f"{label} 누락")
                )

        up = _as_number(row.get("unitPrice"))
        if not (up == up) or up <= 0:  # noqa: PLR0124 — NaN 자기부정 비교로 유효성 체크
            errors.append(
                ValidationError(
                    rowNumber=row_number,
                    groupId=group_id,
                    field="unitPrice",
                    message="단가가 유효하지 않습니다 (양수만 허용)",
                )
            )

        qty = _as_number(row.get("quantity"))
        if not (qty == qty) or qty <= 0:  # noqa: PLR0124
            errors.append(
                ValidationError(
                    rowNumber=row_number,
                    groupId=group_id,
                    field="quantity",
                    message="수량이 유효하지 않습니다 (양수만 허용)",
                )
            )

        if row.get("requestDate") and parse_date(row.get("requestDate")) is None:
            errors.append(
                ValidationError(
                    rowNumber=row_number,
                    groupId=group_id,
                    field="requestDate",
                    message=f"청구일자 파싱 실패: {row.get('requestDate')}",
                )
            )
        if row.get("expenseDate") and parse_date(row.get("expenseDate")) is None:
            errors.append(
                ValidationError(
                    rowNumber=row_number,
                    groupId=group_id,
                    field="expenseDate",
                    message=f"지급일자 파싱 실패: {row.get('expenseDate')}",
                )
            )

    return errors


def _floor_or_zero(value: float) -> int:
    if value != value:  # noqa: PLR0124 — NaN 자기부정
        return 0
    return math.floor(value)


def compute_item_amount(row: dict) -> tuple[int, int, int]:
    """(unitPrice, quantity, amount) — 미리보기/commit 공용. floor10 규칙 미적용(원본 계약)."""
    unit_price = _floor_or_zero(_as_number(row.get("unitPrice")))
    quantity = _floor_or_zero(_as_number(row.get("quantity")))
    return unit_price, quantity, unit_price * quantity


def group_rows(rows: list[dict]) -> dict[str, list[dict]]:
    groups: dict[str, list[dict]] = {}
    for index, row in enumerate(rows):
        key = str(row["groupId"]) if row.get("groupId") not in (None, "") else f"__single_{index}"
        groups.setdefault(key, []).append(row)
    return groups


# ============================================================
# 일괄 업로드 — 실행
# ============================================================


async def execute_bulk_upload(
    session: AsyncSession,
    tenant_id: str,
    rows: list[dict],
    *,
    dry_run: bool,
    applicant: BulkApplicant,
) -> BulkUploadResult:
    validation_errors = validate_rows(rows)
    groups = group_rows(rows)

    budget_errors: list[ValidationError] = []
    budget_cache: dict[str, BudgetHierarchyInfo] = {}
    resolved_by_group: dict[str, BudgetHierarchyInfo] = {}
    preview: list[PreviewItem] = []

    for group_key, group_rows_list in groups.items():
        first = group_rows_list[0]
        row_number = rows.index(first) + 2

        cat = str(first.get("budgetCategory") or "").strip()
        sub = str(first.get("budgetSubcategory") or "").strip()
        det = str(first.get("budgetDetail") or "").strip()
        input_committee = str(first.get("committee") or "").strip()
        input_department = str(first.get("department") or "").strip()

        budget_info: BudgetHierarchyInfo | None = None
        if cat and sub and det:
            cache_key = f"{cat}|{sub}|{det}"
            if cache_key in budget_cache:
                budget_info = budget_cache[cache_key]
            else:
                hier = await lookup_budget_hierarchy(session, tenant_id, cat, sub, det)
                if hier:
                    budget_info = hier
                    budget_cache[cache_key] = budget_info
                else:
                    budget_errors.append(
                        ValidationError(
                            rowNumber=row_number,
                            groupId=group_key,
                            message=f"예산 정보를 찾을 수 없습니다: {cat} / {sub} / {det}",
                        )
                    )
                    continue
        else:
            continue

        if input_committee and input_department:
            valid = await verify_budget_mapping(
                session, tenant_id, input_committee, input_department, cat, sub, det
            )
            if not valid:
                budget_errors.append(
                    ValidationError(
                        rowNumber=row_number,
                        groupId=group_key,
                        field="department",
                        message=(
                            f"위원회/사역팀이 해당 세목에 매핑되어 있지 않습니다: "
                            f"{input_committee} / {input_department} / {cat} / {sub} / {det}"
                        ),
                    )
                )
                continue
            budget_info = BudgetHierarchyInfo(committee=input_committee, department=input_department)

        resolved_by_group[group_key] = budget_info

        request_amount = sum(compute_item_amount(r)[2] for r in group_rows_list)

        preview.append(
            PreviewItem(
                groupId=group_key,
                committee=budget_info.committee,
                department=budget_info.department,
                applicantName=applicant.username,
                itemsCount=len(group_rows_list),
                requestAmount=request_amount,
            )
        )

    all_errors = [*validation_errors, *budget_errors]

    if dry_run:
        return BulkUploadResult(
            dryRun=True,
            totalRows=len(rows),
            totalExpenses=len(groups),
            errors=all_errors,
            preview=preview,
        )

    if all_errors:
        return BulkUploadResult(
            dryRun=False, totalRows=len(rows), totalExpenses=len(groups), errors=all_errors
        )

    created_ids: list[str] = []
    for group_key, group_rows_list in groups.items():
        first = group_rows_list[0]
        budget_info = resolved_by_group.get(group_key)
        if budget_info is None:
            raise RuntimeError("내부 오류: 검증 단계에서 채워졌어야 할 캐시가 비어있습니다.")

        items: list[ExpenseItem] = []
        total_amount = 0
        for idx, r in enumerate(group_rows_list):
            unit_price, quantity, amount = compute_item_amount(r)
            total_amount += amount
            items.append(
                ExpenseItem(
                    tenantId=tenant_id,
                    budgetCategory=str(r["budgetCategory"]).strip(),
                    budgetSubcategory=str(r["budgetSubcategory"]).strip(),
                    budgetDetail=str(r["budgetDetail"]).strip(),
                    description=str(r["description"]).strip(),
                    unitPrice=unit_price,
                    quantity=quantity,
                    amount=amount,
                    order=idx + 1,
                )
            )

        expense = Expense(
            tenantId=tenant_id,
            userId=applicant.userId,
            committee=budget_info.committee,
            department=budget_info.department,
            expenseDate=parse_date(first.get("expenseDate")),
            requestAmount=total_amount,
            requestDate=parse_date(first.get("requestDate")),
            requestTeam="출납팀",
            applicantName=applicant.username,
            applicantTitle=None,
            bankName=str(first["bankName"]).strip(),
            accountNumber=str(first["accountNumber"]).strip(),
            accountHolder=str(first["accountHolder"]).strip(),
        )
        session.add(expense)
        await session.flush()
        for item in items:
            item.expenseId = expense.id
            session.add(item)
        created_ids.append(expense.id)

    await session.commit()

    return BulkUploadResult(
        dryRun=False, totalRows=len(rows), totalExpenses=len(groups), errors=[], createdIds=created_ids
    )


# ============================================================
# 일괄 업로드 템플릿
# ============================================================

_TEMPLATE_HEADERS = list(EXCEL_ROW_HEADERS)
_TEMPLATE_WIDTHS = [10, 14, 14, 15, 18, 20, 30, 10, 8, 12, 12, 12, 18, 12]

_TEMPLATE_SAMPLE_ROWS = [
    [
        1,
        "교육위원회",
        "기획팀",
        "사역지원비",
        "기획비",
        "아웃팅비",
        "기획팀 회의 후 식사",
        10000,
        5,
        "2026-05-01",
        "2026-05-05",
        "우리은행",
        "1002-123-456789",
        "홍길동",
    ],
    [
        1,
        "교육위원회",
        "기획팀",
        "사역지원비",
        "기획비",
        "행사비(전교인행사)",
        "기획팀 회의 다과",
        5000,
        10,
        "2026-05-01",
        "2026-05-05",
        "우리은행",
        "1002-123-456789",
        "홍길동",
    ],
    [
        2,
        "예배위원회",
        "찬양팀",
        "예배사역비",
        "찬양팀운영비",
        "소모품비",
        "마이크 커버 구매",
        3000,
        20,
        "2026-05-02",
        "",
        "국민은행",
        "123-45-6789012",
        "김찬양",
    ],
]

_HEADER_DESCRIPTIONS: dict[str, str] = {
    "groupId": "그룹ID — 같은 ID는 한 지출결의서로 묶임 (비우면 행마다 별건)",
    "committee": "위원회 (예산 매핑과 교차 검증)",
    "department": "사역팀(부) (예산 매핑과 교차 검증)",
    "budgetCategory": "예산(항)",
    "budgetSubcategory": "예산(목)",
    "budgetDetail": "예산(세목)",
    "description": "적요",
    "unitPrice": "단가",
    "quantity": "수량",
    "requestDate": "청구일자 (YYYY-MM-DD)",
    "expenseDate": "지급일자 (YYYY-MM-DD, 선택)",
    "bankName": "은행명 (수취 계좌)",
    "accountNumber": "계좌번호 (수취 계좌)",
    "accountHolder": "예금주 (수취 계좌)",
}

_OPTIONAL_FIELDS = {"groupId", "expenseDate"}


def build_expense_template_workbook() -> WorkbookType:
    workbook = Workbook()
    data_sheet = workbook.active
    data_sheet.title = "업로드데이터"
    data_sheet.append(_TEMPLATE_HEADERS)
    style_header_row(data_sheet)
    set_column_widths(data_sheet, _TEMPLATE_WIDTHS)
    for row in _TEMPLATE_SAMPLE_ROWS:
        data_sheet.append(row)

    desc_sheet = workbook.create_sheet("컬럼설명")
    desc_sheet.append(["컬럼명", "설명", "필수여부"])
    style_header_row(desc_sheet)
    set_column_widths(desc_sheet, [18, 45, 10])
    for h in EXCEL_ROW_HEADERS:
        desc_sheet.append([h, _HEADER_DESCRIPTIONS[h], "선택" if h in _OPTIONAL_FIELDS else "필수"])

    return workbook
