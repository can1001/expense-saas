"""C4 계약 테스트 — users/upload, departments/leaders-upload,
budget-details/year(+auto-assign), budget-details/{id}/description.
(app/api/users/upload, departments/leaders-upload, budget-details/year*,
budget-details/[id]/description 이전)
"""

import io

import pytest_asyncio
from httpx import ASGITransport, AsyncClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import event, select
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine
from sqlalchemy.pool import StaticPool
from sqlmodel import SQLModel

import expense_api.core.models  # noqa: F401
from expense_api.core.db.session import get_session
from expense_api.core.models.budget import (
    BudgetCategory,
    BudgetDetail,
    BudgetDetailYear,
    BudgetSubcategory,
    Committee,
    Department,
    DepartmentBudgetDetail,
)
from expense_api.core.models.tenant import Tenant
from expense_api.core.models.user import Membership, Role, User, UserYearRole
from expense_api.core.security.jwt import hash_password
from expense_api.core.security.rate_limit import _reset_all
from main import app

XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
YEAR = 2026


@pytest_asyncio.fixture
async def client():
    _reset_all()
    engine = create_async_engine(
        "sqlite+aiosqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )

    @event.listens_for(engine.sync_engine, "connect")
    def _fk_on(dbapi_conn, _rec):
        cur = dbapi_conn.cursor()
        cur.execute("PRAGMA foreign_keys=ON")
        cur.close()

    async with engine.begin() as conn:
        await conn.run_sync(SQLModel.metadata.create_all)
    maker = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)

    async with maker() as s:
        t = Tenant(name="데모", subdomain="demo")
        s.add(t)
        await s.flush()
        s.add(
            User(
                tenantId=t.id,
                userid="admin",
                username="관리자",
                password=hash_password("admin123"),
                role="admin",
            )
        )
        await s.commit()

    async def _override():
        async with maker() as s:
            yield s

    app.dependency_overrides[get_session] = _override
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as c:
        c._maker = maker  # type: ignore[attr-defined]
        yield c
    app.dependency_overrides.clear()
    await engine.dispose()


async def _login(client: AsyncClient) -> dict:
    r = await client.post("/api/auth/login", json={"userid": "admin", "password": "admin123"})
    assert r.status_code == 200
    return {"Authorization": f"Bearer {r.json()['token']}"}


def _xlsx_bytes(headers: list[str], rows: list[list]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


# ── /api/users/upload ────────────────────────────────────────────────
async def test_users_upload_template_requires_auth(client: AsyncClient):
    r = await client.get("/api/users/upload")
    assert r.status_code == 401


async def test_users_upload_template_download(client: AsyncClient):
    headers = await _login(client)
    r = await client.get("/api/users/upload", headers=headers)
    assert r.status_code == 200
    assert r.headers["content-type"] == XLSX_CONTENT_TYPE
    assert r.headers["content-disposition"].startswith('attachment; filename="users_template_')

    wb = load_workbook(io.BytesIO(r.content))
    ws = wb["사용자목록"]
    header = [cell.value for cell in ws[1]]
    assert header == [
        "userid (아이디)", "username (이름)", "role (역할)", "department (부서)", "isActive (활성화)",
    ]
    # 관리자 시드 1건이 포함되어야 함
    row2 = [cell.value for cell in ws[2]]
    assert row2[0] == "admin"


async def test_users_upload_requires_file(client: AsyncClient):
    headers = await _login(client)
    r = await client.post("/api/users/upload", headers=headers, data={"mode": "merge"})
    assert r.status_code == 400
    body = r.json()
    assert body["success"] is False
    assert body["error"]["type"] == "VALIDATION_ERROR"


async def test_users_upload_creates_users_and_memberships(client: AsyncClient):
    headers = await _login(client)
    data = _xlsx_bytes(
        ["userid", "username", "role", "department", "isActive"],
        [["newuser1", "새사용자1", "사용자", "재정팀", "Y"]],
    )
    files = {"file": ("users.xlsx", data, XLSX_CONTENT_TYPE)}
    r = await client.post(
        "/api/users/upload", headers=headers, files=files, data={"mode": "merge"}
    )
    assert r.status_code == 200
    body = r.json()
    assert body["success"] is True
    assert body["data"]["summary"]["created"] == 1

    maker = client._maker  # type: ignore[attr-defined]
    async with maker() as s:
        user = (
            await s.execute(select(User).where(User.userid == "newuser1"))
        ).scalars().first()
        assert user is not None
        assert user.role == "user"
        assert user.mustChangePassword is True

        membership = (
            await s.execute(select(Membership).where(Membership.userId == user.id))
        ).scalars().first()
        assert membership is not None
        assert membership.isDefault is True


async def test_users_upload_validation_error_returns_200(client: AsyncClient):
    headers = await _login(client)
    data = _xlsx_bytes(
        ["userid", "username", "role", "department", "isActive"],
        [["", "이름만있음", "사용자", "", "Y"]],
    )
    files = {"file": ("users.xlsx", data, XLSX_CONTENT_TYPE)}
    r = await client.post(
        "/api/users/upload", headers=headers, files=files, data={"mode": "merge"}
    )
    assert r.status_code == 200
    body = r.json()
    assert body["success"] is False
    assert body["error"]["type"] == "VALIDATION_ERROR"
    assert len(body["error"]["fields"]) == 1


async def test_users_upload_dry_run_has_no_side_effects(client: AsyncClient):
    headers = await _login(client)
    data = _xlsx_bytes(
        ["userid", "username", "role", "department", "isActive"],
        [["dryrunuser", "드라이런", "사용자", "", "Y"]],
    )
    files = {"file": ("users.xlsx", data, XLSX_CONTENT_TYPE)}
    r = await client.post(
        "/api/users/upload",
        headers=headers,
        files=files,
        data={"mode": "merge", "dryRun": "true"},
    )
    assert r.status_code == 200
    body = r.json()
    assert body["data"]["dryRun"] is True

    maker = client._maker  # type: ignore[attr-defined]
    async with maker() as s:
        user = (
            await s.execute(select(User).where(User.userid == "dryrunuser"))
        ).scalars().first()
        assert user is None


# ── /api/departments/leaders-upload ────────────────────────────────────
async def _seed_department(client: AsyncClient) -> dict:
    maker = client._maker  # type: ignore[attr-defined]
    async with maker() as s:
        tid = (await s.execute(select(Tenant.id))).scalar_one()
        comm = Committee(tenantId=tid, name="교육위원회", sortOrder=1)
        s.add(comm)
        await s.flush()
        dept = Department(tenantId=tid, committeeId=comm.id, name="유년부", sortOrder=1)
        leader = User(tenantId=tid, userid="leader1", username="정혜종", role="user")
        s.add_all([dept, leader])
        await s.flush()
        s.add(Role(tenantId=tid, code="team_leader", name="팀장"))
        await s.commit()
        return {"committeeId": comm.id, "departmentId": dept.id, "leaderId": leader.id}


async def test_leaders_upload_template_download(client: AsyncClient):
    headers = await _login(client)
    await _seed_department(client)
    r = await client.get(
        "/api/departments/leaders-upload", headers=headers, params={"year": YEAR}
    )
    assert r.status_code == 200
    assert r.headers["content-type"] == XLSX_CONTENT_TYPE
    assert r.headers["content-disposition"].startswith('attachment; filename="leaders_template_')

    wb = load_workbook(io.BytesIO(r.content))
    ws = wb["사역팀장목록"]
    header = [cell.value for cell in ws[1]]
    assert header == ["위원회", "사역팀", "팀장"]


async def test_leaders_upload_requires_file(client: AsyncClient):
    headers = await _login(client)
    r = await client.post("/api/departments/leaders-upload", headers=headers, data={})
    assert r.status_code == 400


async def test_leaders_upload_sets_and_clears_leader(client: AsyncClient):
    headers = await _login(client)
    seed = await _seed_department(client)

    data = _xlsx_bytes(
        ["위원회", "사역팀", "팀장"],
        [["교육위원회", "유년부", "정혜종"]],
    )
    files = {"file": ("leaders.xlsx", data, XLSX_CONTENT_TYPE)}
    r = await client.post(
        "/api/departments/leaders-upload",
        headers=headers,
        files=files,
        data={"year": str(YEAR)},
    )
    assert r.status_code == 200
    body = r.json()
    assert body["success"] is True
    assert body["data"]["summary"]["updated"] == 1

    maker = client._maker  # type: ignore[attr-defined]
    async with maker() as s:
        yr = (
            await s.execute(
                select(UserYearRole).where(
                    UserYearRole.departmentId == seed["departmentId"],
                    UserYearRole.role == "team_leader",
                )
            )
        ).scalars().first()
        assert yr is not None
        assert yr.userId == seed["leaderId"]

    # 팀장 비우기 → 기존 UserYearRole 삭제
    data2 = _xlsx_bytes(
        ["위원회", "사역팀", "팀장"],
        [["교육위원회", "유년부", ""]],
    )
    files2 = {"file": ("leaders2.xlsx", data2, XLSX_CONTENT_TYPE)}
    r2 = await client.post(
        "/api/departments/leaders-upload",
        headers=headers,
        files=files2,
        data={"year": str(YEAR)},
    )
    assert r2.status_code == 200

    async with maker() as s:
        yr = (
            await s.execute(
                select(UserYearRole).where(
                    UserYearRole.departmentId == seed["departmentId"],
                    UserYearRole.role == "team_leader",
                )
            )
        ).scalars().first()
        assert yr is None


async def test_leaders_upload_unknown_department_is_validation_error(client: AsyncClient):
    headers = await _login(client)
    await _seed_department(client)
    data = _xlsx_bytes(
        ["위원회", "사역팀", "팀장"],
        [["교육위원회", "없는부서", "정혜종"]],
    )
    files = {"file": ("leaders.xlsx", data, XLSX_CONTENT_TYPE)}
    r = await client.post(
        "/api/departments/leaders-upload",
        headers=headers,
        files=files,
        data={"year": str(YEAR)},
    )
    assert r.status_code == 200
    body = r.json()
    assert body["success"] is False
    assert body["error"]["type"] == "VALIDATION_ERROR"


# ── /api/budget-details/year, year/auto-assign ──────────────────────────
async def _seed_budget_detail(client: AsyncClient) -> dict:
    maker = client._maker  # type: ignore[attr-defined]
    async with maker() as s:
        tid = (await s.execute(select(Tenant.id))).scalar_one()
        manager = User(tenantId=tid, userid="mgr", username="김담당", role="user")
        s.add(manager)
        comm = Committee(tenantId=tid, name="기획본부", sortOrder=1)
        s.add(comm)
        await s.flush()
        dept = Department(tenantId=tid, committeeId=comm.id, name="재정팀", sortOrder=1)
        cat = BudgetCategory(tenantId=tid, name="사무행정비", sortOrder=1)
        s.add_all([dept, cat])
        await s.flush()
        sub = BudgetSubcategory(tenantId=tid, categoryId=cat.id, name="회의비", sortOrder=1)
        s.add(sub)
        await s.flush()
        detail = BudgetDetail(tenantId=tid, subcategoryId=sub.id, name="간식비", sortOrder=1)
        s.add(detail)
        await s.flush()
        s.add(
            DepartmentBudgetDetail(tenantId=tid, departmentId=dept.id, budgetDetailId=detail.id)
        )
        await s.commit()
        return {
            "departmentId": dept.id,
            "detailId": detail.id,
            "managerId": manager.id,
        }


async def test_budget_details_year_get(client: AsyncClient):
    headers = await _login(client)
    seed = await _seed_budget_detail(client)

    r = await client.get(
        "/api/budget-details/year", headers=headers, params={"year": YEAR}
    )
    assert r.status_code == 200
    body = r.json()
    assert body["year"] == YEAR
    assert body["total"] == 1
    detail = body["details"][0]
    assert detail["id"] == seed["detailId"]
    assert detail["yearSetting"] is None
    assert detail["departments"][0]["name"] == "재정팀"


async def test_budget_details_year_post_upserts(client: AsyncClient):
    headers = await _login(client)
    seed = await _seed_budget_detail(client)

    r = await client.post(
        "/api/budget-details/year",
        headers=headers,
        json={
            "year": YEAR,
            "settings": [
                {
                    "budgetDetailId": seed["detailId"],
                    "managerId": seed["managerId"],
                    "budgetAmount": 500_000,
                }
            ],
        },
    )
    assert r.status_code == 200
    body = r.json()
    assert body["results"][0]["budgetAmount"] == 500_000
    assert body["results"][0]["manager"]["id"] == seed["managerId"]

    r2 = await client.get(
        "/api/budget-details/year", headers=headers, params={"year": YEAR}
    )
    detail = r2.json()["details"][0]
    assert detail["yearSetting"]["budgetAmount"] == 500_000
    assert detail["yearSetting"]["managerName"] == "김담당"


async def test_budget_details_year_auto_assign(client: AsyncClient):
    headers = await _login(client)
    seed = await _seed_budget_detail(client)

    maker = client._maker  # type: ignore[attr-defined]
    async with maker() as s:
        tid = (await s.execute(select(Tenant.id))).scalar_one()
        s.add(
            UserYearRole(
                tenantId=tid,
                userId=seed["managerId"],
                year=YEAR,
                role="team_leader",
                departmentId=seed["departmentId"],
            )
        )
        await s.commit()

    r = await client.post(
        "/api/budget-details/year/auto-assign",
        headers=headers,
        json={"year": YEAR},
    )
    assert r.status_code == 200
    body = r.json()
    assert body["summary"]["created"] == 1
    assert body["results"][0]["managerId"] == seed["managerId"]

    async with maker() as s:
        ys = (
            await s.execute(
                select(BudgetDetailYear).where(
                    BudgetDetailYear.budgetDetailId == seed["detailId"], BudgetDetailYear.year == YEAR
                )
            )
        ).scalars().first()
        assert ys is not None
        assert ys.managerId == seed["managerId"]


async def test_budget_details_year_auto_assign_requires_year(client: AsyncClient):
    headers = await _login(client)
    r = await client.post(
        "/api/budget-details/year/auto-assign", headers=headers, json={}
    )
    assert r.status_code == 400


# ── /api/budget-details/{id}/description ────────────────────────────────
async def test_update_detail_description(client: AsyncClient):
    headers = await _login(client)
    seed = await _seed_budget_detail(client)

    r = await client.patch(
        f"/api/budget-details/{seed['detailId']}/description",
        headers=headers,
        json={"description": "회의 간식, 다과"},
    )
    assert r.status_code == 200
    body = r.json()
    assert body["success"] is True
    assert body["budgetDetail"]["description"] == "회의 간식, 다과"


async def test_update_detail_description_requires_field(client: AsyncClient):
    headers = await _login(client)
    seed = await _seed_budget_detail(client)

    r = await client.patch(
        f"/api/budget-details/{seed['detailId']}/description", headers=headers, json={}
    )
    assert r.status_code == 400


async def test_update_detail_description_404(client: AsyncClient):
    headers = await _login(client)
    r = await client.patch(
        "/api/budget-details/doesnotexist12345678/description",
        headers=headers,
        json={"description": "x"},
    )
    assert r.status_code == 404
