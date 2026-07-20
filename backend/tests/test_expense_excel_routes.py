"""expenses Excel 계열 계약 테스트 — export/excel, bulk-upload, bulk-upload-template.
(app/api/expenses/export/excel, bulk-upload, bulk-upload-template 컷오버, C3)
"""

import io
from datetime import datetime

import pytest_asyncio
from httpx import ASGITransport, AsyncClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import event, select
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine
from sqlalchemy.pool import StaticPool
from sqlmodel import SQLModel

import expense_api.core.models  # noqa: F401
from expense_api.core.db.session import get_session
from expense_api.core.models.budget import (
    BudgetCategory,
    BudgetDetail,
    BudgetSubcategory,
    Committee,
    Department,
    DepartmentBudgetDetail,
)
from expense_api.core.models.expense import Expense, ExpenseItem
from expense_api.core.models.tenant import Tenant
from expense_api.core.models.user import User
from expense_api.core.security.jwt import hash_password
from expense_api.core.security.rate_limit import _reset_all
from main import app

XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

_UPLOAD_HEADERS = [
    "groupId", "committee", "department", "budgetCategory", "budgetSubcategory",
    "budgetDetail", "description", "unitPrice", "quantity", "requestDate",
    "expenseDate", "bankName", "accountNumber", "accountHolder",
]


@pytest_asyncio.fixture
async def client():
    _reset_all()
    engine = create_async_engine(
        "sqlite+aiosqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )

    @event.listens_for(engine.sync_engine, "connect")
    def _fk_on(dbapi_conn, _rec):
        cur = dbapi_conn.cursor()
        cur.execute("PRAGMA foreign_keys=ON")
        cur.close()

    async with engine.begin() as conn:
        await conn.run_sync(SQLModel.metadata.create_all)
    maker = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)

    async with maker() as s:
        t = Tenant(name="데모", subdomain="demo")
        s.add(t)
        await s.flush()
        s.add_all(
            [
                User(
                    tenantId=t.id, userid="admin", username="관리자",
                    password=hash_password("admin123"), role="admin",
                ),
                User(
                    tenantId=t.id, userid="plain", username="일반유저",
                    password=hash_password("plain123"), role="user",
                ),
            ]
        )
        await s.commit()

    async def _override():
        async with maker() as s:
            yield s

    app.dependency_overrides[get_session] = _override
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as c:
        c._maker = maker  # type: ignore[attr-defined]
        yield c
    app.dependency_overrides.clear()
    await engine.dispose()


async def _login(client: AsyncClient, userid: str = "admin", password: str = "admin123") -> dict:
    r = await client.post("/api/auth/login", json={"userid": userid, "password": password})
    assert r.status_code == 200
    return {"Authorization": f"Bearer {r.json()['token']}"}


async def _seed_budget_tree(client: AsyncClient) -> dict:
    maker = client._maker  # type: ignore[attr-defined]
    async with maker() as s:
        tid = (await s.execute(select(Tenant.id))).scalar_one()

        comm = Committee(tenantId=tid, name="교육위원회", sortOrder=1)
        s.add(comm)
        await s.flush()
        dept = Department(tenantId=tid, committeeId=comm.id, name="기획팀", sortOrder=1)
        cat = BudgetCategory(tenantId=tid, name="사역지원비", sortOrder=1)
        s.add_all([dept, cat])
        await s.flush()
        sub = BudgetSubcategory(tenantId=tid, categoryId=cat.id, name="기획비", sortOrder=1)
        s.add(sub)
        await s.flush()
        detail = BudgetDetail(tenantId=tid, subcategoryId=sub.id, name="아웃팅비", sortOrder=1)
        s.add(detail)
        await s.flush()
        s.add(DepartmentBudgetDetail(tenantId=tid, departmentId=dept.id, budgetDetailId=detail.id))
        await s.commit()
        return {"tid": tid, "committeeId": comm.id, "departmentId": dept.id, "detailId": detail.id}


async def _seed_expense(client: AsyncClient, *, status: str = "APPROVED_FINAL") -> dict:
    maker = client._maker  # type: ignore[attr-defined]
    async with maker() as s:
        tid = (await s.execute(select(Tenant.id))).scalar_one()
        admin_id = (await s.execute(select(User.id).where(User.userid == "admin"))).scalar_one()

        expense = Expense(
            tenantId=tid, userId=admin_id, committee="교육위원회", department="기획팀",
            expenseDate=datetime(2026, 5, 5), requestAmount=50_000,
            requestDate=datetime(2026, 5, 1), applicantName="관리자",
            bankName="우리은행", accountNumber="1002-123-456789", accountHolder="홍길동",
            status=status,
        )
        s.add(expense)
        await s.flush()
        s.add(
            ExpenseItem(
                tenantId=tid, expenseId=expense.id, budgetCategory="사역지원비",
                budgetSubcategory="기획비", budgetDetail="아웃팅비",
                description="기획팀 회의 후 식사", unitPrice=10_000, quantity=5,
                amount=50_000, order=1,
            )
        )
        await s.commit()
        return {"tid": tid, "expenseId": expense.id}


def _make_upload_xlsx(rows: list[list]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(_UPLOAD_HEADERS)
    for row in rows:
        sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


_VALID_ROW = [
    1, "교육위원회", "기획팀", "사역지원비", "기획비", "아웃팅비",
    "회의 다과", 10000, 5, "2026-05-01", "2026-05-05",
    "우리은행", "1002-123-456789", "홍길동",
]


# ── GET /export/excel ───────────────────────────────────────────────────
async def test_export_excel_requires_auth(client: AsyncClient):
    r = await client.get("/api/expenses/export/excel")
    assert r.status_code == 401


async def test_export_excel_returns_404_when_no_match(client: AsyncClient):
    headers = await _login(client)
    r = await client.get("/api/expenses/export/excel", headers=headers)
    assert r.status_code == 404
    assert r.json()["detail"] == "내보낼 지출결의서가 없습니다."


async def test_export_excel_default_format(client: AsyncClient):
    headers = await _login(client)
    await _seed_expense(client)

    r = await client.get("/api/expenses/export/excel", headers=headers)
    assert r.status_code == 200
    assert r.headers["content-type"] == XLSX_CONTENT_TYPE
    assert "filename*=UTF-8''" in r.headers["content-disposition"]

    wb = load_workbook(io.BytesIO(r.content))
    ws = wb.active
    assert ws.title == "지출재정"
    header = [c.value for c in ws[1]]
    assert header == ["항", "목", "세목", "세세목", "지급방법", "예금주", "은행", "계좌번호", "금액", "날짜", "메모"]
    data_row = [c.value for c in ws[2]]
    assert data_row == [
        "사역지원비", "기획비", "아웃팅비", None, "이체", "홍길동", "우리은행",
        "1002-123-456789", 50_000, "2026-05-05", "기획팀 회의 후 식사",
    ]


async def test_export_excel_woori_format(client: AsyncClient):
    headers = await _login(client)
    await _seed_expense(client)

    r = await client.get("/api/expenses/export/excel?format=woori", headers=headers)
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb.active
    assert ws.title == "Sheet1"
    row = [c.value for c in ws[1]]
    assert row == ["우리은행", "1002123456789", 50_000, "홍길동", "청연교회", None]


async def test_export_excel_status_filter_excludes_draft(client: AsyncClient):
    headers = await _login(client)
    await _seed_expense(client, status="DRAFT")
    r = await client.get("/api/expenses/export/excel", headers=headers)
    assert r.status_code == 404


# ── GET /bulk-upload-template ───────────────────────────────────────────
async def test_bulk_upload_template_requires_auth(client: AsyncClient):
    r = await client.get("/api/expenses/bulk-upload-template")
    assert r.status_code == 401


async def test_bulk_upload_template_forbidden_for_plain_user(client: AsyncClient):
    headers = await _login(client, "plain", "plain123")
    r = await client.get("/api/expenses/bulk-upload-template", headers=headers)
    assert r.status_code == 403
    assert r.json()["detail"] == "권한이 없습니다."


async def test_bulk_upload_template_returns_xlsx_with_headers(client: AsyncClient):
    headers = await _login(client)
    r = await client.get("/api/expenses/bulk-upload-template", headers=headers)
    assert r.status_code == 200
    assert r.headers["content-type"] == XLSX_CONTENT_TYPE
    assert r.headers["content-disposition"] == (
        'attachment; filename="expense-bulk-upload-template.xlsx"'
    )

    wb = load_workbook(io.BytesIO(r.content))
    ws = wb["업로드데이터"]
    header = [c.value for c in ws[1]]
    assert header == _UPLOAD_HEADERS


# ── POST /bulk-upload ───────────────────────────────────────────────────
async def test_bulk_upload_requires_file(client: AsyncClient):
    headers = await _login(client)
    r = await client.post("/api/expenses/bulk-upload", headers=headers)
    assert r.status_code == 400
    assert r.json()["detail"] == "파일이 필요합니다."


async def test_bulk_upload_rejects_non_xlsx(client: AsyncClient):
    headers = await _login(client)
    files = {"file": ("data.txt", b"not excel", "text/plain")}
    r = await client.post("/api/expenses/bulk-upload", headers=headers, files=files)
    assert r.status_code == 400
    assert r.json()["detail"] == "Excel(.xlsx) 파일만 업로드할 수 있습니다."


async def test_bulk_upload_forbidden_for_plain_user(client: AsyncClient):
    headers = await _login(client, "plain", "plain123")
    files = {"file": ("data.xlsx", _make_upload_xlsx([_VALID_ROW]), XLSX_CONTENT_TYPE)}
    r = await client.post("/api/expenses/bulk-upload", headers=headers, files=files)
    assert r.status_code == 403


async def test_bulk_upload_dry_run_makes_no_db_changes(client: AsyncClient):
    headers = await _login(client)
    await _seed_budget_tree(client)

    files = {"file": ("data.xlsx", _make_upload_xlsx([_VALID_ROW]), XLSX_CONTENT_TYPE)}
    r = await client.post(
        "/api/expenses/bulk-upload", headers=headers, files=files, data={"dryRun": "true"}
    )
    assert r.status_code == 200
    body = r.json()
    assert body["dryRun"] is True
    assert body["totalRows"] == 1
    assert body["totalExpenses"] == 1
    assert body["errors"] == []
    assert body["preview"][0]["committee"] == "교육위원회"
    assert body["preview"][0]["department"] == "기획팀"
    assert body["preview"][0]["requestAmount"] == 50_000
    assert "createdIds" not in body

    maker = client._maker  # type: ignore[attr-defined]
    async with maker() as s:
        count = len((await s.execute(select(Expense))).scalars().all())
        assert count == 0


async def test_bulk_upload_commit_creates_expense(client: AsyncClient):
    headers = await _login(client)
    await _seed_budget_tree(client)

    files = {"file": ("data.xlsx", _make_upload_xlsx([_VALID_ROW]), XLSX_CONTENT_TYPE)}
    r = await client.post("/api/expenses/bulk-upload", headers=headers, files=files)
    assert r.status_code == 200
    body = r.json()
    assert body["dryRun"] is False
    assert body["errors"] == []
    assert len(body["createdIds"]) == 1

    maker = client._maker  # type: ignore[attr-defined]
    async with maker() as s:
        expense = (await s.execute(select(Expense))).scalars().one()
        assert expense.status == "DRAFT"
        assert expense.committee == "교육위원회"
        assert expense.department == "기획팀"
        assert expense.requestAmount == 50_000
        assert expense.applicantName == "관리자"

        item = (await s.execute(select(ExpenseItem))).scalars().one()
        assert item.unitPrice == 10_000
        assert item.quantity == 5
        assert item.amount == 50_000


async def test_bulk_upload_missing_required_field_reports_row_error(client: AsyncClient):
    headers = await _login(client)
    await _seed_budget_tree(client)

    bad_row = list(_VALID_ROW)
    bad_row[6] = ""  # description 누락
    files = {"file": ("data.xlsx", _make_upload_xlsx([bad_row]), XLSX_CONTENT_TYPE)}
    r = await client.post("/api/expenses/bulk-upload", headers=headers, files=files)
    assert r.status_code == 200
    body = r.json()
    assert body["dryRun"] is False
    assert any(e["field"] == "description" for e in body["errors"])
    assert "createdIds" not in body

    maker = client._maker  # type: ignore[attr-defined]
    async with maker() as s:
        count = len((await s.execute(select(Expense))).scalars().all())
        assert count == 0


async def test_bulk_upload_department_mismatch_reports_error(client: AsyncClient):
    headers = await _login(client)
    await _seed_budget_tree(client)

    bad_row = list(_VALID_ROW)
    bad_row[2] = "다른부서"  # department 가 실제 매핑과 불일치
    files = {"file": ("data.xlsx", _make_upload_xlsx([bad_row]), XLSX_CONTENT_TYPE)}
    r = await client.post("/api/expenses/bulk-upload", headers=headers, files=files)
    assert r.status_code == 200
    body = r.json()
    assert any("매핑되어 있지 않습니다" in e["message"] for e in body["errors"])
