"""budget 조회 라우트 계약 테스트 — hierarchy/search/simple/all-details/usage-details/memo-examples
+ committees DELETE. (app/api/budget/* 컷오버)

레거시 Next.js 응답 형태와의 계약 정합을 검증한다.
"""

import io
from datetime import datetime

import pytest_asyncio
from httpx import ASGITransport, AsyncClient
from openpyxl import load_workbook
from sqlalchemy import event
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine
from sqlalchemy.pool import StaticPool
from sqlmodel import SQLModel

import expense_api.core.models  # noqa: F401
from expense_api.core.db.session import get_session
from expense_api.core.models.budget import (
    BudgetCategory,
    BudgetDetail,
    BudgetDetailYear,
    BudgetSubcategory,
    Committee,
    Department,
    DepartmentBudgetDetail,
)
from expense_api.core.models.expense import Expense, ExpenseItem
from expense_api.core.models.tenant import Tenant
from expense_api.core.models.user import User, UserYearRole
from expense_api.core.security.jwt import hash_password
from expense_api.core.security.rate_limit import _reset_all
from main import app

YEAR = 2026


@pytest_asyncio.fixture
async def client():
    _reset_all()
    engine = create_async_engine(
        "sqlite+aiosqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )

    @event.listens_for(engine.sync_engine, "connect")
    def _fk_on(dbapi_conn, _rec):
        cur = dbapi_conn.cursor()
        cur.execute("PRAGMA foreign_keys=ON")
        cur.close()

    async with engine.begin() as conn:
        await conn.run_sync(SQLModel.metadata.create_all)
    maker = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)

    async with maker() as s:
        t = Tenant(name="데모", subdomain="demo")
        s.add(t)
        await s.flush()
        s.add(
            User(
                tenantId=t.id,
                userid="admin",
                username="관리자",
                password=hash_password("admin123"),
                role="admin",
            )
        )
        await s.commit()

    async def _override():
        async with maker() as s:
            yield s

    app.dependency_overrides[get_session] = _override
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as c:
        c._maker = maker  # type: ignore[attr-defined]
        yield c
    app.dependency_overrides.clear()
    await engine.dispose()


async def _login(client: AsyncClient) -> dict:
    r = await client.post("/api/auth/login", json={"userid": "admin", "password": "admin123"})
    assert r.status_code == 200
    return {"Authorization": f"Bearer {r.json()['token']}"}


async def _seed_budget_tree(client: AsyncClient) -> dict:
    """위원회→사역팀→항→목→세목(+연도설정/담당자/부서연결) 트리 시드."""
    maker = client._maker  # type: ignore[attr-defined]
    async with maker() as s:
        from sqlalchemy import select

        tid = (await s.execute(select(Tenant.id))).scalar_one()
        admin_id = (await s.execute(select(User.id).where(User.userid == "admin"))).scalar_one()

        manager = User(tenantId=tid, userid="mgr", username="김담당", role="user")
        s.add(manager)
        await s.flush()

        comm = Committee(tenantId=tid, name="기획본부", sortOrder=1)
        s.add(comm)
        await s.flush()
        dept = Department(tenantId=tid, committeeId=comm.id, name="재정팀", sortOrder=1)
        cat = BudgetCategory(tenantId=tid, name="사무행정비", sortOrder=1)
        s.add_all([dept, cat])
        await s.flush()
        sub = BudgetSubcategory(tenantId=tid, categoryId=cat.id, name="회의비", sortOrder=1)
        s.add(sub)
        await s.flush()
        detail = BudgetDetail(
            tenantId=tid, subcategoryId=sub.id, name="간식비", sortOrder=1,
            description="회의 간식, 다과 구입",
        )
        s.add(detail)
        await s.flush()
        s.add(
            BudgetDetailYear(
                tenantId=tid, budgetDetailId=detail.id, year=YEAR,
                managerId=manager.id, budgetAmount=1_000_000,
            )
        )
        s.add(
            DepartmentBudgetDetail(tenantId=tid, departmentId=dept.id, budgetDetailId=detail.id)
        )

        # 지급완료 지출 (hierarchy usedAmount 집계용)
        paid = Expense(
            tenantId=tid, userId=admin_id, committee="기획본부", department="재정팀",
            expenseDate=datetime(YEAR, 3, 2), requestAmount=30_000,
            requestDate=datetime(YEAR, 3, 1), applicantName="관리자",
            bankName="은행", accountNumber="1-2-3", accountHolder="관리자",
            status="APPROVED_FINAL", paymentStatus="COMPLETED",
        )
        s.add(paid)
        await s.flush()
        s.add(
            ExpenseItem(
                tenantId=tid, expenseId=paid.id, budgetCategory="사무행정비",
                budgetSubcategory="회의비", budgetDetail="간식비",
                description="3월 회의 간식", unitPrice=30_000, quantity=1,
                amount=30_000, order=1,
            )
        )
        await s.commit()
        return {
            "tid": tid, "committeeId": comm.id, "departmentId": dept.id,
            "detailId": detail.id, "managerId": manager.id, "expenseId": paid.id,
        }


# ── hierarchy ─────────────────────────────────────────────────────────
async def test_hierarchy_structure_and_used_amount(client: AsyncClient):
    headers = await _login(client)
    seed = await _seed_budget_tree(client)

    r = await client.get(f"/api/budget/hierarchy?year={YEAR}", headers=headers)
    assert r.status_code == 200
    body = r.json()

    assert body["summary"] == {
        "totalCommittees": 1,
        "totalDepartments": 1,
        "totalDetails": 1,
        "totalBudgetAmount": 1_000_000,
        "unassignedCount": 0,
    }
    committee = body["committees"][0]
    assert committee["name"] == "기획본부"
    detail = committee["departments"][0]["details"][0]
    assert detail["fullPath"] == "사무행정비 > 회의비 > 간식비"
    assert detail["managerName"] == "김담당"
    assert detail["budgetAmount"] == 1_000_000
    assert detail["usedAmount"] == 30_000
    assert body["allCommittees"] == [{"id": seed["committeeId"], "name": "기획본부"}]


async def test_hierarchy_search_filters_out_nonmatching(client: AsyncClient):
    headers = await _login(client)
    await _seed_budget_tree(client)

    r = await client.get(f"/api/budget/hierarchy?year={YEAR}&search=없는세목", headers=headers)
    assert r.status_code == 200
    assert r.json()["committees"] == []
    assert r.json()["summary"]["totalDetails"] == 0


# ── hierarchy/export ─────────────────────────────────────────────────
async def test_hierarchy_export_returns_xlsx_with_data_row(client: AsyncClient):
    headers = await _login(client)
    await _seed_budget_tree(client)

    r = await client.get(f"/api/budget/hierarchy/export?year={YEAR}", headers=headers)
    assert r.status_code == 200
    assert r.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert r.headers["content-disposition"] == f'attachment; filename="budget_{YEAR}.xlsx"'

    wb = load_workbook(io.BytesIO(r.content))
    ws = wb.active
    assert ws.title == f"{YEAR}년 예산현황"
    header = [cell.value for cell in ws[1]]
    assert header == ["위원회", "사역팀", "예산(항)", "예산(목)", "예산(세목)", "담당자", "예산금액"]
    data_row = [cell.value for cell in ws[2]]
    assert data_row == ["기획본부", "재정팀", "사무행정비", "회의비", "간식비", "김담당", 1_000_000]
    # 합계 행 (빈 행 다음)
    total_row = [cell.value for cell in ws[4]]
    assert total_row[4] == "합계"


async def test_hierarchy_export_requires_auth(client: AsyncClient):
    r = await client.get(f"/api/budget/hierarchy/export?year={YEAR}")
    assert r.status_code == 401


# ── search ────────────────────────────────────────────────────────────
async def test_search_returns_hierarchy_info(client: AsyncClient):
    headers = await _login(client)
    seed = await _seed_budget_tree(client)

    r = await client.get(f"/api/budget/search?q=간식&year={YEAR}", headers=headers)
    assert r.status_code == 200
    body = r.json()
    assert body["total"] == 1 and body["showing"] == 1
    result = body["results"][0]
    assert result["detail"] == "간식비"
    assert result["fullPath"] == "사무행정비 > 회의비 > 간식비"
    assert result["managerName"] == "김담당"
    assert result["hierarchy"]["committee"] == "기획본부"
    assert result["hierarchy"]["department"] == "재정팀"

    # 부서 필터: 다른 부서 id 로는 결과 없음
    r2 = await client.get(
        f"/api/budget/search?q=간식&year={YEAR}&departmentId={seed['committeeId']}",
        headers=headers,
    )
    assert r2.json()["total"] == 0


async def test_search_empty_query_returns_empty(client: AsyncClient):
    headers = await _login(client)
    r = await client.get("/api/budget/search", headers=headers)
    assert r.status_code == 200
    assert r.json() == {"results": [], "total": 0}


# ── simple ────────────────────────────────────────────────────────────
async def test_simple_requires_finance_head(client: AsyncClient):
    headers = await _login(client)
    await _seed_budget_tree(client)

    # 재정팀장 역할이 없으면 옵션 없음
    r = await client.post("/api/budget/simple", json={"year": YEAR}, headers=headers)
    assert r.status_code == 200
    assert r.json() == {"field": "categories", "options": []}


async def test_simple_cascade_with_finance_head(client: AsyncClient):
    headers = await _login(client)
    seed = await _seed_budget_tree(client)
    maker = client._maker  # type: ignore[attr-defined]
    async with maker() as s:
        s.add(
            UserYearRole(
                tenantId=seed["tid"], userId=seed["managerId"], year=YEAR, role="finance_head"
            )
        )
        await s.commit()

    r = await client.post("/api/budget/simple", json={"year": YEAR}, headers=headers)
    assert r.json() == {"field": "categories", "options": ["사무행정비"]}

    r = await client.post(
        "/api/budget/simple", json={"year": YEAR, "category": "사무행정비"}, headers=headers
    )
    assert r.json() == {"field": "subcategories", "options": ["회의비"]}

    r = await client.post(
        "/api/budget/simple",
        json={"year": YEAR, "category": "사무행정비", "subcategory": "회의비"},
        headers=headers,
    )
    assert r.json() == {"field": "details", "options": ["간식비"]}


# ── simple/all-details ────────────────────────────────────────────────
async def test_all_details_includes_parents_and_manager(client: AsyncClient):
    headers = await _login(client)
    await _seed_budget_tree(client)

    r = await client.get(f"/api/budget/simple/all-details?year={YEAR}", headers=headers)
    assert r.status_code == 200
    details = r.json()["details"]
    assert details == [
        {
            "name": "간식비",
            "category": "사무행정비",
            "subcategory": "회의비",
            "managerId": details[0]["managerId"],
            "managerName": "김담당",
        }
    ]


# ── usage-details ─────────────────────────────────────────────────────
async def test_usage_details_lists_approved_items(client: AsyncClient):
    headers = await _login(client)
    seed = await _seed_budget_tree(client)

    params = (
        f"budgetCategory=사무행정비&budgetSubcategory=회의비&budgetDetail=간식비&year={YEAR}"
    )
    r = await client.get(f"/api/budget/usage-details?{params}", headers=headers)
    assert r.status_code == 200
    body = r.json()
    assert body["count"] == 1 and body["totalAmount"] == 30_000
    item = body["items"][0]
    assert item["expenseId"] == seed["expenseId"]
    assert item["requestDate"] == f"{YEAR}-03-01"
    assert item["status"] == "APPROVED_FINAL"

    # excludeExpenseId 로 이중 차감 방지
    r2 = await client.get(
        f"/api/budget/usage-details?{params}&excludeExpenseId={seed['expenseId']}",
        headers=headers,
    )
    assert r2.json()["count"] == 0


async def test_usage_details_requires_params(client: AsyncClient):
    headers = await _login(client)
    r = await client.get("/api/budget/usage-details?budgetCategory=항만", headers=headers)
    assert r.status_code == 400


# ── memo-examples ─────────────────────────────────────────────────────
async def test_memo_examples_by_id_and_name(client: AsyncClient):
    headers = await _login(client)
    seed = await _seed_budget_tree(client)

    r = await client.get(
        f"/api/budget/memo-examples?budgetDetailId={seed['detailId']}", headers=headers
    )
    assert r.status_code == 200
    assert r.json() == {
        "examples": ["회의 간식", "다과 구입"],
        "budgetDetail": {"id": seed["detailId"], "name": "간식비"},
    }

    r2 = await client.get("/api/budget/memo-examples?budgetDetailName=간식비", headers=headers)
    assert r2.json()["budgetDetail"]["name"] == "간식비"

    r3 = await client.get("/api/budget/memo-examples?budgetDetailName=없는세목", headers=headers)
    assert r3.json() == {"examples": [], "budgetDetail": None}

    r4 = await client.get("/api/budget/memo-examples", headers=headers)
    assert r4.status_code == 400


# ── committees DELETE ─────────────────────────────────────────────────
async def test_delete_committee_blocked_when_departments_exist(client: AsyncClient):
    headers = await _login(client)
    seed = await _seed_budget_tree(client)

    r = await client.delete(f"/api/committees/{seed['committeeId']}", headers=headers)
    assert r.status_code == 400
    assert "하위 사역팀" in r.json()["detail"]


async def test_delete_committee_success_when_empty(client: AsyncClient):
    headers = await _login(client)
    seed = await _seed_budget_tree(client)
    maker = client._maker  # type: ignore[attr-defined]
    async with maker() as s:
        empty = Committee(tenantId=seed["tid"], name="빈위원회", sortOrder=9)
        s.add(empty)
        await s.commit()
        empty_id = empty.id

    r = await client.delete(f"/api/committees/{empty_id}", headers=headers)
    assert r.status_code == 200
    assert r.json() == {"success": True}

    r2 = await client.delete(f"/api/committees/{empty_id}", headers=headers)
    assert r2.status_code == 404
