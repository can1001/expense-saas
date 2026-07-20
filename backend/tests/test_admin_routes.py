"""admin 대시보드/연도 설정 현황/보고서/실행·이력 라우트 계약 테스트. (app/api/admin/dashboard,
app/api/admin/year-setup-status 컷오버 — D1;
app/api/admin/budget-execution, cumulative-report, quarterly-report(+export) 컷오버 — D2;
app/api/admin/hr-admin-execution, manager-exceptions, change-history 컷오버 — D3)
"""

import io
from datetime import datetime
from urllib.parse import quote

import pytest_asyncio
from httpx import ASGITransport, AsyncClient
from openpyxl import load_workbook
from sqlalchemy import event
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine
from sqlalchemy.pool import StaticPool
from sqlmodel import SQLModel

import expense_api.core.models  # noqa: F401
from expense_api.core.db.session import get_session
from expense_api.core.models.budget import (
    BudgetCategory,
    BudgetDetail,
    BudgetDetailYear,
    Committee,
    Department,
    DepartmentBudgetDetail,
)
from expense_api.core.models.budget import BudgetDetailYearHistory, BudgetSubcategory
from expense_api.core.models.expense import Expense, ExpenseItem
from expense_api.core.models.tenant import Tenant
from expense_api.core.models.user import User, UserYearRole, UserYearRoleHistory
from expense_api.core.security.jwt import hash_password
from expense_api.core.security.rate_limit import _reset_all
from main import app

YEAR = 2026


@pytest_asyncio.fixture
async def client():
    _reset_all()
    engine = create_async_engine(
        "sqlite+aiosqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )

    @event.listens_for(engine.sync_engine, "connect")
    def _fk_on(dbapi_conn, _rec):
        cur = dbapi_conn.cursor()
        cur.execute("PRAGMA foreign_keys=ON")
        cur.close()

    async with engine.begin() as conn:
        await conn.run_sync(SQLModel.metadata.create_all)
    maker = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)

    async with maker() as s:
        t = Tenant(name="데모", subdomain="demo")
        s.add(t)
        await s.flush()
        s.add(
            User(
                tenantId=t.id,
                userid="admin",
                username="관리자",
                password=hash_password("admin123"),
                role="admin",
            )
        )
        s.add(
            User(
                tenantId=t.id,
                userid="user1",
                username="사용자1",
                password=hash_password("user123"),
                role="user",
            )
        )
        await s.commit()

    async def _override():
        async with maker() as s:
            yield s

    app.dependency_overrides[get_session] = _override
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as c:
        c._maker = maker  # type: ignore[attr-defined]
        yield c
    app.dependency_overrides.clear()
    await engine.dispose()


async def _login(client: AsyncClient, userid: str = "admin", password: str = "admin123") -> dict:
    r = await client.post("/api/auth/login", json={"userid": userid, "password": password})
    assert r.status_code == 200
    return {"Authorization": f"Bearer {r.json()['token']}"}


async def _seed(client: AsyncClient) -> dict:
    """위원회→사역팀→항→목→세목(연도설정+담당자) + 지출 2건 + 역할 시드."""
    maker = client._maker  # type: ignore[attr-defined]
    async with maker() as s:
        from sqlalchemy import select

        tid = (await s.execute(select(Tenant.id))).scalar_one()
        admin_id = (await s.execute(select(User.id).where(User.userid == "admin"))).scalar_one()
        user_id = (await s.execute(select(User.id).where(User.userid == "user1"))).scalar_one()

        manager = User(tenantId=tid, userid="mgr", username="김담당", role="user")
        s.add(manager)
        await s.flush()

        comm = Committee(tenantId=tid, name="기획본부", sortOrder=1)
        s.add(comm)
        await s.flush()
        dept = Department(tenantId=tid, committeeId=comm.id, name="재정팀", sortOrder=1)
        cat = BudgetCategory(tenantId=tid, name="사무행정비", sortOrder=1)
        s.add_all([dept, cat])
        await s.flush()
        sub = BudgetSubcategory(tenantId=tid, categoryId=cat.id, name="회의비", sortOrder=1)
        s.add(sub)
        await s.flush()
        detail_done = BudgetDetail(tenantId=tid, subcategoryId=sub.id, name="간식비", sortOrder=1)
        detail_missing = BudgetDetail(tenantId=tid, subcategoryId=sub.id, name="교통비", sortOrder=2)
        s.add_all([detail_done, detail_missing])
        await s.flush()
        s.add(
            BudgetDetailYear(
                tenantId=tid, budgetDetailId=detail_done.id, year=YEAR,
                managerId=manager.id, budgetAmount=1_000_000, usedAmount=300_000,
            )
        )
        s.add_all(
            [
                DepartmentBudgetDetail(tenantId=tid, departmentId=dept.id, budgetDetailId=detail_done.id),
                DepartmentBudgetDetail(tenantId=tid, departmentId=dept.id, budgetDetailId=detail_missing.id),
            ]
        )
        s.add(UserYearRole(tenantId=tid, userId=user_id, year=YEAR, role="user"))

        now = datetime.now()
        pending = Expense(
            tenantId=tid, userId=admin_id, committee="기획본부", department="재정팀",
            expenseDate=now, requestAmount=50_000, requestDate=now, applicantName="관리자",
            bankName="은행", accountNumber="1-2-3", accountHolder="관리자",
            status="PENDING", paymentStatus="PENDING",
        )
        approved = Expense(
            tenantId=tid, userId=admin_id, committee="기획본부", department="재정팀",
            expenseDate=now, requestAmount=120_000, requestDate=now, applicantName="관리자",
            bankName="은행", accountNumber="1-2-3", accountHolder="관리자",
            status="APPROVED_FINAL", paymentStatus="COMPLETED",
        )
        s.add_all([pending, approved])
        await s.commit()

    return {"tenantId": tid}


async def test_dashboard_kpi(client: AsyncClient):
    headers = await _login(client)
    await _seed(client)

    r = await client.get(f"/api/admin/dashboard?year={YEAR}", headers=headers)
    assert r.status_code == 200
    body = r.json()
    assert body["year"] == YEAR
    assert body["kpi"]["totalBudget"] == 1_000_000
    assert body["kpi"]["totalUsed"] == 300_000
    assert body["kpi"]["executionRate"] == 30.0
    assert body["kpi"]["pendingApprovals"] == 1
    assert body["kpi"]["pendingPayments"] == 0
    assert body["yearly"]["totalExpense"] == 120_000
    assert body["yearly"]["expenseCount"] == 1
    assert len(body["recentExpenses"]) == 2


async def test_dashboard_requires_permission(client: AsyncClient):
    headers = await _login(client, "user1", "user123")
    r = await client.get("/api/admin/dashboard", headers=headers)
    assert r.status_code == 403


async def test_year_setup_status(client: AsyncClient):
    headers = await _login(client)
    await _seed(client)

    r = await client.get(f"/api/admin/year-setup-status?year={YEAR}", headers=headers)
    assert r.status_code == 200
    body = r.json()
    assert body["year"] == YEAR
    assert body["summary"]["roleSetup"]["total"] == 3
    assert body["summary"]["roleSetup"]["completed"] == 1
    assert body["summary"]["managerAssignment"]["total"] == 2
    assert body["summary"]["managerAssignment"]["completed"] == 1
    assert body["summary"]["budgetInput"]["completed"] == 1
    assert body["summary"]["budgetInput"]["totalAmount"] == 1_000_000
    missing_detail_names = {m["name"] for m in body["missing"]["managers"]}
    assert "교통비" in missing_detail_names
    missing_budget_names = {m["name"] for m in body["missing"]["budgets"]}
    assert "교통비" in missing_budget_names


async def test_year_setup_status_requires_permission(client: AsyncClient):
    headers = await _login(client, "user1", "user123")
    r = await client.get("/api/admin/year-setup-status", headers=headers)
    assert r.status_code == 403


# ── D2: 보고서 ──────────────────────────────────────────────────────


async def _seed_report_tree(client: AsyncClient) -> dict:
    """위원회 2개(사역/행정)×부서×예산세목 + 지출 4건(분기 분산) 시드."""
    maker = client._maker  # type: ignore[attr-defined]
    async with maker() as s:
        from sqlalchemy import select

        tid = (await s.execute(select(Tenant.id))).scalar_one()
        admin_id = (await s.execute(select(User.id).where(User.userid == "admin"))).scalar_one()

        s.add(
            User(
                tenantId=tid, userid="assistant1", username="간사1",
                password=hash_password("asst123"), role="admin_assistant",
            )
        )

        comm1 = Committee(tenantId=tid, name="기획본부", sortOrder=1)
        comm2 = Committee(tenantId=tid, name="행정위원회", sortOrder=2)
        s.add_all([comm1, comm2])
        await s.flush()
        dept1 = Department(tenantId=tid, committeeId=comm1.id, name="재정팀", sortOrder=1)
        dept2 = Department(tenantId=tid, committeeId=comm2.id, name="행정팀", sortOrder=1)
        cat = BudgetCategory(tenantId=tid, name="사무행정비", sortOrder=1)
        s.add_all([dept1, dept2, cat])
        await s.flush()
        sub = BudgetSubcategory(tenantId=tid, categoryId=cat.id, name="회의비", sortOrder=1)
        s.add(sub)
        await s.flush()
        detail1 = BudgetDetail(tenantId=tid, subcategoryId=sub.id, name="간식비", sortOrder=1)
        detail2 = BudgetDetail(tenantId=tid, subcategoryId=sub.id, name="다과비", sortOrder=2)
        s.add_all([detail1, detail2])
        await s.flush()
        s.add_all(
            [
                BudgetDetailYear(tenantId=tid, budgetDetailId=detail1.id, year=YEAR, budgetAmount=1_000_000),
                BudgetDetailYear(tenantId=tid, budgetDetailId=detail2.id, year=YEAR, budgetAmount=400_000),
                DepartmentBudgetDetail(tenantId=tid, departmentId=dept1.id, budgetDetailId=detail1.id),
                DepartmentBudgetDetail(tenantId=tid, departmentId=dept2.id, budgetDetailId=detail2.id),
            ]
        )

        def _expense(committee, department, when, amount, status, payment_status):
            return Expense(
                tenantId=tid, userId=admin_id, committee=committee, department=department,
                expenseDate=when, requestAmount=amount, requestDate=when, applicantName="관리자",
                bankName="은행", accountNumber="1-2-3", accountHolder="관리자",
                status=status, paymentStatus=payment_status,
            )

        e1 = _expense("기획본부", "재정팀", datetime(YEAR, 2, 10), 100_000, "APPROVED_FINAL", "COMPLETED")
        e2 = _expense("기획본부", "재정팀", datetime(YEAR, 8, 5), 200_000, "APPROVED_FINAL", "PENDING")
        e3 = _expense("행정위원회", "행정팀", datetime(YEAR, 5, 1), 50_000, "APPROVED_FINAL", "COMPLETED")
        e4 = _expense("기획본부", "재정팀", datetime(YEAR, 2, 15), 999_999, "PENDING", "PENDING")
        s.add_all([e1, e2, e3, e4])
        await s.flush()
        s.add_all(
            [
                ExpenseItem(
                    tenantId=tid, expenseId=e1.id, budgetCategory="사무행정비", budgetSubcategory="회의비",
                    budgetDetail="간식비", description="2월지출", unitPrice=100_000, quantity=1,
                    amount=100_000, order=1,
                ),
                ExpenseItem(
                    tenantId=tid, expenseId=e2.id, budgetCategory="사무행정비", budgetSubcategory="회의비",
                    budgetDetail="간식비", description="8월지출", unitPrice=200_000, quantity=1,
                    amount=200_000, order=1,
                ),
                ExpenseItem(
                    tenantId=tid, expenseId=e3.id, budgetCategory="사무행정비", budgetSubcategory="회의비",
                    budgetDetail="다과비", description="5월지출", unitPrice=50_000, quantity=1,
                    amount=50_000, order=1,
                ),
            ]
        )
        await s.commit()
    return {"tenantId": tid}


async def test_budget_execution_summary(client: AsyncClient):
    headers = await _login(client)
    await _seed_report_tree(client)

    r = await client.get(f"/api/admin/budget-execution?year={YEAR}", headers=headers)
    assert r.status_code == 200
    body = r.json()
    assert body["year"] == YEAR
    assert body["summary"] == {
        "totalBudget": 1_000_000,
        "totalSpent": 300_000,
        "executionRate": 30,
        "grandTotalBudget": 1_400_000,
        "grandTotalSpent": 350_000,
        "grandTotalExecutionRate": 25,
        "ministryBudgetRatio": 71,
    }
    assert len(body["committees"]) == 1
    committee = body["committees"][0]
    assert committee["name"] == "기획본부"
    assert committee["budget"] == 1_000_000
    assert committee["spent"] == 300_000
    dept = committee["departments"][0]
    assert dept["name"] == "재정팀"
    assert dept["budget"] == 1_000_000
    assert dept["spent"] == 300_000
    assert dept["executionRate"] == 30


async def test_budget_execution_requires_permission(client: AsyncClient):
    headers = await _login(client, "user1", "user123")
    r = await client.get("/api/admin/budget-execution", headers=headers)
    assert r.status_code == 403


async def test_cumulative_report_summary(client: AsyncClient):
    headers = await _login(client)
    await _seed_report_tree(client)

    r = await client.get(f"/api/admin/cumulative-report?year={YEAR}&toQuarter=4", headers=headers)
    assert r.status_code == 200
    body = r.json()
    assert body["year"] == YEAR
    assert body["toQuarter"] == 4
    assert body["summary"] == {
        "totalBudget": 1_400_000,
        "cumulativeSpent": 350_000,
        "remaining": 1_050_000,
        "executionRate": 25.0,
    }
    assert body["quarterlyBreakdown"] == [
        {"quarter": 1, "spent": 100_000, "ratio": 7.1},
        {"quarter": 2, "spent": 50_000, "ratio": 3.6},
        {"quarter": 3, "spent": 200_000, "ratio": 14.3},
        {"quarter": 4, "spent": 0, "ratio": 0},
    ]
    assert body["byDepartment"] == [
        {
            "committee": "기획본부", "department": "재정팀",
            "budget": 1_000_000, "cumulativeSpent": 300_000,
            "remaining": 700_000, "executionRate": 30.0,
        },
        {
            "committee": "행정위원회", "department": "행정팀",
            "budget": 400_000, "cumulativeSpent": 50_000,
            "remaining": 350_000, "executionRate": 12.5,
        },
    ]


async def test_cumulative_report_requires_permission(client: AsyncClient):
    headers = await _login(client, "user1", "user123")
    r = await client.get("/api/admin/cumulative-report", headers=headers)
    assert r.status_code == 403


async def test_quarterly_report_summary(client: AsyncClient):
    headers = await _login(client)
    await _seed_report_tree(client)

    r = await client.get(f"/api/admin/quarterly-report?year={YEAR}&quarter=3", headers=headers)
    assert r.status_code == 200
    body = r.json()
    assert body["year"] == YEAR
    assert body["quarter"] == 3
    assert body["period"] == {"startDate": f"{YEAR}-07-01", "endDate": f"{YEAR}-09-30"}
    assert body["summary"] == {
        "totalExpenses": 1,
        "totalAmount": 200_000,
        "completedAmount": 0,
        "pendingAmount": 200_000,
    }
    assert body["budgetSummary"] == {
        "totalBudget": 1_400_000,
        "yearlySpent": 350_000,
        "yearlyRemaining": 1_050_000,
        "yearlyExecutionRate": 25.0,
        "quarterlyBudget": 350_000,
        "quarterlySpent": 200_000,
        "quarterlyRemaining": 150_000,
        "quarterlyExecutionRate": 57.1,
    }
    assert body["byMonth"] == [
        {"month": 7, "monthLabel": "7월", "count": 0, "amount": 0, "ratio": 0},
        {"month": 8, "monthLabel": "8월", "count": 1, "amount": 200_000, "ratio": 100.0},
        {"month": 9, "monthLabel": "9월", "count": 0, "amount": 0, "ratio": 0},
    ]
    assert body["filterOptions"] == {
        "departments": [{"committee": "기획본부", "department": "재정팀"}],
        "categories": ["사무행정비"],
    }

    dept_names = [d["committee"] for d in body["byDepartment"]]
    assert dept_names == ["기획본부", "행정위원회"]
    gihoek = body["byDepartment"][0]
    assert gihoek["amount"] == 200_000
    assert gihoek["departments"][0]["department"] == "재정팀"
    assert gihoek["departments"][0]["amount"] == 200_000
    detail = gihoek["departments"][0]["categoryDetails"][0]["subcategories"][0]["details"][0]
    assert detail["detail"] == "간식비"
    assert detail["amount"] == 200_000

    assert len(body["byCategory"]) == 1
    cat = body["byCategory"][0]
    assert cat["category"] == "사무행정비"
    assert cat["spentAmount"] == 200_000
    assert cat["budgetAmount"] == 1_400_000
    assert cat["yearlySpentAmount"] == 350_000
    assert cat["quarterlyBudget"] == 350_000
    assert cat["quarterlyExecutionRate"] == 57.1
    sub = cat["subcategories"][0]
    assert sub["subcategory"] == "회의비"
    detail_names = [d["detail"] for d in sub["details"]]
    assert detail_names == ["간식비", "다과비"]
    snack = next(d for d in sub["details"] if d["detail"] == "간식비")
    assert snack["spentAmount"] == 200_000
    assert snack["yearlySpentAmount"] == 300_000
    treat = next(d for d in sub["details"] if d["detail"] == "다과비")
    assert treat["spentAmount"] == 0
    assert treat["yearlySpentAmount"] == 50_000


async def test_quarterly_report_requires_permission(client: AsyncClient):
    headers = await _login(client, "user1", "user123")
    r = await client.get("/api/admin/quarterly-report", headers=headers)
    assert r.status_code == 403


async def test_quarterly_report_export_xlsx(client: AsyncClient):
    headers = await _login(client)
    await _seed_report_tree(client)

    r = await client.get(f"/api/admin/quarterly-report/export?year={YEAR}&quarter=3", headers=headers)
    assert r.status_code == 200
    assert r.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"분기별회계보고_{YEAR}년_3분기.xlsx"
    assert r.headers["content-disposition"] == f"attachment; filename*=UTF-8''{quote(filename)}"

    wb = load_workbook(io.BytesIO(r.content))
    assert wb.sheetnames == ["요약", "월별 지출", "부서별 지출", "분기별예산대비지출", "세목별상세"]
    summary = wb["요약"]
    assert [c.value for c in summary[1]] == ["항목", "값"]
    dept_sheet = wb["부서별 지출"]
    assert [c.value for c in dept_sheet[1]] == ["위원회", "사역팀(부)", "건수", "금액", "비율(%)"]
    assert [c.value for c in dept_sheet[2]] == ["기획본부", "재정팀", 1, 200_000, 100.0]


async def test_quarterly_report_export_requires_report_export_permission(client: AsyncClient):
    await _seed_report_tree(client)
    headers = await _login(client, "assistant1", "asst123")

    r_view = await client.get(f"/api/admin/quarterly-report?year={YEAR}&quarter=3", headers=headers)
    assert r_view.status_code == 200

    r = await client.get(f"/api/admin/quarterly-report/export?year={YEAR}&quarter=3", headers=headers)
    assert r.status_code == 403
    assert r.json()["detail"] == "권한이 없습니다."


# ── D3: 실행·이력 ──────────────────────────────────────────────────────


async def _seed_hr_admin_tree(client: AsyncClient) -> dict:
    """인사위/행정위 위원회 각 1 세목 + 지출 각 1건 시드."""
    maker = client._maker  # type: ignore[attr-defined]
    async with maker() as s:
        from sqlalchemy import select

        tid = (await s.execute(select(Tenant.id))).scalar_one()
        admin_id = (await s.execute(select(User.id).where(User.userid == "admin"))).scalar_one()

        comm_p = Committee(tenantId=tid, name="인사위원회", sortOrder=1)
        comm_a = Committee(tenantId=tid, name="행정위원회", sortOrder=2)
        s.add_all([comm_p, comm_a])
        await s.flush()
        dept_p = Department(tenantId=tid, committeeId=comm_p.id, name="인사팀", sortOrder=1)
        dept_a = Department(tenantId=tid, committeeId=comm_a.id, name="총무팀", sortOrder=1)
        cat = BudgetCategory(tenantId=tid, name="운영비", sortOrder=1)
        s.add_all([dept_p, dept_a, cat])
        await s.flush()
        sub_p = BudgetSubcategory(tenantId=tid, categoryId=cat.id, name="인건비", sortOrder=1)
        sub_a = BudgetSubcategory(tenantId=tid, categoryId=cat.id, name="사무비", sortOrder=2)
        s.add_all([sub_p, sub_a])
        await s.flush()
        detail_p = BudgetDetail(tenantId=tid, subcategoryId=sub_p.id, name="급여", sortOrder=1)
        detail_a = BudgetDetail(tenantId=tid, subcategoryId=sub_a.id, name="소모품비", sortOrder=1)
        s.add_all([detail_p, detail_a])
        await s.flush()
        s.add_all(
            [
                BudgetDetailYear(tenantId=tid, budgetDetailId=detail_p.id, year=YEAR, budgetAmount=1_000_000),
                BudgetDetailYear(tenantId=tid, budgetDetailId=detail_a.id, year=YEAR, budgetAmount=500_000),
                DepartmentBudgetDetail(tenantId=tid, departmentId=dept_p.id, budgetDetailId=detail_p.id),
                DepartmentBudgetDetail(tenantId=tid, departmentId=dept_a.id, budgetDetailId=detail_a.id),
            ]
        )

        when = datetime(YEAR, 3, 1)
        e_p = Expense(
            tenantId=tid, userId=admin_id, committee="인사위원회", department="인사팀",
            expenseDate=when, requestAmount=300_000, requestDate=when, applicantName="관리자",
            bankName="은행", accountNumber="1-2-3", accountHolder="관리자",
            status="APPROVED_FINAL", paymentStatus="COMPLETED",
        )
        e_a = Expense(
            tenantId=tid, userId=admin_id, committee="행정위원회", department="총무팀",
            expenseDate=when, requestAmount=100_000, requestDate=when, applicantName="관리자",
            bankName="은행", accountNumber="1-2-3", accountHolder="관리자",
            status="APPROVED_FINAL", paymentStatus="COMPLETED",
        )
        s.add_all([e_p, e_a])
        await s.flush()
        s.add_all(
            [
                ExpenseItem(
                    tenantId=tid, expenseId=e_p.id, budgetCategory="운영비", budgetSubcategory="인건비",
                    budgetDetail="급여", description="급여", unitPrice=300_000, quantity=1,
                    amount=300_000, order=1,
                ),
                ExpenseItem(
                    tenantId=tid, expenseId=e_a.id, budgetCategory="운영비", budgetSubcategory="사무비",
                    budgetDetail="소모품비", description="소모품", unitPrice=100_000, quantity=1,
                    amount=100_000, order=1,
                ),
            ]
        )
        await s.commit()
    return {"tenantId": tid}


async def test_hr_admin_execution_summary(client: AsyncClient):
    headers = await _login(client)
    await _seed_hr_admin_tree(client)

    r = await client.get(f"/api/admin/hr-admin-execution?year={YEAR}", headers=headers)
    assert r.status_code == 200
    body = r.json()
    assert body["year"] == YEAR
    assert body["personnel"]["total"] == {"budget": 1_000_000, "spent": 300_000, "executionRate": 30}
    assert body["personnel"]["items"] == [
        {"name": "인건비", "budget": 1_000_000, "spent": 300_000, "executionRate": 30}
    ]
    assert body["admin"]["total"] == {"budget": 500_000, "spent": 100_000, "executionRate": 20}
    assert body["admin"]["groups"] == [
        {
            "name": "운영비",
            "items": [{"name": "사무비", "budget": 500_000, "spent": 100_000, "executionRate": 20}],
            "subtotal": {"budget": 500_000, "spent": 100_000, "executionRate": 20},
        }
    ]
    assert body["combined"] == {"budget": 1_500_000, "spent": 400_000, "executionRate": 27}
    assert body["overall"] == {"totalBudget": 1_500_000, "hrAdminBudget": 1_500_000, "hrAdminRatio": 100}


async def test_hr_admin_execution_requires_permission(client: AsyncClient):
    headers = await _login(client, "user1", "user123")
    r = await client.get("/api/admin/hr-admin-execution", headers=headers)
    assert r.status_code == 403


async def _seed_manager_exceptions_tree(client: AsyncClient) -> dict:
    """세목 2개(담당자=팀장 일치/불일치) 시드."""
    maker = client._maker  # type: ignore[attr-defined]
    async with maker() as s:
        from sqlalchemy import select

        tid = (await s.execute(select(Tenant.id))).scalar_one()

        leader = User(tenantId=tid, userid="leader1", username="김팀장", role="team_leader")
        manager2 = User(tenantId=tid, userid="mgr2", username="박담당", role="user")
        s.add_all([leader, manager2])
        await s.flush()

        comm = Committee(tenantId=tid, name="선교위원회", sortOrder=1)
        s.add(comm)
        await s.flush()
        dept = Department(tenantId=tid, committeeId=comm.id, name="선교팀", sortOrder=1)
        cat = BudgetCategory(tenantId=tid, name="사업비", sortOrder=1)
        s.add_all([dept, cat])
        await s.flush()
        sub = BudgetSubcategory(tenantId=tid, categoryId=cat.id, name="행사비", sortOrder=1)
        s.add(sub)
        await s.flush()
        detail_ok = BudgetDetail(tenantId=tid, subcategoryId=sub.id, name="다과비", sortOrder=1)
        detail_exc = BudgetDetail(tenantId=tid, subcategoryId=sub.id, name="교통비", sortOrder=2)
        s.add_all([detail_ok, detail_exc])
        await s.flush()

        s.add(
            UserYearRole(tenantId=tid, userId=leader.id, year=YEAR, role="team_leader", departmentId=dept.id)
        )
        s.add_all(
            [
                DepartmentBudgetDetail(tenantId=tid, departmentId=dept.id, budgetDetailId=detail_ok.id),
                DepartmentBudgetDetail(tenantId=tid, departmentId=dept.id, budgetDetailId=detail_exc.id),
            ]
        )
        bdy_exc = BudgetDetailYear(
            tenantId=tid, budgetDetailId=detail_exc.id, year=YEAR, managerId=manager2.id, budgetAmount=50_000
        )
        s.add_all(
            [
                BudgetDetailYear(
                    tenantId=tid, budgetDetailId=detail_ok.id, year=YEAR, managerId=leader.id, budgetAmount=100_000
                ),
                bdy_exc,
            ]
        )
        await s.commit()

    return {
        "tenantId": tid,
        "leaderId": leader.id,
        "leaderName": leader.username,
        "manager2Id": manager2.id,
        "manager2Name": manager2.username,
        "detailExcId": detail_exc.id,
        "bdyExcId": bdy_exc.id,
    }


async def test_manager_exceptions_summary(client: AsyncClient):
    headers = await _login(client)
    seeded = await _seed_manager_exceptions_tree(client)

    r = await client.get(f"/api/admin/manager-exceptions?year={YEAR}", headers=headers)
    assert r.status_code == 200
    body = r.json()
    assert body["year"] == YEAR
    assert body["summary"] == {"totalDetails": 2, "exceptionCount": 1, "exceptionRate": 50.0}
    assert body["exceptions"] == [
        {
            "budgetDetailId": seeded["detailExcId"],
            "budgetDetailYearId": seeded["bdyExcId"],
            "committee": "선교위원회",
            "department": "선교팀",
            "category": "사업비",
            "subcategory": "행사비",
            "detail": "교통비",
            "teamLeader": {"id": seeded["leaderId"], "name": seeded["leaderName"]},
            "manager": {"id": seeded["manager2Id"], "name": seeded["manager2Name"]},
        }
    ]


async def test_manager_exceptions_requires_permission(client: AsyncClient):
    headers = await _login(client, "user1", "user123")
    r = await client.get("/api/admin/manager-exceptions", headers=headers)
    assert r.status_code == 403


async def _seed_change_history(client: AsyncClient) -> dict:
    """역할 변경 이력 3건(연도 다름 1건 포함) + 예산 변경 이력 1건 시드."""
    maker = client._maker  # type: ignore[attr-defined]
    async with maker() as s:
        from sqlalchemy import select

        tid = (await s.execute(select(Tenant.id))).scalar_one()
        user1_id = (await s.execute(select(User.id).where(User.userid == "user1"))).scalar_one()

        s.add_all(
            [
                UserYearRoleHistory(
                    tenantId=tid, userId=user1_id, year=YEAR, action="CREATE",
                    changedBy="관리자", newRole="team_leader",
                    changedAt=datetime(YEAR, 1, 1, 10, 0, 0),
                ),
                UserYearRoleHistory(
                    tenantId=tid, userId=user1_id, year=YEAR, action="UPDATE",
                    changedBy="관리자", previousRole="team_leader", newRole="accountant",
                    changedAt=datetime(YEAR, 2, 1, 10, 0, 0),
                ),
                UserYearRoleHistory(
                    tenantId=tid, userId="ghost-user-id", year=YEAR - 1, action="DELETE",
                    changedBy="관리자", previousRole="user",
                    changedAt=datetime(YEAR - 1, 12, 1, 9, 0, 0),
                ),
            ]
        )
        s.add(
            BudgetDetailYearHistory(
                tenantId=tid, budgetDetailId="bd-1", budgetDetailName="간식비", year=YEAR,
                action="UPDATE", changedBy="관리자", previousBudgetAmt=100_000, newBudgetAmt=200_000,
                changedAt=datetime(YEAR, 3, 1, 9, 0, 0),
            )
        )
        await s.commit()
    return {"tenantId": tid}


async def test_change_history_roles_filtered_by_year(client: AsyncClient):
    headers = await _login(client)
    await _seed_change_history(client)

    r = await client.get(f"/api/admin/change-history?type=roles&year={YEAR}", headers=headers)
    assert r.status_code == 200
    body = r.json()
    assert body["total"] == {"roles": 2}
    assert "budgetHistory" not in body
    history = body["roleHistory"]
    assert len(history) == 2
    assert history[0]["action"] == "UPDATE"
    assert history[0]["userName"] == "사용자1"
    assert history[0]["userLoginId"] == "user1"
    assert history[1]["action"] == "CREATE"


async def test_change_history_all_default(client: AsyncClient):
    headers = await _login(client)
    await _seed_change_history(client)

    r = await client.get("/api/admin/change-history", headers=headers)
    assert r.status_code == 200
    body = r.json()
    assert body["total"] == {"roles": 3, "budgets": 1}
    assert len(body["roleHistory"]) == 3
    ghost_entry = next(h for h in body["roleHistory"] if h["userId"] == "ghost-user-id")
    assert ghost_entry["userName"] == "ghost-user-id"
    assert "userLoginId" not in ghost_entry
    assert body["budgetHistory"][0]["budgetDetailName"] == "간식비"


async def test_change_history_requires_permission(client: AsyncClient):
    headers = await _login(client, "user1", "user123")
    r = await client.get("/api/admin/change-history", headers=headers)
    assert r.status_code == 403
