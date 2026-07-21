"""budget/upload 계약 테스트 — Excel 템플릿 다운로드 + replace/merge/append 업로드.
(app/api/budget/upload 컷오버, lib/budget-upload.ts 712L 포팅)
"""

import io

import pytest_asyncio
from httpx import ASGITransport, AsyncClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import event, select
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine
from sqlalchemy.pool import StaticPool
from sqlmodel import SQLModel

import expense_api.core.models  # noqa: F401
from expense_api.core.db.session import get_session
from expense_api.core.models.budget import (
    BudgetCategory,
    BudgetDetail,
    BudgetSubcategory,
    Committee,
    Department,
)
from expense_api.core.models.tenant import Tenant
from expense_api.core.models.user import User
from expense_api.core.security.jwt import hash_password
from expense_api.core.security.rate_limit import _reset_all
from main import app

XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@pytest_asyncio.fixture
async def client():
    _reset_all()
    engine = create_async_engine(
        "sqlite+aiosqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )

    @event.listens_for(engine.sync_engine, "connect")
    def _fk_on(dbapi_conn, _rec):
        cur = dbapi_conn.cursor()
        cur.execute("PRAGMA foreign_keys=ON")
        cur.close()

    async with engine.begin() as conn:
        await conn.run_sync(SQLModel.metadata.create_all)
    maker = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)

    async with maker() as s:
        t = Tenant(name="데모", subdomain="demo")
        s.add(t)
        await s.flush()
        s.add(
            User(
                tenantId=t.id,
                userid="admin",
                username="관리자",
                password=hash_password("admin123"),
                role="admin",
            )
        )
        await s.commit()

    async def _override():
        async with maker() as s:
            yield s

    app.dependency_overrides[get_session] = _override
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as c:
        c._maker = maker  # type: ignore[attr-defined]
        yield c
    app.dependency_overrides.clear()
    await engine.dispose()


async def _login(client: AsyncClient) -> dict:
    r = await client.post("/api/auth/login", json={"userid": "admin", "password": "admin123"})
    assert r.status_code == 200
    return {"Authorization": f"Bearer {r.json()['token']}"}


def _make_xlsx(rows: list[list]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "위원회", "사역팀(부)", "예산(항)", "예산(목)", "예산(세목)",
            "담당자", "계정코드", "항목 내역", "활성화", "연도", "예산금액",
        ]
    )
    for row in rows:
        sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


_ROW_A = ["위원회A", "부서A", "항A", "목A", "세목A", "매니저1", "AC1", "내역1", "true", 2026, 100_000]
_ROW_B = ["위원회A", "부서A", "항A", "목A", "세목B", "매니저2", "AC2", "내역2", "true", 2026, 200_000]


async def _counts(client: AsyncClient) -> dict[str, int]:
    maker = client._maker  # type: ignore[attr-defined]
    async with maker() as s:
        committees = (await s.execute(select(Committee))).scalars().all()
        departments = (await s.execute(select(Department))).scalars().all()
        categories = (await s.execute(select(BudgetCategory))).scalars().all()
        subcategories = (await s.execute(select(BudgetSubcategory))).scalars().all()
        details = (await s.execute(select(BudgetDetail))).scalars().all()
        return {
            "committees": len(committees),
            "departments": len(departments),
            "categories": len(categories),
            "subcategories": len(subcategories),
            "details": len(details),
        }


# ── GET /upload — 템플릿 다운로드 ───────────────────────────────────────
async def test_template_download_returns_xlsx_with_headers(client: AsyncClient):
    headers = await _login(client)
    r = await client.get("/api/budget/upload", headers=headers)
    assert r.status_code == 200
    assert r.headers["content-type"] == XLSX_CONTENT_TYPE
    assert r.headers["content-disposition"].startswith('attachment; filename="budget_template_')

    wb = load_workbook(io.BytesIO(r.content))
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    assert header == [
        "위원회", "사역팀(부)", "예산(항)", "예산(목)", "예산(세목)",
        "담당자", "계정코드", "항목 내역", "활성화", "연도", "예산금액",
    ]


async def test_template_download_requires_auth(client: AsyncClient):
    r = await client.get("/api/budget/upload")
    assert r.status_code == 401


# ── POST /upload — 검증 오류 ─────────────────────────────────────────────
async def test_upload_requires_file(client: AsyncClient):
    headers = await _login(client)
    r = await client.post("/api/budget/upload", headers=headers, data={"mode": "merge"})
    assert r.status_code == 400
    body = r.json()
    assert body["success"] is False
    assert body["error"]["type"] == "VALIDATION"
    assert body["error"]["fields"][0]["fieldName"] == "file"


async def test_upload_rejects_invalid_file_type(client: AsyncClient):
    headers = await _login(client)
    files = {"file": ("data.txt", b"not excel", "text/plain")}
    r = await client.post("/api/budget/upload", headers=headers, files=files)
    assert r.status_code == 400
    assert "지원하지 않는 파일 형식" in r.json()["message"]


async def test_upload_rejects_invalid_mode(client: AsyncClient):
    headers = await _login(client)
    xlsx = _make_xlsx([_ROW_A])
    files = {"file": ("budget.xlsx", xlsx, XLSX_CONTENT_TYPE)}
    r = await client.post(
        "/api/budget/upload", headers=headers, files=files, data={"mode": "invalid"}
    )
    assert r.status_code == 400
    assert "잘못된 업로드 모드" in r.json()["message"]


async def test_upload_reports_row_validation_errors(client: AsyncClient):
    headers = await _login(client)
    xlsx = _make_xlsx([["위원회A", "", "항A", "목A", "세목A"]])  # department 누락
    files = {"file": ("budget.xlsx", xlsx, XLSX_CONTENT_TYPE)}
    r = await client.post("/api/budget/upload", headers=headers, files=files)
    assert r.status_code == 400
    body = r.json()
    assert "파싱 오류" in body["message"]
    assert body["error"]["fields"][0]["fieldName"] == "row_2_department"


# ── POST /upload — dryRun ────────────────────────────────────────────────
async def test_dry_run_reports_summary_without_db_changes(client: AsyncClient):
    headers = await _login(client)
    xlsx = _make_xlsx([_ROW_A, _ROW_B])
    files = {"file": ("budget.xlsx", xlsx, XLSX_CONTENT_TYPE)}
    r = await client.post(
        "/api/budget/upload",
        headers=headers,
        files=files,
        data={"mode": "merge", "dryRun": "true"},
    )
    assert r.status_code == 200
    body = r.json()
    assert body["success"] is True
    assert body["code"] == "UPLOAD_SUCCESS"
    assert body["data"]["dryRun"] is True
    assert body["data"]["summary"] == {
        "totalRows": 2, "created": 2, "updated": 0, "skipped": 0, "errors": 0,
    }

    counts = await _counts(client)
    assert counts == {
        "committees": 0, "departments": 0, "categories": 0, "subcategories": 0, "details": 0,
    }


# ── POST /upload — merge ────────────────────────────────────────────────
async def test_merge_creates_then_updates_existing_detail(client: AsyncClient):
    headers = await _login(client)
    xlsx = _make_xlsx([_ROW_A])
    files = {"file": ("budget.xlsx", xlsx, XLSX_CONTENT_TYPE)}

    r1 = await client.post(
        "/api/budget/upload", headers=headers, files=files, data={"mode": "merge"}
    )
    assert r1.status_code == 200
    assert r1.json()["data"]["summary"]["created"] == 1

    counts = await _counts(client)
    assert counts["committees"] == 1 and counts["details"] == 1

    updated_row = ["위원회A", "부서A", "항A", "목A", "세목A", "매니저1", "AC9", "변경됨", "true", 2026, 999_999]
    xlsx2 = _make_xlsx([updated_row])
    files2 = {"file": ("budget.xlsx", xlsx2, XLSX_CONTENT_TYPE)}
    r2 = await client.post(
        "/api/budget/upload", headers=headers, files=files2, data={"mode": "merge"}
    )
    assert r2.status_code == 200
    assert r2.json()["data"]["summary"] == {
        "totalRows": 1, "created": 0, "updated": 1, "skipped": 0, "errors": 0,
    }

    maker = client._maker  # type: ignore[attr-defined]
    async with maker() as s:
        detail = (await s.execute(select(BudgetDetail))).scalars().one()
        assert detail.accountCode == "AC9"
        assert detail.description == "변경됨"


# ── POST /upload — append ───────────────────────────────────────────────
async def test_append_skips_existing_detail(client: AsyncClient):
    headers = await _login(client)
    xlsx = _make_xlsx([_ROW_A])
    files = {"file": ("budget.xlsx", xlsx, XLSX_CONTENT_TYPE)}
    r1 = await client.post(
        "/api/budget/upload", headers=headers, files=files, data={"mode": "merge"}
    )
    assert r1.json()["data"]["summary"]["created"] == 1

    xlsx2 = _make_xlsx([_ROW_A, _ROW_B])
    files2 = {"file": ("budget.xlsx", xlsx2, XLSX_CONTENT_TYPE)}
    r2 = await client.post(
        "/api/budget/upload", headers=headers, files=files2, data={"mode": "append"}
    )
    assert r2.status_code == 200
    assert r2.json()["data"]["summary"] == {
        "totalRows": 2, "created": 1, "updated": 0, "skipped": 1, "errors": 0,
    }

    counts = await _counts(client)
    assert counts["details"] == 2


# ── POST /upload — replace ──────────────────────────────────────────────
async def test_replace_clears_previous_categories_and_details(client: AsyncClient):
    headers = await _login(client)
    xlsx = _make_xlsx([_ROW_A, _ROW_B])
    files = {"file": ("budget.xlsx", xlsx, XLSX_CONTENT_TYPE)}
    r1 = await client.post(
        "/api/budget/upload", headers=headers, files=files, data={"mode": "merge"}
    )
    assert r1.json()["data"]["summary"]["created"] == 2

    replacement_row = ["위원회A", "부서A", "항B", "목B", "세목C", "매니저3", "AC3", "내역3", "true", 2026, 50_000]
    xlsx2 = _make_xlsx([replacement_row])
    files2 = {"file": ("budget.xlsx", xlsx2, XLSX_CONTENT_TYPE)}
    r2 = await client.post(
        "/api/budget/upload", headers=headers, files=files2, data={"mode": "replace"}
    )
    assert r2.status_code == 200
    assert r2.json()["data"]["summary"] == {
        "totalRows": 1, "created": 1, "updated": 0, "skipped": 0, "errors": 0,
    }

    maker = client._maker  # type: ignore[attr-defined]
    async with maker() as s:
        categories = (await s.execute(select(BudgetCategory))).scalars().all()
        details = (await s.execute(select(BudgetDetail))).scalars().all()
        # Committee/Department 는 유지
        committees = (await s.execute(select(Committee))).scalars().all()
        assert len(committees) == 1
        assert [c.name for c in categories] == ["항B"]
        assert [d.name for d in details] == ["세목C"]
